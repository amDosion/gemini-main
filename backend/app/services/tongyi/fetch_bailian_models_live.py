#!/usr/bin/env python3
"""
阿里云百炼模型列表实时爬取脚本
真正从HTML结构动态提取模型信息，不使用预定义模型列表
输出格式：类别、模型名称、定位/说明
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import json
from typing import List, Dict, Tuple, Optional


class BailianModelScraper:
    """百炼模型列表实时爬取器 - 动态解析版"""
    
    def __init__(self):
        self.base_url = "https://help.aliyun.com/zh/model-studio/models"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
        }
        
    def fetch_page(self) -> str:
        """获取页面HTML内容"""
        print(f"正在获取页面: {self.base_url}")
        try:
            response = requests.get(self.base_url, headers=self.headers, timeout=60)
            response.raise_for_status()
            response.encoding = 'utf-8'
            print(f"页面获取成功，内容长度: {len(response.text)} 字符")
            return response.text
        except requests.RequestException as e:
            print(f"获取页面失败: {e}")
            return ""
    
    def is_model_name(self, text: str) -> bool:
        """
        判断文本是否像模型名称
        通过特征判断，而非预定义列表
        """
        if not text or len(text) < 3 or len(text) > 60:
            return False
        
        text = text.strip().lower()
        
        # 排除明显不是模型名的
        exclude_keywords = [
            '稳定版', '最新版', '快照版', '预览版', '免费', '价格', '成本',
            '说明', '版本', '上下文', '输入', '输出', 'token', '长度',
            '使用方法', 'api参考', '在线体验', '思考模式', '计费',
            '元/张', '元/秒', '每千', '有效期', '百炼', '阿里云',
            '中国内地', '全球', '国际', '美国', '新加坡', '北京',
        ]
        for kw in exclude_keywords:
            if kw in text:
                return False
        
        # 模型名通常包含这些特征
        # 1. 包含版本号模式 (v1, v2, 2.5, 3.5等)
        # 2. 包含模型类型后缀 (-instruct, -chat, -plus, -turbo, -flash等)
        # 3. 包含参数规模 (7b, 14b, 32b, 72b等)
        # 4. 以常见前缀开头
        
        model_indicators = [
            # 版本号
            r'-v\d', r'v\d\.', r'\d+\.\d+', 
            # 后缀
            r'-instruct', r'-chat', r'-plus', r'-turbo', r'-flash',
            r'-max', r'-lite', r'-preview', r'-latest', r'-thinking',
            # 参数规模
            r'\d+b', r'\d+B',
            # 任务类型
            r't2i', r'i2i', r't2v', r'i2v', r'vl', r'ocr', r'tts', r'asr',
        ]
        
        for pattern in model_indicators:
            if re.search(pattern, text, re.IGNORECASE):
                return True
        
        # 检查是否以已知前缀开头（这里只检查格式，不硬编码具体模型）
        # 模型名通常是 "前缀-后缀" 或 "前缀数字-后缀" 格式
        if re.match(r'^[a-z][\w.-]+-[\w.-]+$', text):
            return True
        
        return False
    
    def extract_from_tables(self, soup: BeautifulSoup) -> List[Dict]:
        """从表格中提取模型信息"""
        models = []
        tables = soup.find_all('table')
        print(f"找到 {len(tables)} 个表格")
        
        current_section = "未分类"
        
        for table in tables:
            # 尝试找到表格前的标题作为分类
            prev = table.find_previous(['h2', 'h3', 'h4'])
            if prev:
                section_text = prev.get_text(strip=True)
                if len(section_text) < 30:
                    current_section = section_text
            
            rows = table.find_all('tr')
            header_row = None
            model_col_idx = -1
            desc_col_idx = -1
            
            for row_idx, row in enumerate(rows):
                cells = row.find_all(['td', 'th'])
                cell_texts = [c.get_text(separator=' ', strip=True) for c in cells]
                
                # 检测表头行，找到"模型名称"列
                if row_idx == 0 or header_row is None:
                    for i, text in enumerate(cell_texts):
                        text_lower = text.lower()
                        if '模型名称' in text or 'model' in text_lower:
                            model_col_idx = i
                            header_row = row_idx
                        if '说明' in text or '描述' in text or '版本' in text:
                            desc_col_idx = i
                    continue
                
                # 如果找到了模型名称列，从该列提取
                if model_col_idx >= 0 and model_col_idx < len(cell_texts):
                    cell_text = cell_texts[model_col_idx]
                    # 可能一个单元格有多个模型名（换行分隔）
                    potential_names = re.split(r'[\n\r]+', cell_text)
                    for name in potential_names:
                        name = name.strip()
                        # 进一步清理，去掉括号内容和中文说明
                        name = re.sub(r'\s*[\(（].*?[\)）]', '', name)
                        name = re.sub(r'\s+.*$', '', name)  # 取第一个空格前的部分
                        name = name.strip()
                        
                        if self.is_model_name(name):
                            desc = ""
                            if desc_col_idx >= 0 and desc_col_idx < len(cell_texts):
                                desc = cell_texts[desc_col_idx][:200]
                            models.append({
                                "类别": current_section,
                                "模型名称": name,
                                "定位/说明": desc
                            })
                else:
                    # 没有明确的模型名称列，遍历所有单元格查找
                    for i, text in enumerate(cell_texts):
                        # 清理文本
                        clean_text = re.sub(r'\s*[\(（].*?[\)）]', '', text)
                        words = clean_text.split()
                        for word in words:
                            word = word.strip()
                            if self.is_model_name(word):
                                desc = ""
                                if i + 1 < len(cell_texts):
                                    desc = cell_texts[i + 1][:200]
                                models.append({
                                    "类别": current_section,
                                    "模型名称": word,
                                    "定位/说明": desc
                                })
                                break
        
        return models
    
    def extract_from_text(self, soup: BeautifulSoup) -> List[Dict]:
        """从页面文本中提取模型名称（作为补充）"""
        models = []
        
        # 获取所有文本内容
        text = soup.get_text()
        
        # 用通用正则匹配可能的模型名
        # 模型名通常格式：小写字母开头，包含字母数字和点横线
        pattern = r'\b([a-z][a-z0-9]*(?:[\.-][a-z0-9]+)+)\b'
        
        matches = re.findall(pattern, text, re.IGNORECASE)
        
        for match in matches:
            if self.is_model_name(match):
                models.append({
                    "类别": "从文本提取",
                    "模型名称": match,
                    "定位/说明": ""
                })
        
        return models
    
    def categorize_by_name(self, model_name: str) -> str:
        """根据模型名称特征自动分类"""
        name_lower = model_name.lower()
        
        # 图像相关
        if any(kw in name_lower for kw in ['image', 't2i', 'i2i', 'sketch', 'paint', 
                                            'background', 'segment', 'erase', 'poster',
                                            'wordart', 'flux', 'diffusion', 'virtual-model',
                                            'facechain', 'aitryon', 'shoe-model']):
            return "图像生成模型"
        
        # 视频相关
        if any(kw in name_lower for kw in ['t2v', 'i2v', 'flf2v', 'video', 'animate',
                                            'digital-human', 'emote', 'portrait', 'emoji',
                                            'retalk', 'ref2v', 'vace']):
            return "视频生成模型"
        
        # 语音相关
        if any(kw in name_lower for kw in ['tts', 'asr', 'audio', 'voice', 'speech',
                                            'sambert', 'paraformer', 'sensevoice', 'gummy']):
            return "语音模型"
        
        # 向量相关
        if any(kw in name_lower for kw in ['embedding', 'rerank']):
            return "向量及其他模型"
        
        # 其他特殊模型
        if any(kw in name_lower for kw in ['farui', 'intent', 'roleplay']):
            return "向量及其他模型"
        
        # 默认文本生成
        return "文本生成模型"
    
    def get_subcategory(self, model_name: str) -> str:
        """根据模型名称获取子类别"""
        name_lower = model_name.lower()
        
        # 根据前缀和特征推断子类别
        if name_lower.startswith('qwen3-max') or name_lower.startswith('qwen-max'):
            return "通义千问Max"
        if name_lower.startswith('qwen3-plus') or name_lower.startswith('qwen-plus'):
            return "通义千问Plus"
        if name_lower.startswith('qwen3-flash') or name_lower.startswith('qwen-flash'):
            return "通义千问Flash"
        if name_lower.startswith('qwen-turbo'):
            return "通义千问Turbo"
        if name_lower.startswith('qwen-long'):
            return "通义千问Long"
        if 'omni' in name_lower and 'realtime' in name_lower:
            return "通义千问Omni-Realtime"
        if 'omni' in name_lower:
            return "通义千问Omni"
        if name_lower.startswith('qvq'):
            return "QVQ视觉推理"
        if name_lower.startswith('qwq'):
            return "QwQ推理模型"
        if '-vl' in name_lower or name_lower.startswith('qwen-vl') or name_lower.startswith('qwen3-vl'):
            return "通义千问VL"
        if 'ocr' in name_lower:
            return "通义千问OCR"
        if 'audio' in name_lower and 'qwen' in name_lower:
            return "通义千问Audio"
        if 'math' in name_lower:
            return "数学模型"
        if 'coder' in name_lower:
            return "代码模型"
        if name_lower.startswith('qwen-mt') or 'mt-' in name_lower:
            return "翻译模型"
        if 'doc-turbo' in name_lower:
            return "数据挖掘"
        if 'deep-research' in name_lower:
            return "深入研究"
        if name_lower.startswith('deepseek'):
            return "DeepSeek"
        if name_lower.startswith('kimi') or 'moonshot' in name_lower:
            return "Kimi"
        if name_lower.startswith('glm'):
            return "GLM"
        if name_lower.startswith('abab'):
            return "MiniMax"
        if name_lower.startswith('wan') or name_lower.startswith('wanx'):
            if 't2i' in name_lower:
                return "通义万相文生图"
            if 'i2i' in name_lower or 'imageedit' in name_lower:
                return "通义万相图像编辑"
            if 't2v' in name_lower:
                return "文生视频"
            if 'i2v' in name_lower:
                return "图生视频"
            return "通义万相"
        if 'qwen-image' in name_lower:
            if 'edit' in name_lower:
                return "通义千问图像编辑"
            return "通义千问文生图"
        if name_lower.startswith('stable-diffusion'):
            return "Stable Diffusion"
        if name_lower.startswith('flux'):
            return "FLUX"
        if 'cosyvoice' in name_lower:
            return "CosyVoice语音合成"
        if 'sambert' in name_lower:
            return "Sambert语音合成"
        if 'paraformer' in name_lower:
            return "Paraformer语音识别"
        if 'sensevoice' in name_lower:
            return "SenseVoice语音识别"
        if 'embedding' in name_lower:
            return "文本向量"
        if 'rerank' in name_lower:
            return "重排序"
        
        # 通用Qwen开源模型
        if name_lower.startswith('qwen3-'):
            return "Qwen3开源"
        if name_lower.startswith('qwen2.5-'):
            return "Qwen2.5开源"
        if name_lower.startswith('qwen2-'):
            return "Qwen2开源"
        if name_lower.startswith('qwen1.5-'):
            return "Qwen1.5开源"
        
        return "其他"
    
    def scrape(self) -> Dict[str, List[Dict]]:
        """执行爬取"""
        html_content = self.fetch_page()
        if not html_content:
            return {}
        
        soup = BeautifulSoup(html_content, 'html.parser')
        
        # 从表格提取
        table_models = self.extract_from_tables(soup)
        print(f"从表格提取: {len(table_models)} 个")
        
        # 从文本提取（补充）
        text_models = self.extract_from_text(soup)
        print(f"从文本提取: {len(text_models)} 个")
        
        # 合并去重
        seen = set()
        all_models = []
        for model in table_models + text_models:
            name = model["模型名称"].lower()
            if name not in seen:
                seen.add(name)
                # 重新分类
                main_cat = self.categorize_by_name(model["模型名称"])
                sub_cat = self.get_subcategory(model["模型名称"])
                all_models.append({
                    "类别": sub_cat,
                    "模型名称": model["模型名称"],
                    "定位/说明": model.get("定位/说明", ""),
                    "_main_category": main_cat
                })
        
        print(f"去重后: {len(all_models)} 个")
        
        # 按主分类分组
        categorized = {
            "文本生成模型": [],
            "图像生成模型": [],
            "视频生成模型": [],
            "语音模型": [],
            "向量及其他模型": []
        }
        
        for model in all_models:
            main_cat = model.pop("_main_category")
            categorized[main_cat].append(model)
        
        # 排序
        for cat in categorized:
            categorized[cat].sort(key=lambda x: (x["类别"], x["模型名称"]))
        
        return categorized
    
    def save_to_excel(self, models: Dict[str, List[Dict]], output_path: str):
        """保存到Excel"""
        wb = Workbook()
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        first_sheet = True
        for sheet_name, model_list in models.items():
            if not model_list:
                continue
            
            if first_sheet:
                ws = wb.active
                ws.title = sheet_name[:31]
                first_sheet = False
            else:
                ws = wb.create_sheet(sheet_name[:31])
            
            headers = ["类别", "模型名称", "定位/说明"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            for row_idx, model in enumerate(model_list, 2):
                ws.cell(row=row_idx, column=1, value=model["类别"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=model["模型名称"]).border = thin_border
                ws.cell(row=row_idx, column=3, value=model["定位/说明"]).border = thin_border
            
            ws.column_dimensions['A'].width = 22
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 65
        
        wb.save(output_path)
        print(f"Excel已保存: {output_path}")
    
    def save_to_json(self, models: Dict[str, List[Dict]], output_path: str):
        """保存到JSON"""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(models, f, ensure_ascii=False, indent=2)
        print(f"JSON已保存: {output_path}")
    
    def save_to_csv(self, models: Dict[str, List[Dict]], output_path: str):
        """保存到CSV"""
        all_rows = []
        for sheet_name, model_list in models.items():
            for model in model_list:
                all_rows.append({
                    "分类": sheet_name,
                    "类别": model["类别"],
                    "模型名称": model["模型名称"],
                    "定位/说明": model["定位/说明"]
                })
        
        df = pd.DataFrame(all_rows)
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"CSV已保存: {output_path}")


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='阿里云百炼模型列表实时爬取工具')
    parser.add_argument('--output', '-o', default='bailian_models', help='输出文件名（不含扩展名）')
    parser.add_argument('--format', '-f', choices=['excel', 'json', 'csv', 'all'], default='excel', help='输出格式')
    
    args = parser.parse_args()
    
    scraper = BailianModelScraper()
    models = scraper.scrape()
    
    if not models or all(len(v) == 0 for v in models.values()):
        print("未能获取模型列表")
        return
    
    if args.format == 'excel' or args.format == 'all':
        scraper.save_to_excel(models, f"{args.output}.xlsx")
    
    if args.format == 'json' or args.format == 'all':
        scraper.save_to_json(models, f"{args.output}.json")
    
    if args.format == 'csv' or args.format == 'all':
        scraper.save_to_csv(models, f"{args.output}.csv")
    
    print("\n=== 统计 ===")
    total = 0
    for cat, lst in models.items():
        print(f"{cat}: {len(lst)} 个")
        total += len(lst)
    print(f"总计: {total} 个")


if __name__ == "__main__":
    main()
