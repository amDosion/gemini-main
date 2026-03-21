"""
Workflow API Router - 工作流执行 + Agent CRUD (Phase 1)

提供：
- POST /api/workflows/execute    执行工作流
- GET  /api/workflows/mode-presets            列出模式工作流预设
- GET  /api/workflows/mode-presets/{mode_id}  获取单个模式工作流预设
- GET  /api/agents               列出 Agent
- POST /api/agents               创建 Agent
- PUT  /api/agents/{id}          更新 Agent
- DELETE /api/agents/{id}        删除 Agent
"""

import asyncio
import inspect
import json
import time
import logging
import copy
import re
import io
import mimetypes
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Tuple
from pathlib import Path
from urllib.parse import urlparse

from fastapi import APIRouter, HTTPException, Depends, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict
from sqlalchemy.orm import Session
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError

from ...core.config import settings
from ...core.database import get_db, SessionLocal
from ...core.dependencies import require_current_user
from ...models.db_models import (
    AgentRegistry,
    WorkflowExecution,
    NodeExecution,
    WorkflowTemplate,
    ConfigProfile,
    ActiveStorage,
    generate_uuid,
)
from ...utils.case_converter import to_camel_case
from ...services.agent.agent_llm_service import AgentLLMService
from ...services.agent.workflow_engine import WorkflowEngine
from ...services.agent.workflow_history_image_service import (
    build_workflow_image_previews as _build_workflow_image_previews_shared,
    build_workflow_images_zip as _build_workflow_images_zip_shared,
    resolve_trusted_workflow_history_base_url as _resolve_trusted_workflow_history_base_url_shared,
)
from ...services.agent.workflow_history_media_service import (
    build_workflow_media_previews as _build_workflow_media_previews_shared,
    build_workflow_media_zip as _build_workflow_media_zip_shared,
    download_workflow_media_binary as _download_workflow_media_binary_shared,
    resolve_workflow_media_item as _resolve_workflow_media_item_shared,
)
from ...services.agent.workflow_result_contract import (
    extract_cost_summary as _extract_cost_summary_shared,
    extract_runtime_hints as _extract_runtime_hints_shared,
    extract_text_preview as _extract_text_preview_shared,
    extract_trace_summary as _extract_trace_summary_shared,
    normalize_runtime_hint as _normalize_runtime_hint_shared,
    pick_primary_runtime as _pick_primary_runtime_shared,
)
from ...services.agent.workflow_runtime_store import create_workflow_runtime_store
from ...services.agent.workflow_runtime_helpers import (
    RuntimeEventMetrics,
    build_checkpoint_summary as _build_checkpoint_summary_helper,
    build_resume_request_payload as _build_resume_request_payload_helper,
    build_runtime_metrics_snapshot as _build_runtime_metrics_snapshot_helper,
    create_pause_checkpoint as _create_pause_checkpoint_helper,
    record_runtime_event_publish,
)
from ...services.gemini.base.video_common import is_google_provider_video_uri
from ...services.agent.workflow_payload_normalizer import (
    _clamp_optional_int,
    _normalize_agent_default_task_type,
    _normalize_agent_task_type,
    _normalize_analysis_type,
    _normalize_optional_choice,
    _normalize_string_list,
    _normalize_workflow_input_payload,
    _normalize_workflow_nodes,
    _validate_and_normalize_agent_card,
    _validate_workflow_execute_payload,
)
from ...services.agent.agent_seed_service import ensure_seed_agents, get_default_seed_agents
from ...services.gemini.agent.workflow_template_service import WorkflowTemplateService
from ...services.common.attachment_service import AttachmentService
from ...services.common.reference_image_catalog import (
    is_placeholder_reference_image_url,
    pick_reference_image,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["workflows"])


# ==================== Request Models ====================

class WorkflowExecuteRequest(BaseModel):
    nodes: List[Dict[str, Any]]
    edges: List[Dict[str, Any]]
    input: Optional[Dict[str, Any]] = None
    meta: Optional[Dict[str, Any]] = None
    async_mode: Optional[bool] = False


class CreateAgentRequest(BaseModel):
    model_config = ConfigDict(protected_namespaces=("model_validate", "model_dump"))

    name: str
    description: Optional[str] = ""
    agent_type: Optional[str] = "custom"
    provider_id: str
    model_id: str
    system_prompt: Optional[str] = ""
    temperature: Optional[float] = 0.7
    max_tokens: Optional[int] = 4096
    icon: Optional[str] = "🤖"
    color: Optional[str] = "#14b8a6"
    agent_card: Optional[Dict[str, Any]] = None


class UpdateAgentRequest(BaseModel):
    model_config = ConfigDict(protected_namespaces=("model_validate", "model_dump"))

    name: Optional[str] = None
    description: Optional[str] = None
    agent_type: Optional[str] = None
    provider_id: Optional[str] = None
    model_id: Optional[str] = None
    system_prompt: Optional[str] = None
    temperature: Optional[float] = None
    max_tokens: Optional[int] = None
    icon: Optional[str] = None
    color: Optional[str] = None
    status: Optional[str] = None
    agent_card: Optional[Dict[str, Any]] = None


class WorkflowTemplateCreateRequest(BaseModel):
    name: str
    description: Optional[str] = None
    category: Optional[str] = "general"
    workflow_type: Optional[str] = "graph"
    tags: Optional[List[str]] = None
    config: Dict[str, Any]
    is_public: Optional[bool] = False


class WorkflowTemplateUpdateRequest(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    workflow_type: Optional[str] = None
    tags: Optional[List[str]] = None
    config: Optional[Dict[str, Any]] = None
    is_public: Optional[bool] = None


class WorkflowTemplateCopyRequest(BaseModel):
    name: Optional[str] = None


class WorkflowTemplateCategoryCreateRequest(BaseModel):
    name: str


class WorkflowTemplateRebuildRequest(BaseModel):
    recreate_starters: Optional[bool] = True


class WorkflowResetRequest(BaseModel):
    recreate_starters: Optional[bool] = True


ALLOWED_AGENT_DEFAULT_TASK_TYPES = {
    "chat",
    "image-gen",
    "image-edit",
    "video-gen",
    "audio-gen",
    "vision-understand",
    "data-analysis",
}
ALLOWED_AGENT_LIST_TASK_FILTERS = {"all", *ALLOWED_AGENT_DEFAULT_TASK_TYPES}

WORKFLOW_STATUS_ALIASES = {
    "queued": "pending",
    "in_progress": "running",
    "done": "completed",
    "success": "completed",
    "error": "failed",
    "canceled": "cancelled",
    "paused": "paused",
    "pause_requested": "running",
    "resumed": "running",
}
WORKFLOW_ALLOWED_STATUSES = {"pending", "running", "paused", "completed", "failed", "cancelled"}
WORKFLOW_TERMINAL_STATUSES = {"completed", "failed", "cancelled"}
WORKFLOW_ALLOWED_TRANSITIONS: Dict[str, set[str]] = {
    "pending": {"running", "failed", "cancelled", "paused"},
    "running": {"paused", "completed", "failed", "cancelled"},
    "paused": {"running", "failed", "cancelled"},
    "completed": set(),
    "failed": set(),
    "cancelled": set(),
}

NODE_STATUS_ALIASES = {
    "queued": "pending",
    "in_progress": "running",
    "done": "completed",
    "success": "completed",
    "error": "failed",
    "cancelled": "skipped",
    "canceled": "skipped",
}
NODE_ALLOWED_STATUSES = {"pending", "running", "completed", "failed", "skipped"}
NODE_ALLOWED_TRANSITIONS: Dict[str, set[str]] = {
    "pending": {"running", "failed", "skipped"},
    "running": {"completed", "failed", "skipped"},
    "completed": set(),
    "failed": set(),
    "skipped": set(),
}

WORKFLOW_IDEMPOTENCY_META_FIELDS = ("idempotencyKey", "idempotency_key")
WORKFLOW_IDEMPOTENCY_HEADER_FIELDS = ("idempotency-key", "x-idempotency-key")
WORKFLOW_IDEMPOTENCY_MAX_KEY_LENGTH = 128
WORKFLOW_IDEMPOTENCY_SCAN_LIMIT = 500

ALLOWED_WORKFLOW_AGENT_TASK_TYPES = {
    "chat",
    "image-gen",
    "image-edit",
    "video-gen",
    "audio-gen",
    "vision-understand",
    "data-analysis",
}
ALLOWED_WORKFLOW_ANALYSIS_TYPES = {
    "comprehensive",
    "statistics",
    "correlation",
    "trends",
    "distribution",
}
ALLOWED_IMAGE_OUTPUT_MIME_TYPES = {"image/png", "image/jpeg", "image/webp"}
ALLOWED_OUTPUT_FORMATS = {"text", "json", "markdown"}
VALID_AGENT_TYPES = {"custom", "adk", "google-adk", "seed", "interactions"}
AGENT_TYPE_ALIASES = {
    "google-adk": "adk",
}
ALLOWED_IMAGE_EDIT_MODES = {
    "image-chat-edit",
    "image-mask-edit",
    "image-inpainting",
    "image-background-edit",
    "image-recontext",
    "image-outpainting",
}


def _normalize_agent_registry_type(value: Any) -> str:
    normalized = str(value or "custom").strip().lower()
    return AGENT_TYPE_ALIASES.get(normalized, normalized)


def _agent_type_requires_google_provider(agent_type: str) -> bool:
    return str(agent_type or "").strip().lower() == "adk"

WORKFLOW_EXECUTION_CLIENT_POLICY = {
    "sse_idle_threshold_ms": 15000,
    "polling_interval_ms": 8000,
    "hard_timeout_ms": 30 * 60 * 1000,
}
EXECUTION_STATE_PUSH_MIN_INTERVAL_MS = 400
WORKFLOW_HISTORY_PREVIEW_MAX_BYTES_PER_IMAGE = 5 * 1024 * 1024
WORKFLOW_HISTORY_PREVIEW_MAX_TOTAL_BYTES = 20 * 1024 * 1024
WORKFLOW_HISTORY_AUDIO_PREVIEW_MAX_BYTES = 64 * 1024 * 1024
WORKFLOW_HISTORY_VIDEO_PREVIEW_MAX_BYTES = 128 * 1024 * 1024

WORKFLOW_STATUS_STREAM_TERMINAL_EVENTS = {
    "workflow_complete",
    "workflow_failed",
    "workflow_cancelled",
    "workflow_paused",
}

AGENT_MODEL_SELECTION_POLICY = {
    "strategy": "provider_default_then_first_compatible",
    "tasks": ["chat", "data-analysis", "vision-understand", "image-gen", "image-edit", "video-gen", "audio-gen"],
}


class InvalidWorkflowStateTransitionError(ValueError):
    """Raised when workflow or node status transition violates state-machine rules."""


def _normalize_idempotency_key(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) > WORKFLOW_IDEMPOTENCY_MAX_KEY_LENGTH:
        return text[:WORKFLOW_IDEMPOTENCY_MAX_KEY_LENGTH]
    return text


def _normalize_workflow_status(value: Any, default: str = "pending") -> str:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return default
    mapped = WORKFLOW_STATUS_ALIASES.get(normalized, normalized)
    if mapped in WORKFLOW_ALLOWED_STATUSES:
        return mapped
    return default


def _normalize_node_status(value: Any, default: str = "pending") -> str:
    normalized = str(value or "").strip().lower()
    if not normalized:
        return default
    mapped = NODE_STATUS_ALIASES.get(normalized, normalized)
    if mapped in NODE_ALLOWED_STATUSES:
        return mapped
    return default


def _resolve_terminal_workflow_event(status: str) -> str:
    normalized = _normalize_workflow_status(status, default="failed")
    if normalized == "completed":
        return "workflow_complete"
    if normalized == "cancelled":
        return "workflow_cancelled"
    return "workflow_failed"


def _normalize_agent_list_status(value: Any) -> str:
    normalized = str(value or "").strip().lower()
    return normalized if normalized in {"active", "inactive"} else ""


def _normalize_agent_task_filter(value: Any) -> str:
    normalized = str(value or "").strip()
    if not normalized:
        return "all"
    normalized_task = _normalize_agent_default_task_type(normalized)
    if normalized_task in ALLOWED_AGENT_LIST_TASK_FILTERS:
        return normalized_task
    return "all"


def _extract_agent_default_task_type(agent_payload: Dict[str, Any]) -> str:
    if not isinstance(agent_payload, dict):
        return "chat"
    agent_card = agent_payload.get("agentCard")
    if not isinstance(agent_card, dict):
        agent_card = agent_payload.get("agent_card")
    defaults = agent_card.get("defaults") if isinstance(agent_card, dict) else None
    raw_task_type = defaults.get("defaultTaskType") if isinstance(defaults, dict) else None
    normalized = _normalize_agent_default_task_type(raw_task_type)
    return normalized if normalized in ALLOWED_AGENT_DEFAULT_TASK_TYPES else "chat"


def _build_agent_task_counts(agent_payloads: List[Dict[str, Any]]) -> Dict[str, int]:
    counts = {task_type: 0 for task_type in ALLOWED_AGENT_LIST_TASK_FILTERS}
    counts["all"] = len(agent_payloads)
    for agent_payload in agent_payloads:
        counts[_extract_agent_default_task_type(agent_payload)] += 1
    return counts


def _filter_agent_payloads_by_task(
    agent_payloads: List[Dict[str, Any]],
    task_type: Any,
) -> List[Dict[str, Any]]:
    normalized_task = _normalize_agent_task_filter(task_type)
    if normalized_task == "all":
        return list(agent_payloads)
    return [
        agent_payload for agent_payload in agent_payloads
        if _extract_agent_default_task_type(agent_payload) == normalized_task
    ]


def _resolve_workflow_final_status(status: str) -> str:
    normalized = _normalize_workflow_status(status)
    if normalized == "paused":
        return "workflow_paused"
    return normalized


def _assert_allowed_transition(
    *,
    current_status: str,
    target_status: str,
    allowed_transitions: Dict[str, set[str]],
    subject: str,
) -> str:
    if target_status == current_status:
        return current_status

    allowed_next = allowed_transitions.get(current_status, set())
    if target_status in allowed_next:
        return target_status

    raise InvalidWorkflowStateTransitionError(
        f"Illegal {subject} status transition: {current_status} -> {target_status}"
    )


def _transition_workflow_status(execution: WorkflowExecution, target_status: str) -> str:
    current_status = _normalize_workflow_status(execution.status)
    normalized_target = _normalize_workflow_status(target_status, default="")
    if not normalized_target:
        raise InvalidWorkflowStateTransitionError(
            f"Illegal workflow status transition: {current_status} -> {target_status}"
        )

    _assert_allowed_transition(
        current_status=current_status,
        target_status=normalized_target,
        allowed_transitions=WORKFLOW_ALLOWED_TRANSITIONS,
        subject="workflow",
    )
    execution.status = normalized_target
    return normalized_target


def _transition_node_status(node_execution: NodeExecution, target_status: str) -> str:
    current_status = _normalize_node_status(node_execution.status)
    normalized_target = _normalize_node_status(target_status, default="")
    if not normalized_target:
        raise InvalidWorkflowStateTransitionError(
            f"Illegal node status transition: {current_status} -> {target_status}"
        )

    _assert_allowed_transition(
        current_status=current_status,
        target_status=normalized_target,
        allowed_transitions=NODE_ALLOWED_TRANSITIONS,
        subject="node",
    )
    node_execution.status = normalized_target
    return normalized_target


def _build_mode_workflow_preset(
    *,
    mode_id: str,
    name: str,
    description: str,
    tool_name: str,
    tool_label: str,
    tool_description: str,
    requires_image: bool = False,
    prompt_hint: str = "",
    prompt_example: Any = None,
    tool_overrides: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    node_key = mode_id.replace("_", "-").replace("/", "-")
    start_id = f"start-{node_key}"
    input_text_id = f"input-text-{node_key}"
    input_image_id = f"input-image-{node_key}"
    tool_id = f"tool-{node_key}"
    end_id = f"end-{node_key}"
    default_image_url = pick_reference_image(seed=mode_id)

    tool_data: Dict[str, Any] = {
        "type": "tool",
        "label": tool_label,
        "description": tool_description,
        "icon": "🧩" if requires_image else "🖼️",
        "iconColor": "bg-indigo-500" if requires_image else "bg-pink-500",
        "toolName": tool_name,
    }
    if requires_image:
        tool_data["toolArgsTemplate"] = json.dumps(
            {
                "imageUrl": "{{input.imageUrl}}",
                "editPrompt": "{{input.task}}",
                "mode": mode_id,
            },
            ensure_ascii=False,
        )
        tool_data["toolReferenceImageUrl"] = default_image_url
    else:
        tool_data["toolArgsTemplate"] = json.dumps(
            {"prompt": "{{input.task}}"},
            ensure_ascii=False,
        )
        tool_data["toolNumberOfImages"] = 1
        tool_data["toolAspectRatio"] = "1:1"

    if tool_overrides:
        tool_data.update(tool_overrides)

    nodes: List[Dict[str, Any]] = [
        {
            "id": start_id,
            "type": "start",
            "position": {"x": 80, "y": 220},
            "data": {
                "type": "start",
                "label": "开始",
                "description": "接收模式输入",
                "icon": "🚀",
                "iconColor": "bg-blue-500",
            },
        },
        {
            "id": input_text_id,
            "type": "input_text",
            "position": {"x": 320, "y": 220},
            "data": {
                "type": "input_text",
                "label": "文本输入",
                "description": "配置任务提示词（input.task）",
                "icon": "📝",
                "iconColor": "bg-emerald-500",
                "startTask": "",
            },
        },
    ]
    edges: List[Dict[str, Any]] = [
        {"id": f"edge-{node_key}-1", "source": start_id, "target": input_text_id},
    ]

    previous_id = input_text_id
    tool_x = 820
    end_x = 1080
    if requires_image:
        nodes.append(
            {
                "id": input_image_id,
                "type": "input_image",
                "position": {"x": 560, "y": 220},
                "data": {
                    "type": "input_image",
                    "label": "图片输入",
                    "description": "上传原图或填写 imageUrl",
                    "icon": "🖼️",
                    "iconColor": "bg-lime-500",
                    "startImageUrl": default_image_url,
                },
            }
        )
        edges.append({"id": f"edge-{node_key}-2", "source": input_text_id, "target": input_image_id})
        previous_id = input_image_id
    else:
        tool_x = 580
        end_x = 840

    nodes.extend([
        {
            "id": tool_id,
            "type": "tool",
            "position": {"x": tool_x, "y": 220},
            "data": tool_data,
        },
        {
            "id": end_id,
            "type": "end",
            "position": {"x": end_x, "y": 220},
            "data": {
                "type": "end",
                "label": "结束",
                "description": "输出执行结果",
                "icon": "🏁",
                "iconColor": "bg-green-500",
            },
        },
    ])
    edges.extend([
        {"id": f"edge-{node_key}-3", "source": previous_id, "target": tool_id},
        {"id": f"edge-{node_key}-4", "source": tool_id, "target": end_id},
    ])

    normalized_prompt_example = prompt_example
    if requires_image and isinstance(prompt_example, dict):
        normalized_prompt_example = dict(prompt_example)
        image_url = normalized_prompt_example.get("imageUrl")
        if image_url is None:
            image_url = normalized_prompt_example.get("image_url")
        if is_placeholder_reference_image_url(image_url):
            normalized_prompt_example["imageUrl"] = default_image_url
            normalized_prompt_example.pop("image_url", None)

    return {
        "id": mode_id,
        "mode": mode_id,
        "name": name,
        "description": description,
        "requiresImage": requires_image,
        "promptHint": prompt_hint,
        "promptExample": normalized_prompt_example,
        "workflow": {
            "schemaVersion": 2,
            "nodes": nodes,
            "edges": edges,
        },
    }


MODE_WORKFLOW_PRESET_ITEMS: List[Dict[str, Any]] = [
    _build_mode_workflow_preset(
        mode_id="image-gen",
        name="模式流程 · 图片生成",
        description="与 image-gen 模式一致：文生图工具节点自动执行。",
        tool_name="image_generate",
        tool_label="图片生成",
        tool_description="根据输入提示词生成图片",
        requires_image=False,
        prompt_hint="在“文本输入”节点填写生成需求；也支持 JSON 输入。",
        prompt_example="生成一张产品主图：白底、柔光、写实风格，突出产品细节",
    ),
    _build_mode_workflow_preset(
        mode_id="image-chat-edit",
        name="模式流程 · 图片编辑（对话式）",
        description="与 image-chat-edit 模式一致：输入原图 + 编辑指令后自动执行。",
        tool_name="image_chat_edit",
        tool_label="图片编辑（对话式）",
        tool_description="基于参考图执行对话式编辑",
        requires_image=True,
        prompt_hint="在“文本输入 + 图片输入”节点填写 task 和 imageUrl；也支持 JSON 输入。",
        prompt_example={
            "task": "将背景替换为纯白摄影棚，保留主体细节",
            "imageUrl": "https://example.com/original-image.png",
        },
        tool_overrides={"toolEditMode": "image-chat-edit"},
    ),
    _build_mode_workflow_preset(
        mode_id="image-mask-edit",
        name="模式流程 · 蒙版编辑",
        description="与 image-mask-edit 模式一致：按蒙版区域进行局部修改。",
        tool_name="image_mask_edit",
        tool_label="蒙版编辑",
        tool_description="在局部区域执行图像编辑",
        requires_image=True,
        prompt_hint="在“文本输入 + 图片输入”节点填写 task 和 imageUrl；mask 可在高级参数补充。",
        prompt_example={
            "task": "将选中区域替换为木纹材质，保持光照一致",
            "imageUrl": "https://example.com/original-image.png",
        },
        tool_overrides={"toolEditMode": "image-mask-edit"},
    ),
    _build_mode_workflow_preset(
        mode_id="image-inpainting",
        name="模式流程 · 局部重绘",
        description="与 image-inpainting 模式一致：按指令重绘局部内容。",
        tool_name="image_inpainting",
        tool_label="局部重绘",
        tool_description="基于参考图进行局部内容重绘",
        requires_image=True,
        prompt_hint="在“文本输入 + 图片输入”节点填写 task 和 imageUrl；也支持 JSON 输入。",
        prompt_example={
            "task": "移除画面中多余物体，并补全背景",
            "imageUrl": "https://example.com/original-image.png",
        },
        tool_overrides={"toolEditMode": "image-inpainting"},
    ),
    _build_mode_workflow_preset(
        mode_id="image-background-edit",
        name="模式流程 · 背景替换",
        description="与 image-background-edit 模式一致：替换背景并保持主体。",
        tool_name="image_background_edit",
        tool_label="背景替换",
        tool_description="替换背景并保持主体结构",
        requires_image=True,
        prompt_hint="在“文本输入 + 图片输入”节点填写 task 和 imageUrl；也支持 JSON 输入。",
        prompt_example={
            "task": "把背景改为现代客厅场景，主体位置不变",
            "imageUrl": "https://example.com/original-image.png",
        },
        tool_overrides={"toolEditMode": "image-background-edit"},
    ),
    _build_mode_workflow_preset(
        mode_id="image-recontext",
        name="模式流程 · 场景重构",
        description="与 image-recontext 模式一致：在新语境中重构图像场景。",
        tool_name="image_recontext",
        tool_label="场景重构",
        tool_description="在不同语境下重构画面元素",
        requires_image=True,
        prompt_hint="在“文本输入 + 图片输入”节点填写 task 和 imageUrl；也支持 JSON 输入。",
        prompt_example={
            "task": "将产品放到户外露营场景，增加自然光氛围",
            "imageUrl": "https://example.com/original-image.png",
        },
        tool_overrides={"toolEditMode": "image-recontext"},
    ),
    _build_mode_workflow_preset(
        mode_id="image-outpainting",
        name="模式流程 · 图片扩展",
        description="与 image-outpainting 模式一致：基于原图向外扩展画面。",
        tool_name="image_outpaint",
        tool_label="图片扩展",
        tool_description="扩展原图边界并补全内容",
        requires_image=True,
        prompt_hint="在“文本输入 + 图片输入”节点填写 task 和 imageUrl；也支持 JSON 输入。",
        prompt_example={
            "task": "向左右各扩展画面 30%，保持透视和光照一致",
            "imageUrl": "https://example.com/original-image.png",
        },
        tool_overrides={"toolEditMode": "image-outpainting", "toolAspectRatio": "16:9"},
    ),
    {
        "id": "video-gen",
        "mode": "video-gen",
        "name": "模式流程 · 视频生成",
        "description": "与 video-gen 模式一致：可直接走当前活跃 Provider 的文生视频 / 图生视频。",
        "requiresImage": False,
        "promptHint": "在“文本输入”节点填写视频脚本；也可以给“视频生成”节点补充首帧参考图、源视频和掩码参数。",
        "promptExample": "生成一支 8 秒产品短片：白色耳机盒在浅灰桌面上缓慢旋转，柔和侧光，镜头平稳推进，适合电商首屏展示。",
        "workflow": {
            "schemaVersion": 2,
            "nodes": [
                {
                    "id": "start-video-gen-mode",
                    "type": "start",
                    "position": {"x": 80, "y": 220},
                    "data": {
                        "type": "start",
                        "label": "开始",
                        "description": "接收视频生成任务",
                        "icon": "🚀",
                        "iconColor": "bg-blue-500",
                    },
                },
                {
                    "id": "input-text-video-gen-mode",
                    "type": "input_text",
                    "position": {"x": 320, "y": 220},
                    "data": {
                        "type": "input_text",
                        "label": "文本输入",
                        "description": "配置视频提示词（input.task）",
                        "icon": "📝",
                        "iconColor": "bg-emerald-500",
                        "startTask": "",
                    },
                },
                {
                    "id": "agent-video-gen-mode",
                    "type": "agent",
                    "position": {"x": 580, "y": 220},
                    "data": {
                        "type": "agent",
                        "label": "视频生成",
                        "description": "自动使用当前活跃配置里的视频生成模型",
                        "icon": "🎬",
                        "iconColor": "bg-fuchsia-500",
                        "inlineUseActiveProfile": True,
                        "agentTaskType": "video-gen",
                        "agentAspectRatio": "16:9",
                        "agentResolutionTier": "1K",
                        "agentVideoDurationSeconds": 8,
                        "inputMapping": (
                            "视频需求：{{input-text-video-gen-mode.output.text}}\n\n"
                            "请补全镜头语言、运动节奏、主体稳定性和光线细节，然后直接生成视频。"
                        ),
                        "instructions": "输出适合产品展示或叙事短片的视频，保证主体稳定、镜头流畅、避免无关文字和水印。",
                    },
                },
                {
                    "id": "end-video-gen-mode",
                    "type": "end",
                    "position": {"x": 840, "y": 220},
                    "data": {
                        "type": "end",
                        "label": "结束",
                        "description": "输出视频结果",
                        "icon": "🏁",
                        "iconColor": "bg-green-500",
                    },
                },
            ],
            "edges": [
                {"id": "edge-video-gen-mode-1", "source": "start-video-gen-mode", "target": "input-text-video-gen-mode"},
                {"id": "edge-video-gen-mode-2", "source": "input-text-video-gen-mode", "target": "agent-video-gen-mode"},
                {"id": "edge-video-gen-mode-3", "source": "agent-video-gen-mode", "target": "end-video-gen-mode"},
            ],
        },
    },
]

MODE_WORKFLOW_PRESETS: Dict[str, Dict[str, Any]] = {
    item["id"]: item for item in MODE_WORKFLOW_PRESET_ITEMS
}

MODE_WORKFLOW_PRESET_ALIASES: Dict[str, str] = {
    "image_gen": "image-gen",
    "image-edit": "image-chat-edit",
    "image_edit": "image-chat-edit",
    "image_outpainting": "image-outpainting",
    "product-recontext": "image-recontext",
}

def _normalize_mode_preset_id(mode_id: str) -> str:
    normalized = str(mode_id or "").strip().lower().replace("_", "-")
    if not normalized:
        return ""
    return MODE_WORKFLOW_PRESET_ALIASES.get(normalized, normalized)


def _build_mode_preset_summary(preset: Dict[str, Any]) -> Dict[str, Any]:
    workflow = preset.get("workflow") if isinstance(preset.get("workflow"), dict) else {}
    nodes = workflow.get("nodes") if isinstance(workflow.get("nodes"), list) else []
    edges = workflow.get("edges") if isinstance(workflow.get("edges"), list) else []
    return {
        "id": preset.get("id"),
        "mode": preset.get("mode"),
        "name": preset.get("name"),
        "description": preset.get("description"),
        "requires_image": bool(preset.get("requiresImage")),
        "prompt_hint": preset.get("promptHint") or "",
        "prompt_example": preset.get("promptExample"),
        "node_count": len(nodes),
        "edge_count": len(edges),
    }


# ==================== Workflow Execution Runtime ====================

@dataclass
class ExecutionStatePushRuntimeMetrics:
    emitted: int = 0
    dropped: int = 0
    skipped: int = 0
    last_emitted_at: int = 0
    last_dropped_at: int = 0
    last_skipped_at: int = 0


@dataclass
class LocalExecutionRuntime:
    subscribers: List[asyncio.Queue] = field(default_factory=list)
    metrics: RuntimeEventMetrics = field(default_factory=RuntimeEventMetrics)
    execution_state_push_metrics: ExecutionStatePushRuntimeMetrics = field(default_factory=ExecutionStatePushRuntimeMetrics)
    execution_state_cache: Optional[Dict[str, Any]] = None
    execution_state_last_emitted: Optional[Dict[str, Any]] = None
    execution_state_last_emitted_at: int = 0


_execution_runtime_local: Dict[str, LocalExecutionRuntime] = {}
_execution_tasks: Dict[str, asyncio.Task] = {}
_execution_idempotency_locks: Dict[str, asyncio.Lock] = {}
_execution_resume_locks: Dict[str, asyncio.Lock] = {}
_workflow_runtime_store = create_workflow_runtime_store()


def _get_local_runtime(execution_id: str) -> LocalExecutionRuntime:
    runtime = _execution_runtime_local.get(execution_id)
    if runtime:
        return runtime

    runtime = LocalExecutionRuntime()
    _execution_runtime_local[execution_id] = runtime
    return runtime


def _build_execution_state_push_metrics_snapshot(
    metrics: Optional[ExecutionStatePushRuntimeMetrics],
) -> Dict[str, Any]:
    return {
        "emitted": max(0, _safe_int(getattr(metrics, "emitted", 0), default=0)),
        "dropped": max(0, _safe_int(getattr(metrics, "dropped", 0), default=0)),
        "skipped": max(0, _safe_int(getattr(metrics, "skipped", 0), default=0)),
        "last_emitted_at": max(0, _safe_int(getattr(metrics, "last_emitted_at", 0), default=0)),
        "last_dropped_at": max(0, _safe_int(getattr(metrics, "last_dropped_at", 0), default=0)),
        "last_skipped_at": max(0, _safe_int(getattr(metrics, "last_skipped_at", 0), default=0)),
    }


def _build_runtime_metrics_snapshot(execution_id: str) -> Dict[str, Any]:
    runtime = _execution_runtime_local.get(execution_id)
    snapshot = _build_runtime_metrics_snapshot_helper(
        None if runtime is None else runtime.metrics,
        subscriber_count=0 if runtime is None else len(runtime.subscribers),
    )
    snapshot["execution_state_push"] = _build_execution_state_push_metrics_snapshot(
        None if runtime is None else runtime.execution_state_push_metrics
    )
    return snapshot


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _build_execution_state_payload(
    snapshot: Dict[str, Any],
    *,
    runtime_updated_at: int = 0,
    checkpoint_payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    normalized_status = _normalize_workflow_status(snapshot.get("status"))
    final_status = _resolve_workflow_final_status(normalized_status)
    is_terminal = normalized_status in WORKFLOW_TERMINAL_STATUSES or normalized_status == "paused"

    started_at = _safe_int(snapshot.get("startedAt"), default=0)
    completed_at = _safe_int(snapshot.get("completedAt"), default=0)
    state_version = max(_safe_int(runtime_updated_at, default=0), started_at, completed_at)

    payload: Dict[str, Any] = {
        **snapshot,
        "status": normalized_status,
        "finalStatus": final_status,
        "isTerminal": is_terminal,
        "stateVersion": state_version,
        "clientPolicy": dict(WORKFLOW_EXECUTION_CLIENT_POLICY),
    }

    if normalized_status == "paused":
        payload["checkpoint"] = _build_checkpoint_summary(checkpoint_payload)
        payload["pausedAt"] = runtime_updated_at if _safe_int(runtime_updated_at, default=0) > 0 else None

    return payload


def _cache_execution_state_payload(execution_id: str, payload: Dict[str, Any]) -> None:
    runtime = _get_local_runtime(execution_id)
    runtime.execution_state_cache = copy.deepcopy(payload if isinstance(payload, dict) else {})


def _resolve_execution_state_event_version(data: Dict[str, Any], *, fallback: int) -> int:
    candidates = [fallback]
    for key in ("stateVersion", "timestamp", "completedAt", "pausedAt", "resumedAt", "pauseRequestedAt", "updatedAt"):
        value = _safe_int(data.get(key), default=0)
        if value > 0:
            candidates.append(value)
    return max(candidates)


def _build_execution_state_change_key(payload: Dict[str, Any], *, include_version: bool = True) -> Dict[str, Any]:
    status = _normalize_workflow_status(payload.get("status"), default="pending")
    derived_is_terminal = status in WORKFLOW_TERMINAL_STATUSES or status == "paused"
    key: Dict[str, Any] = {
        "status": status,
        "finalStatus": _resolve_workflow_final_status(status),
        "isTerminal": bool(payload.get("isTerminal")) or derived_is_terminal,
        "nodeStatuses": copy.deepcopy(payload.get("nodeStatuses")) if isinstance(payload.get("nodeStatuses"), dict) else {},
        "nodeProgress": copy.deepcopy(payload.get("nodeProgress")) if isinstance(payload.get("nodeProgress"), dict) else {},
        "nodeErrors": copy.deepcopy(payload.get("nodeErrors")) if isinstance(payload.get("nodeErrors"), dict) else {},
    }
    if include_version:
        key["stateVersion"] = _safe_int(payload.get("stateVersion"), default=0)
    return key


def _has_execution_state_change(
    previous_payload: Optional[Dict[str, Any]],
    current_payload: Dict[str, Any],
    *,
    include_version: bool = True,
) -> bool:
    if not isinstance(previous_payload, dict):
        return True
    return _build_execution_state_change_key(previous_payload, include_version=include_version) != _build_execution_state_change_key(
        current_payload,
        include_version=include_version,
    )


def _is_execution_state_terminal(payload: Dict[str, Any]) -> bool:
    status = _normalize_workflow_status(payload.get("status"), default="running")
    return bool(payload.get("isTerminal")) or status in WORKFLOW_TERMINAL_STATUSES or status == "paused"


def _should_emit_execution_state(
    runtime: LocalExecutionRuntime,
    payload: Dict[str, Any],
    *,
    emitted_at: int,
) -> bool:
    if not _has_execution_state_change(runtime.execution_state_last_emitted, payload, include_version=True):
        return False
    if _is_execution_state_terminal(payload):
        return True
    last_emitted_at = _safe_int(runtime.execution_state_last_emitted_at, default=0)
    if last_emitted_at <= 0:
        return True
    return emitted_at - last_emitted_at >= EXECUTION_STATE_PUSH_MIN_INTERVAL_MS


def _mark_execution_state_emitted(runtime: LocalExecutionRuntime, payload: Dict[str, Any], *, emitted_at: int) -> None:
    runtime.execution_state_last_emitted = copy.deepcopy(payload if isinstance(payload, dict) else {})
    runtime.execution_state_last_emitted_at = max(0, _safe_int(emitted_at, default=0))


def _record_execution_state_push_delivery(
    runtime: LocalExecutionRuntime,
    *,
    emitted: int,
    dropped: int,
    emitted_at: int,
) -> None:
    metrics = runtime.execution_state_push_metrics
    safe_emitted = max(0, _safe_int(emitted, default=0))
    safe_dropped = max(0, _safe_int(dropped, default=0))
    now = max(0, _safe_int(emitted_at, default=0))

    metrics.emitted += safe_emitted
    metrics.dropped += safe_dropped

    if safe_emitted > 0:
        metrics.last_emitted_at = now
    if safe_dropped > 0:
        metrics.last_dropped_at = now


def _record_execution_state_push_skipped(
    runtime: LocalExecutionRuntime,
    *,
    skipped_at: int,
) -> None:
    metrics = runtime.execution_state_push_metrics
    metrics.skipped += 1
    metrics.last_skipped_at = max(0, _safe_int(skipped_at, default=0))


def _build_execution_state_from_runtime_event(
    execution_id: str,
    event_type: str,
    data: Dict[str, Any],
    *,
    emitted_at: int,
) -> Dict[str, Any]:
    runtime = _get_local_runtime(execution_id)
    payload = data if isinstance(data, dict) else {}
    state = runtime.execution_state_cache

    if not isinstance(state, dict):
        initial_status = _normalize_workflow_status(payload.get("status"), default="running")
        state = {
            "executionId": execution_id,
            "status": initial_status,
            "startedAt": None,
            "completedAt": None,
            "error": None,
            "result": None,
            "nodeStatuses": {},
            "nodeProgress": {},
            "nodeResults": {},
            "nodeErrors": {},
            "clientPolicy": dict(WORKFLOW_EXECUTION_CLIENT_POLICY),
        }
        state["finalStatus"] = _resolve_workflow_final_status(initial_status)
        state["isTerminal"] = initial_status in WORKFLOW_TERMINAL_STATUSES or initial_status == "paused"
        state["stateVersion"] = _resolve_execution_state_event_version(payload, fallback=emitted_at)
        runtime.execution_state_cache = state

    node_statuses = state.get("nodeStatuses")
    if not isinstance(node_statuses, dict):
        node_statuses = {}
        state["nodeStatuses"] = node_statuses

    node_progress = state.get("nodeProgress")
    if not isinstance(node_progress, dict):
        node_progress = {}
        state["nodeProgress"] = node_progress

    node_errors = state.get("nodeErrors")
    if not isinstance(node_errors, dict):
        node_errors = {}
        state["nodeErrors"] = node_errors

    node_results = state.get("nodeResults")
    if not isinstance(node_results, dict):
        node_results = {}
        state["nodeResults"] = node_results

    previous_change_key = _build_execution_state_change_key(state, include_version=False)

    node_id = str(payload.get("nodeId") or "").strip()
    if node_id:
        if event_type == "node_start":
            node_statuses[node_id] = "running"
            node_progress[node_id] = max(0, min(100, _safe_int(payload.get("progress"), default=0)))
            node_errors.pop(node_id, None)
        elif event_type == "node_progress":
            previous_progress = _safe_int(node_progress.get(node_id), default=0)
            node_progress[node_id] = max(0, min(100, _safe_int(payload.get("progress"), default=previous_progress)))
            current_node_status = _normalize_node_status(node_statuses.get(node_id), default="pending")
            if current_node_status not in {"completed", "failed", "skipped"}:
                node_statuses[node_id] = "running"
        elif event_type == "node_complete":
            node_statuses[node_id] = "completed"
            node_progress[node_id] = 100
            node_errors.pop(node_id, None)
            if "output" in payload:
                node_results[node_id] = payload.get("output")
        elif event_type == "node_error":
            node_statuses[node_id] = "failed"
            node_progress[node_id] = 100
            error_message = str(payload.get("error") or "").strip()
            if error_message:
                node_errors[node_id] = error_message
        elif event_type == "node_skipped":
            node_statuses[node_id] = "skipped"
            node_progress[node_id] = 100
            node_errors.pop(node_id, None)

    status_overrides = {
        "workflow_complete": "completed",
        "workflow_failed": "failed",
        "workflow_cancelled": "cancelled",
        "workflow_paused": "paused",
        "workflow_resumed": "running",
        "workflow_pause_requested": "running",
    }
    current_status = _normalize_workflow_status(state.get("status"), default="running")
    next_status = status_overrides.get(event_type)
    if not next_status and event_type.startswith("workflow_"):
        candidate_status = _normalize_workflow_status(payload.get("status"), default="")
        if candidate_status:
            next_status = candidate_status
    if not next_status and event_type.startswith("node_"):
        if current_status not in WORKFLOW_TERMINAL_STATUSES and current_status != "paused":
            next_status = "running"
    if next_status:
        current_status = next_status

    if event_type == "workflow_complete":
        if "result" in payload:
            state["result"] = payload.get("result")
        state["error"] = None
    elif event_type in {"workflow_failed", "workflow_cancelled"}:
        if "error" in payload:
            state["error"] = payload.get("error")
    elif event_type == "workflow_resumed":
        state["completedAt"] = None
        state["error"] = None
        state["result"] = None

    if "completedAt" in payload:
        state["completedAt"] = payload.get("completedAt")

    if current_status == "paused":
        if "checkpoint" in payload:
            state["checkpoint"] = payload.get("checkpoint")
        paused_at = _safe_int(payload.get("pausedAt"), default=0)
        if paused_at > 0:
            state["pausedAt"] = paused_at
    else:
        state.pop("checkpoint", None)
        state.pop("pausedAt", None)

    state["status"] = current_status
    state["finalStatus"] = _resolve_workflow_final_status(current_status)
    state["isTerminal"] = current_status in WORKFLOW_TERMINAL_STATUSES or current_status == "paused"

    previous_version = _safe_int(state.get("stateVersion"), default=0)
    has_key_change_excluding_version = _build_execution_state_change_key(state, include_version=False) != previous_change_key

    if has_key_change_excluding_version:
        version_candidate = _resolve_execution_state_event_version(payload, fallback=previous_version + 1)
        if version_candidate <= previous_version:
            version_candidate = previous_version + 1
    else:
        incoming_version = _safe_int(payload.get("stateVersion"), default=0)
        version_candidate = incoming_version if incoming_version > previous_version else previous_version

    state["stateVersion"] = max(
        version_candidate,
        _safe_int(state.get("startedAt"), default=0),
        _safe_int(state.get("completedAt"), default=0),
    )
    state["runtimeMetrics"] = _build_runtime_metrics_snapshot(execution_id)

    runtime.execution_state_cache = state
    return copy.deepcopy(state)


def _get_idempotency_lock(user_id: str, idempotency_key: str) -> asyncio.Lock:
    lock_key = f"{user_id}:{idempotency_key}"
    lock = _execution_idempotency_locks.get(lock_key)
    if lock:
        return lock

    lock = asyncio.Lock()
    _execution_idempotency_locks[lock_key] = lock
    return lock


def _get_resume_lock(execution_id: str) -> asyncio.Lock:
    lock = _execution_resume_locks.get(execution_id)
    if lock:
        return lock

    lock = asyncio.Lock()
    _execution_resume_locks[execution_id] = lock
    return lock


def _schedule_runtime_store_sync(
    awaitable: Any,
    *,
    execution_id: str,
    action: str,
) -> None:
    async def _runner() -> None:
        try:
            await awaitable
        except Exception:
            logger.debug(
                "[Workflow] background runtime-store %s failed for execution=%s",
                action,
                execution_id,
                exc_info=True,
            )

    try:
        loop = asyncio.get_running_loop()
    except RuntimeError:
        if inspect.iscoroutine(awaitable):
            awaitable.close()
        return

    loop.create_task(_runner())


async def _cleanup_execution_runtime(
    execution_id: str,
    *,
    clear_store: bool,
    clear_local_runtime: bool,
    clear_task: bool = True,
    clear_resume_lock: bool = True,
) -> None:
    if clear_store:
        try:
            await _workflow_runtime_store.clear(execution_id)
        except Exception:
            logger.debug("[Workflow] runtime store clear failed for execution=%s", execution_id, exc_info=True)

    if clear_local_runtime:
        _execution_runtime_local.pop(execution_id, None)
    if clear_task:
        _execution_tasks.pop(execution_id, None)
    if clear_resume_lock:
        _execution_resume_locks.pop(execution_id, None)


async def _publish_runtime_event(execution_id: str, event_type: str, data: Dict[str, Any]):
    runtime = _get_local_runtime(execution_id)
    now = int(time.time() * 1000)

    try:
        await _workflow_runtime_store.touch(execution_id)
    except Exception:
        logger.debug("[Workflow] runtime store touch failed for execution=%s", execution_id, exc_info=True)

    delivered = 0
    dropped = 0
    delivered_legacy_queues: List[asyncio.Queue] = []
    for queue in list(runtime.subscribers):
        try:
            queue.put_nowait({"event": event_type, "data": data})
            delivered += 1
            delivered_legacy_queues.append(queue)
        except asyncio.QueueFull:
            dropped += 1

    metrics = record_runtime_event_publish(
        runtime.metrics,
        delivered=delivered,
        dropped=dropped,
        emitted_at=now,
    )

    execution_state_payload = _build_execution_state_from_runtime_event(
        execution_id,
        event_type,
        data,
        emitted_at=now,
    )
    emitted_execution_state = 0
    dropped_execution_state = 0
    if not delivered_legacy_queues:
        _record_execution_state_push_skipped(runtime, skipped_at=now)
    elif _should_emit_execution_state(runtime, execution_state_payload, emitted_at=now):
        execution_state_message = {"event": "execution_state", "data": execution_state_payload}
        for queue in delivered_legacy_queues:
            try:
                queue.put_nowait(execution_state_message)
                emitted_execution_state += 1
            except asyncio.QueueFull:
                dropped_execution_state += 1
        _record_execution_state_push_delivery(
            runtime,
            emitted=emitted_execution_state,
            dropped=dropped_execution_state,
            emitted_at=now,
        )
        if emitted_execution_state > 0:
            _mark_execution_state_emitted(runtime, execution_state_payload, emitted_at=now)
    else:
        _record_execution_state_push_skipped(runtime, skipped_at=now)

    if dropped > 0:
        logger.warning(
            "[Workflow] SSE queue full, dropping event: execution=%s event=%s dropped=%s total_dropped=%s subscribers=%s",
            execution_id,
            event_type,
            dropped,
            metrics.dropped_event_count,
            len(runtime.subscribers),
        )
    if dropped_execution_state > 0:
        logger.debug(
            "[Workflow] SSE queue full, dropping execution_state: execution=%s source_event=%s dropped=%s subscribers=%s",
            execution_id,
            event_type,
            dropped_execution_state,
            len(runtime.subscribers),
        )


def _format_sse(event: str, data: Dict[str, Any]) -> str:
    return f"event: {event}\ndata: {json.dumps(data, ensure_ascii=False)}\n\n"


def _safe_json_loads(raw: Optional[str], default: Any = None):
    if not raw:
        return default
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return default


def _resolve_execute_idempotency_key(meta: Optional[Dict[str, Any]], request: Optional[Request]) -> str:
    if request is not None:
        headers = request.headers
        for field in WORKFLOW_IDEMPOTENCY_HEADER_FIELDS:
            candidate = _normalize_idempotency_key(headers.get(field))
            if candidate:
                return candidate

    if isinstance(meta, dict):
        for field in WORKFLOW_IDEMPOTENCY_META_FIELDS:
            candidate = _normalize_idempotency_key(meta.get(field))
            if candidate:
                return candidate
    return ""


def _extract_execution_meta(execution: WorkflowExecution) -> Dict[str, Any]:
    workflow_payload = _safe_json_loads(execution.workflow_json, {})
    if not isinstance(workflow_payload, dict):
        return {}
    meta_payload = workflow_payload.get("meta")
    if not isinstance(meta_payload, dict):
        return {}
    return meta_payload


def _extract_execution_idempotency_key(execution: WorkflowExecution) -> str:
    meta_payload = _extract_execution_meta(execution)
    for field in WORKFLOW_IDEMPOTENCY_META_FIELDS:
        candidate = _normalize_idempotency_key(meta_payload.get(field))
        if candidate:
            return candidate
    return ""


def _find_execution_by_idempotency_key(
    db: Session,
    *,
    user_id: str,
    idempotency_key: str,
) -> Optional[WorkflowExecution]:
    if not idempotency_key:
        return None

    # 新路径：优先走持久化字段查询（跨实例唯一约束配套）。
    try:
        direct_match = db.query(WorkflowExecution).filter(
            WorkflowExecution.user_id == user_id,
            WorkflowExecution.idempotency_key == idempotency_key,
        ).order_by(
            WorkflowExecution.started_at.desc(),
            WorkflowExecution.id.desc(),
        ).first()
        if direct_match:
            return direct_match
    except Exception:
        # 兼容旧 schema：字段尚未迁移完成时回退到历史扫描逻辑。
        pass

    # 兼容旧路径：从 workflow_json.meta 中扫描幂等键。
    recent_executions = db.query(WorkflowExecution).filter(
        WorkflowExecution.user_id == user_id,
    ).order_by(
        WorkflowExecution.started_at.desc(),
        WorkflowExecution.id.desc(),
    ).limit(WORKFLOW_IDEMPOTENCY_SCAN_LIMIT).all()

    for execution in recent_executions:
        if _extract_execution_idempotency_key(execution) == idempotency_key:
            return execution
    return None


def _build_idempotent_execute_response(execution: WorkflowExecution) -> Dict[str, Any]:
    status = _normalize_workflow_status(execution.status)
    response: Dict[str, Any] = {
        "execution_id": execution.id,
        "status": status,
        "idempotency_replay": True,
    }

    if status in WORKFLOW_TERMINAL_STATUSES:
        raw_result = _safe_json_loads(execution.result_json, None)
        response["result"] = _serialize_workflow_result(raw_result) if raw_result is not None else None
        response["events"] = []
        if status in {"failed", "cancelled"} and execution.error:
            response["error"] = execution.error

    return response


def _serialize_workflow_result(payload: Any) -> Any:
    """工作流 result 的局部序列化：统一为 camelCase（不依赖中间件对 result 递归）"""
    if payload is None:
        return payload
    try:
        return to_camel_case(payload)
    except Exception:
        logger.warning("[Workflow] Failed to serialize result payload to camelCase", exc_info=True)
        return payload


def _sanitize_history_detail_payload(
    payload: Any,
    max_text_chars: int = 2000,
    max_list_items: int = 40,
    max_dict_items: int = 80,
    depth: int = 0,
) -> Any:
    """
    历史详情返回前的轻量脱敏/裁剪：
    - 将 data:image / data:audio / data:video 以及 base64 二进制 data URL 替换为占位符，避免 10~100MB JSON 响应压垮前端。
    - 超长字符串进行截断，避免日志/模型原始响应撑爆面板。
    """
    if depth > 32:
        return "[depth-limit-reached]"

    if payload is None:
        return None

    if isinstance(payload, str):
        text = payload.strip()
        lowered = text.lower()
        if lowered.startswith("data:image/"):
            return "[inline-image-omitted; use history images preview/download endpoint]"
        if lowered.startswith("data:audio/"):
            return "[inline-audio-omitted; use media download/link endpoint]"
        if lowered.startswith("data:video/"):
            return "[inline-video-omitted; use media download/link endpoint]"
        if lowered.startswith("data:") and ";base64," in lowered:
            return "[inline-binary-data-url-omitted]"
        compact = text.replace("\n", "").replace("\r", "")
        if (
            len(compact) >= 2048
            and re.fullmatch(r"[A-Za-z0-9+/=]+", compact)
            and " " not in compact
        ):
            return f"[base64-blob-omitted; length={len(compact)}]"
        if len(payload) > max_text_chars:
            omitted = len(payload) - max_text_chars
            return f"{payload[:max_text_chars]}...[truncated {omitted} chars]"
        return payload

    if isinstance(payload, list):
        sanitized_items = [
            _sanitize_history_detail_payload(
                item,
                max_text_chars=max_text_chars,
                max_list_items=max_list_items,
                max_dict_items=max_dict_items,
                depth=depth + 1,
            )
            for item in payload[:max_list_items]
        ]
        if len(payload) > max_list_items:
            sanitized_items.append({
                "_truncated": True,
                "remainingItems": len(payload) - max_list_items,
            })
        return sanitized_items

    if isinstance(payload, dict):
        sanitized: Dict[str, Any] = {}
        items = list(payload.items())
        for index, (key, value) in enumerate(items):
            if index >= max_dict_items:
                sanitized["_truncated"] = {
                    "remainingKeys": len(items) - max_dict_items,
                }
                break
            sanitized[key] = _sanitize_history_detail_payload(
                value,
                max_text_chars=max_text_chars,
                max_list_items=max_list_items,
                max_dict_items=max_dict_items,
                depth=depth + 1,
            )
        return sanitized

    return payload


def _serialize_event_payload(event_type: str, data: Dict[str, Any]) -> Dict[str, Any]:
    """SSE 事件负载序列化，确保 result/output 等字段使用 camelCase"""
    serialized = dict(data or {})
    if event_type == "node_complete" and "output" in serialized:
        serialized["output"] = _serialize_workflow_result(serialized.get("output"))
        runtime_hints = _extract_runtime_hints(serialized.get("output"))
        primary_runtime = _pick_primary_runtime(runtime_hints)
        if primary_runtime and not serialized.get("runtime"):
            serialized["runtime"] = primary_runtime
    if event_type == "workflow_complete" and "result" in serialized:
        serialized["result"] = _serialize_workflow_result(serialized.get("result"))
        runtime_hints = _extract_runtime_hints(serialized.get("result"))
        primary_runtime = _pick_primary_runtime(runtime_hints)
        if runtime_hints:
            serialized["runtimeHints"] = runtime_hints
        if primary_runtime:
            serialized["primaryRuntime"] = primary_runtime
    return serialized


def _normalize_image_candidate(value: Any) -> Optional[str]:
    if not isinstance(value, str):
        return None
    candidate = value.strip()
    if not candidate:
        return None

    lowered = candidate.lower()
    clean_path = lowered.split("?")[0].split("#")[0]
    image_exts = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp", ".svg")
    non_image_exts = (".xlsx", ".xls", ".csv", ".json", ".txt", ".pdf", ".doc", ".docx", ".zip", ".tar", ".gz")

    if lowered.startswith(("data:image/", "blob:")):
        return candidate

    if lowered.startswith(("http://", "https://")):
        if clean_path.endswith(image_exts):
            return candidate
        if clean_path.endswith(non_image_exts):
            return None
        if any(token in clean_path for token in ("/image", "/images", "/attachments", "/uploads", "/generated", "/edited", "/expanded")):
            return candidate
        return None

    if candidate.startswith("/"):
        if clean_path.endswith(image_exts):
            return candidate
        return None

    if clean_path.endswith(image_exts):
        return candidate
    return None


def _extract_image_urls(payload: Any) -> List[str]:
    urls: List[str] = []
    seen = set()
    source_hints = {
        "sourceimageurl",
        "source_image_url",
        "referenceimageurl",
        "reference_image_url",
        "inputimageurl",
        "input_image_url",
        "originalimageurl",
        "original_image_url",
        "startimageurl",
        "start_image_url",
        "raw",
        "rawurl",
        "raw_url",
        "dataurl",
        "data_url",
    }
    source_container_hints = {
        "referenceimages",
        "reference_images",
        "referenceimage",
        "reference_image",
        "sourceimages",
        "source_images",
        "source",
        "source_input",
        "sourceinput",
        "original",
        "originalimage",
        "original_image",
        "input",
        "inputs",
    }

    def push(candidate: Any, key_hint: str = ""):
        normalized_hint = str(key_hint or "").strip().lower()
        if normalized_hint in source_hints:
            return
        normalized = _normalize_image_candidate(candidate)
        if normalized and normalized not in seen:
            seen.add(normalized)
            urls.append(normalized)

    def push_image_items(items: Any):
        if isinstance(items, list):
            for item in items:
                if isinstance(item, dict):
                    push(item.get("url"), key_hint="imageUrl")
                    push(item.get("imageUrl"), key_hint="imageUrl")
                    push(item.get("image_url"), key_hint="image_url")
                else:
                    push(item, key_hint="imageUrl")
            return
        if items is not None:
            push(items, key_hint="imageUrls")

    if isinstance(payload, dict):
        # 优先展示模型产出的图片字段，避免把 sourceImageUrl 识别为结果图。
        push(payload.get("imageUrl"), key_hint="imageUrl")
        push(payload.get("image_url"), key_hint="image_url")
        push_image_items(payload.get("imageUrls"))
        push_image_items(payload.get("image_urls"))
        push_image_items(payload.get("images"))
        if urls:
            return urls

    def walk(value: Any, key_hint: str = ""):
        push(value, key_hint=key_hint)
        if isinstance(value, list):
            for item in value:
                walk(item, key_hint=key_hint)
            return
        if isinstance(value, dict):
            for key, item in value.items():
                normalized_key = str(key).strip().lower()
                if normalized_key in source_container_hints:
                    continue
                walk(item, key_hint=normalized_key)

    walk(payload)
    return urls


def _normalize_result_media_candidate(value: Any, media_kind: str, key_hint: str = "") -> Optional[str]:
    if not isinstance(value, str):
        return None
    candidate = value.strip()
    if not candidate:
        return None

    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"audio", "video"}:
        return None

    lowered = candidate.lower()
    clean_path = lowered.split("?", 1)[0].split("#", 1)[0]
    file_exts = (
        (".aac", ".flac", ".m4a", ".mp3", ".oga", ".ogg", ".opus", ".pcm", ".wav", ".weba")
        if normalized_kind == "audio"
        else (".avi", ".m4v", ".mkv", ".mov", ".mp4", ".mpeg", ".mpg", ".ogv", ".webm")
    )
    non_media_exts = (
        ".bmp",
        ".csv",
        ".doc",
        ".docx",
        ".gif",
        ".jpeg",
        ".jpg",
        ".json",
        ".pdf",
        ".png",
        ".svg",
        ".txt",
        ".webp",
        ".xls",
        ".xlsx",
        ".zip",
    )

    if lowered.startswith(f"data:{normalized_kind}/"):
        return candidate

    hint = str(key_hint or "").strip().lower()
    if "source" in hint or "reference" in hint or "input" in hint or "original" in hint or "start" in hint:
        return None

    if normalized_kind == "video":
        if candidate.startswith(("files/", "gs://")):
            return candidate
        if lowered.startswith("https://") and "/files/" in clean_path:
            return candidate

    path_tokens = (
        ("/audio", "/speech", "/voice", "/tts", "/narration")
        if normalized_kind == "audio"
        else ("/video", "/movie", "/clip", "/veo", "/sora")
    )

    if lowered.startswith(("http://", "https://")):
        if normalized_kind in hint:
            return candidate
        if clean_path.endswith(file_exts):
            return candidate
        if clean_path.endswith(non_media_exts):
            return None
        if any(token in clean_path for token in path_tokens):
            return candidate
        return None

    if candidate.startswith("/api/"):
        if normalized_kind in hint:
            return candidate
        if clean_path.endswith(file_exts):
            return candidate
        if any(token in clean_path for token in path_tokens):
            return candidate
        return None

    return None


def _extract_media_urls(payload: Any, media_kind: str) -> List[str]:
    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"audio", "video"}:
        return []

    urls: List[str] = []
    seen = set()
    source_hints = {
        f"source{normalized_kind}url",
        f"source_{normalized_kind}_url",
        f"reference{normalized_kind}url",
        f"reference_{normalized_kind}_url",
        f"input{normalized_kind}url",
        f"input_{normalized_kind}_url",
        f"original{normalized_kind}url",
        f"original_{normalized_kind}_url",
        f"start{normalized_kind}url",
        f"start_{normalized_kind}_url",
        "raw",
        "rawurl",
        "raw_url",
        "dataurl",
        "data_url",
    }
    if normalized_kind == "video":
        source_hints.update(
            {
                "provider_file_name",
                "providerfilename",
                "provider_file_uri",
                "providerfileuri",
                "gcs_uri",
                "gcsuri",
                "file_uri",
                "fileuri",
                "google_file_uri",
                "googlefileuri",
            }
        )
    source_container_hints = {
        f"reference{normalized_kind}s",
        f"reference_{normalized_kind}s",
        f"reference{normalized_kind}",
        f"reference_{normalized_kind}",
        f"source{normalized_kind}s",
        f"source_{normalized_kind}s",
        f"source{normalized_kind}",
        f"source_{normalized_kind}",
        "source",
        "source_input",
        "sourceinput",
        "original",
        f"original{normalized_kind}",
        f"original_{normalized_kind}",
        "input",
        "inputs",
    }

    def push(candidate: Any, key_hint: str = ""):
        normalized_hint = str(key_hint or "").strip().lower()
        if normalized_hint in source_hints:
            return
        normalized = _normalize_result_media_candidate(candidate, normalized_kind, key_hint=key_hint)
        if normalized and normalized not in seen:
            seen.add(normalized)
            urls.append(normalized)

    def push_media_items(items: Any):
        if isinstance(items, list):
            for item in items:
                if isinstance(item, dict):
                    push(item.get("url"), key_hint=f"{normalized_kind}Url")
                    push(item.get(f"{normalized_kind}Url"), key_hint=f"{normalized_kind}Url")
                    push(item.get(f"{normalized_kind}_url"), key_hint=f"{normalized_kind}_url")
                else:
                    push(item, key_hint=f"{normalized_kind}Url")
            return
        if items is not None:
            push(items, key_hint=f"{normalized_kind}Urls")

    if isinstance(payload, dict):
        if _payload_dict_matches_media_kind(payload, normalized_kind):
            push(payload.get("url"), key_hint=f"{normalized_kind}Url")
        push(payload.get(f"{normalized_kind}Url"), key_hint=f"{normalized_kind}Url")
        push(payload.get(f"{normalized_kind}_url"), key_hint=f"{normalized_kind}_url")
        push_media_items(payload.get(f"{normalized_kind}Urls"))
        push_media_items(payload.get(f"{normalized_kind}_urls"))
        push_media_items(payload.get(f"{normalized_kind}s"))
        if urls:
            return urls

    def walk(value: Any, key_hint: str = ""):
        push(value, key_hint=key_hint)
        if isinstance(value, list):
            for item in value:
                walk(item, key_hint=key_hint)
            return
        if isinstance(value, dict):
            for key, item in value.items():
                normalized_key = str(key).strip().lower()
                if normalized_key in source_container_hints:
                    continue
                walk(item, key_hint=normalized_key)

    walk(payload)
    return urls


def _extract_audio_urls(payload: Any) -> List[str]:
    return _extract_media_urls(payload, "audio")


def _extract_video_urls(payload: Any) -> List[str]:
    return _extract_media_urls(payload, "video")


def _extract_media_urls_by_kind(payload: Any, media_kind: str) -> List[str]:
    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind == "image":
        return _extract_image_urls(payload)
    if normalized_kind == "audio":
        return _extract_audio_urls(payload)
    if normalized_kind == "video":
        return _extract_video_urls(payload)
    return []


def _is_non_result_workflow_node_id(node_id: Any) -> bool:
    normalized = str(node_id or "").strip().lower()
    if not normalized:
        return False
    return (
        normalized.startswith("start")
        or normalized.startswith("input-")
        or normalized.startswith("input_")
    )


def _extract_workflow_output_media_urls(payload: Any, media_kind: str) -> List[str]:
    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"image", "audio", "video"}:
        return []

    extract_urls = lambda value: _extract_media_urls_by_kind(value, normalized_kind)
    if not isinstance(payload, dict):
        return extract_urls(payload)

    looks_like_workflow_result = any(
        key in payload
        for key in ("finalOutput", "final_output", "outputs", "outputsMap", "finalNodeId", "final_node_id", "nodeStates")
    )

    final_output = payload.get("finalOutput")
    if final_output is None:
        final_output = payload.get("final_output")
    if final_output is not None:
        final_output_urls = extract_urls(final_output)
        if final_output_urls:
            return final_output_urls

    outputs_payload = payload.get("outputs")
    if outputs_payload is None:
        outputs_payload = payload.get("outputsMap")
    if isinstance(outputs_payload, dict):
        final_node_id = str(payload.get("finalNodeId") or payload.get("final_node_id") or "").strip()
        if final_node_id and not _is_non_result_workflow_node_id(final_node_id):
            final_node_urls = extract_urls(outputs_payload.get(final_node_id))
            if final_node_urls:
                return final_node_urls

        collected_urls: List[str] = []
        seen_urls = set()
        for node_id, node_output in outputs_payload.items():
            if _is_non_result_workflow_node_id(node_id):
                continue
            for media_url in extract_urls(node_output):
                if media_url in seen_urls:
                    continue
                seen_urls.add(media_url)
                collected_urls.append(media_url)
        if collected_urls:
            return collected_urls

    if looks_like_workflow_result:
        return []
    return extract_urls(payload)


def _guess_image_mime_type(image_url: str) -> str:
    raw = str(image_url or "").strip()
    if not raw:
        return "image/png"

    lowered = raw.lower()
    if lowered.startswith("data:image/"):
        header = raw.split(",", 1)[0]
        mime = str(header.split(":", 1)[1].split(";", 1)[0] if ":" in header else "").strip().lower()
        if mime.startswith("image/"):
            return mime
        return "image/png"

    clean_url = raw.split("?", 1)[0].split("#", 1)[0]
    guessed, _ = mimetypes.guess_type(clean_url)
    normalized = str(guessed or "").strip().lower()
    if normalized.startswith("image/"):
        return normalized
    return "image/png"


def _guess_media_mime_type(media_url: str, media_kind: str) -> str:
    raw = str(media_url or "").strip()
    normalized_kind = str(media_kind or "").strip().lower()
    default_mime = "video/mp4" if normalized_kind == "video" else "audio/mpeg"
    if not raw:
        return default_mime

    lowered = raw.lower()
    if lowered.startswith(f"data:{normalized_kind}/"):
        header = raw.split(",", 1)[0]
        mime = str(header.split(":", 1)[1].split(";", 1)[0] if ":" in header else "").strip().lower()
        if mime.startswith(f"{normalized_kind}/"):
            return mime
        return default_mime

    clean_url = raw.split("?", 1)[0].split("#", 1)[0]
    guessed, _ = mimetypes.guess_type(clean_url)
    normalized = str(guessed or "").strip().lower()
    if normalized.startswith(f"{normalized_kind}/"):
        return normalized
    return default_mime


def _replace_payload_image_urls(payload: Any, replacements: Dict[str, str]) -> Any:
    if not replacements:
        return payload
    if isinstance(payload, str):
        return replacements.get(payload, payload)
    if isinstance(payload, list):
        return [_replace_payload_image_urls(item, replacements) for item in payload]
    if isinstance(payload, dict):
        return {key: _replace_payload_image_urls(value, replacements) for key, value in payload.items()}
    return payload


def _guess_workflow_media_mime_type(media_url: str, media_kind: str) -> str:
    raw = str(media_url or "").strip()
    normalized_kind = str(media_kind or "").strip().lower()
    if not raw:
        return "video/mp4" if normalized_kind == "video" else "audio/mpeg"

    lowered = raw.lower()
    if lowered.startswith(f"data:{normalized_kind}/"):
        header = raw.split(",", 1)[0]
        mime = str(header.split(":", 1)[1].split(";", 1)[0] if ":" in header else "").strip().lower()
        if mime.startswith(f"{normalized_kind}/"):
            return mime

    clean_url = raw.split("?", 1)[0].split("#", 1)[0]
    guessed, _ = mimetypes.guess_type(clean_url)
    normalized = str(guessed or "").strip().lower()
    if normalized.startswith(f"{normalized_kind}/"):
        return normalized

    if normalized_kind == "video":
        return "video/mp4"
    return "audio/mpeg"


def _normalize_workflow_persistable_media_source(value: Any, media_kind: str) -> Optional[str]:
    if not isinstance(value, str):
        return None
    candidate = value.strip()
    if not candidate:
        return None
    lowered = candidate.lower()
    normalized_kind = str(media_kind or "").strip().lower()

    if "{{" in candidate or "}}" in candidate:
        return None
    if lowered.startswith(f"data:{normalized_kind}/"):
        return candidate
    if lowered.startswith(("http://", "https://", "/api/")):
        return candidate
    if normalized_kind == "video" and is_google_provider_video_uri(candidate):
        return candidate
    return None


def _extract_workflow_media_entry(payload: Any, media_kind: str) -> Optional[Dict[str, str]]:
    if not isinstance(payload, dict):
        return None

    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"audio", "video"}:
        return None

    media_url_keys = (
        ("videoUrl", "video_url", "url", "video")
        if normalized_kind == "video"
        else ("audioUrl", "audio_url", "url")
    )
    list_keys = (
        ("videoUrls", "video_urls", "videos", "resultPreviewVideoUrls")
        if normalized_kind == "video"
        else ("audioUrls", "audio_urls", "audios", "resultPreviewAudioUrls")
    )
    mime_type = str(payload.get("mime_type") or payload.get("mimeType") or "").strip().lower()

    source = ""
    for key in media_url_keys:
        candidate = _normalize_workflow_persistable_media_source(payload.get(key), normalized_kind)
        if candidate:
            source = candidate
            break

    if not source:
        for key in list_keys:
            value = payload.get(key)
            if isinstance(value, list):
                for item in value:
                    candidate = _normalize_workflow_persistable_media_source(item, normalized_kind)
                    if candidate:
                        source = candidate
                        break
            else:
                candidate = _normalize_workflow_persistable_media_source(value, normalized_kind)
                if candidate:
                    source = candidate
            if source:
                break

    provider_file_name = ""
    provider_file_uri = ""
    gcs_uri = ""
    file_uri = ""
    if normalized_kind == "video":
        provider_file_name = str(payload.get("provider_file_name") or payload.get("providerFileName") or "").strip()
        provider_file_uri = str(payload.get("provider_file_uri") or payload.get("providerFileUri") or "").strip()
        gcs_uri = str(payload.get("gcs_uri") or payload.get("gcsUri") or "").strip()
        file_uri = str(payload.get("file_uri") or payload.get("fileUri") or "").strip()

    if not source:
        source = (
            _normalize_workflow_persistable_media_source(file_uri, normalized_kind)
            or _normalize_workflow_persistable_media_source(gcs_uri, normalized_kind)
            or _normalize_workflow_persistable_media_source(provider_file_uri, normalized_kind)
            or _normalize_workflow_persistable_media_source(provider_file_name, normalized_kind)
            or ""
        )

    has_media_shape = any(key in payload for key in (*media_url_keys, *list_keys))
    if normalized_kind == "video":
        has_media_shape = has_media_shape or any(
            key in payload for key in ("provider_file_name", "providerFileName", "provider_file_uri", "providerFileUri", "gcs_uri", "gcsUri", "file_uri", "fileUri")
        )

    if not source:
        return None
    if not has_media_shape and not mime_type.startswith(f"{normalized_kind}/"):
        return None

    return {
        "source": source,
        "mime_type": mime_type or _guess_workflow_media_mime_type(source, normalized_kind),
        "filename": str(payload.get("filename") or payload.get("name") or "").strip(),
        "provider_file_name": provider_file_name,
        "provider_file_uri": provider_file_uri,
        "gcs_uri": gcs_uri,
        "file_uri": file_uri,
    }


async def _persist_workflow_result_media(
    db: Session,
    execution_id: str,
    user_id: str,
    result_payload: Any,
    media_kind: str,
) -> Tuple[Any, Dict[str, str]]:
    if result_payload is None:
        return result_payload, {}

    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"audio", "video"}:
        return result_payload, {}

    active_storage = db.query(ActiveStorage).filter(
        ActiveStorage.user_id == user_id
    ).first()
    storage_id = active_storage.storage_id if active_storage and active_storage.storage_id else None
    attachment_service = AttachmentService(db)
    processed_sources: Dict[str, Dict[str, Any]] = {}
    replacements: Dict[str, str] = {}

    async def persist_source(entry: Dict[str, str], index: int) -> Optional[Dict[str, Any]]:
        source = str(entry.get("source") or "").strip()
        if not source or source.startswith("/api/temp-images/") or source.startswith("blob:"):
            return None
        cached = processed_sources.get(source)
        if cached is not None:
            return cached
        processed = await attachment_service.process_ai_result(
            ai_url=source,
            mime_type=str(entry.get("mime_type") or "").strip() or _guess_workflow_media_mime_type(source, normalized_kind),
            session_id=execution_id,
            message_id=execution_id,
            user_id=user_id,
            prefix=f"workflow-{normalized_kind}-result-{index:02d}",
            storage_id=storage_id,
            filename=str(entry.get("filename") or "").strip() or None,
            file_uri=str(entry.get("file_uri") or "").strip() or None,
            provider_file_name=str(entry.get("provider_file_name") or "").strip() or None,
            provider_file_uri=str(entry.get("provider_file_uri") or "").strip() or None,
            gcs_uri=str(entry.get("gcs_uri") or "").strip() or None,
        )
        processed_sources[source] = processed
        replacements[source] = str(processed.get("display_url") or "").strip() or source
        return processed

    async def walk(value: Any, key_hint: str = "", counter: Optional[List[int]] = None) -> Any:
        counter = counter or [0]
        normalized_hint = str(key_hint or "").strip()
        provider_hints = {
            "provider_file_name",
            "providerFileName",
            "provider_file_uri",
            "providerFileUri",
            "gcs_uri",
            "gcsUri",
            "file_uri",
            "fileUri",
            "google_file_uri",
            "googleFileUri",
        }
        direct_url_hints = {
            f"{normalized_kind}Url",
            f"{normalized_kind}_url",
            "url",
        }
        list_hints = {
            f"{normalized_kind}Urls",
            f"{normalized_kind}_urls",
            f"{normalized_kind}s",
            f"resultPreview{normalized_kind.capitalize()}Urls",
        }

        if isinstance(value, dict):
            entry = _extract_workflow_media_entry(value, normalized_kind)
            if entry:
                counter[0] += 1
                processed = await persist_source(entry, counter[0])
                if processed:
                    display_url = str(processed.get("display_url") or entry["source"]).strip()
                    updated = dict(value)
                    updated[f"{normalized_kind}Url"] = display_url
                    updated[f"{normalized_kind}Urls"] = [display_url]
                    if str(updated.get("url") or "").strip():
                        updated["url"] = display_url
                    updated["attachment_id"] = processed.get("attachment_id")
                    updated["upload_status"] = processed.get("status")
                    updated["task_id"] = processed.get("task_id")
                    updated["cloud_url"] = processed.get("cloud_url") or ""
                    updated["mime_type"] = processed.get("mime_type") or entry.get("mime_type") or updated.get("mime_type") or updated.get("mimeType")
                    updated["mimeType"] = updated.get("mime_type") or updated.get("mimeType")
                    if processed.get("filename"):
                        updated["filename"] = processed.get("filename")
                    if processed.get("file_uri") and "file_uri" not in updated and "fileUri" not in updated:
                        updated["file_uri"] = processed.get("file_uri")
                    if processed.get("google_file_uri") and "google_file_uri" not in updated and "googleFileUri" not in updated:
                        updated["google_file_uri"] = processed.get("google_file_uri")
                    return {
                        key: await walk(item, key, counter)
                        for key, item in updated.items()
                    }
            return {
                key: await walk(item, key, counter)
                for key, item in value.items()
            }

        if isinstance(value, list):
            return [await walk(item, normalized_hint, counter) for item in value]

        if not isinstance(value, str):
            return value

        if normalized_hint in provider_hints:
            return value

        candidate = _normalize_workflow_persistable_media_source(value, normalized_kind)
        if not candidate:
            return value

        if normalized_hint and normalized_hint not in direct_url_hints and normalized_hint not in list_hints:
            return value

        counter[0] += 1
        processed = await persist_source(
            {
                "source": candidate,
                "mime_type": _guess_workflow_media_mime_type(candidate, normalized_kind),
                "filename": "",
                "provider_file_name": "",
                "provider_file_uri": "",
                "gcs_uri": "",
                "file_uri": "",
            },
            counter[0],
        )
        if not processed:
            return value
        return str(processed.get("display_url") or value).strip() or value

    try:
        persisted_payload = await walk(result_payload)
    except Exception as exc:
        logger.warning(f"[Workflow] Failed to persist result {normalized_kind} via worker pool: {exc}")
        return result_payload, replacements

    return persisted_payload, replacements


def _build_result_preview_media_urls(
    execution_id: str,
    media_kind: str,
    replacements: Dict[str, str],
) -> List[str]:
    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"audio", "video"} or not replacements:
        return []
    return [
        f"/api/workflows/history/{execution_id}/{normalized_kind}/items/{index}"
        for index in range(1, len(replacements) + 1)
    ]


def _payload_dict_matches_media_kind(payload: Dict[str, Any], media_kind: str) -> bool:
    normalized_kind = str(media_kind or "").strip().lower()
    if normalized_kind not in {"audio", "video"}:
        return False

    mime_type = str(payload.get("mimeType") or payload.get("mime_type") or "").strip().lower()
    if mime_type.startswith(f"{normalized_kind}/"):
        return True

    direct_keys = {
        f"{normalized_kind}Url",
        f"{normalized_kind}_url",
        f"{normalized_kind}Urls",
        f"{normalized_kind}_urls",
    }
    if any(key in payload for key in direct_keys):
        return True

    if normalized_kind == "video" and any(
        key in payload
        for key in (
            "provider_file_name",
            "providerFileName",
            "provider_file_uri",
            "providerFileUri",
            "gcs_uri",
            "gcsUri",
            "file_uri",
            "fileUri",
            "google_file_uri",
            "googleFileUri",
        )
    ):
        return True

    type_hint = str(payload.get("type") or payload.get("kind") or "").strip().lower()
    return type_hint == normalized_kind


def _replace_payload_media_urls(
    payload: Any,
    replacements: Dict[str, str],
    media_kind: str,
    *,
    key_hint: str = "",
    in_media_container: bool = False,
) -> Any:
    if not replacements:
        return payload

    normalized_kind = str(media_kind or "").strip().lower()
    normalized_hint = str(key_hint or "").strip().lower()
    preserve_keys = {
        "provider_file_name",
        "providerfilename",
        "provider_file_uri",
        "providerfileuri",
        "gcs_uri",
        "gcsuri",
        "file_uri",
        "fileuri",
        "google_file_uri",
        "googlefileuri",
    }
    media_list_keys = {
        f"{normalized_kind}urls",
        f"{normalized_kind}_urls",
        f"{normalized_kind}s",
        f"resultpreview{normalized_kind}urls",
    }
    media_url_keys = {
        f"{normalized_kind}url",
        f"{normalized_kind}_url",
        "url",
    }

    if isinstance(payload, str):
        if payload not in replacements:
            return payload
        if normalized_hint in preserve_keys:
            return payload
        if in_media_container or not normalized_hint or normalized_hint in media_list_keys or normalized_hint in media_url_keys:
            return replacements.get(payload, payload)
        return payload

    if isinstance(payload, list):
        next_in_media_container = in_media_container or normalized_hint in media_list_keys
        return [
            _replace_payload_media_urls(
                item,
                replacements,
                normalized_kind,
                key_hint="",
                in_media_container=next_in_media_container,
            )
            for item in payload
        ]

    if isinstance(payload, dict):
        dict_is_media_container = in_media_container or _payload_dict_matches_media_kind(payload, normalized_kind)
        return {
            key: _replace_payload_media_urls(
                value,
                replacements,
                normalized_kind,
                key_hint=str(key),
                in_media_container=dict_is_media_container,
            )
            for key, value in payload.items()
        }

    return payload


async def _persist_workflow_result_images(
    db: Session,
    execution_id: str,
    user_id: str,
    result_payload: Any,
) -> Tuple[Any, Dict[str, str]]:
    if result_payload is None:
        return result_payload, {}

    extracted_urls = _extract_image_urls(result_payload)
    if not extracted_urls:
        return result_payload, {}

    candidates: List[str] = []
    seen = set()
    for url in extracted_urls:
        raw = str(url or "").strip()
        if not raw or raw in seen:
            continue
        lowered = raw.lower()
        if lowered.startswith("/api/temp-images/"):
            continue
        if lowered.startswith("blob:"):
            continue
        if not lowered.startswith(("data:image/", "http://", "https://")):
            continue
        seen.add(raw)
        candidates.append(raw)

    if not candidates:
        return result_payload, {}

    active_storage = db.query(ActiveStorage).filter(
        ActiveStorage.user_id == user_id
    ).first()
    storage_id = active_storage.storage_id if active_storage and active_storage.storage_id else None
    attachment_service = AttachmentService(db)
    replacements: Dict[str, str] = {}

    for index, image_url in enumerate(candidates, start=1):
        try:
            processed = await attachment_service.process_ai_result(
                ai_url=image_url,
                mime_type=_guess_image_mime_type(image_url),
                session_id=execution_id,
                message_id=execution_id,
                user_id=user_id,
                prefix=f"workflow-result-{index:02d}",
                storage_id=storage_id,
            )
            display_url = str(processed.get("display_url") or "").strip()
            replacements[image_url] = display_url or image_url
        except Exception as exc:
            logger.warning(f"[Workflow] Failed to persist result image via worker pool: {exc}")

    if not replacements:
        return result_payload, {}

    return _replace_payload_image_urls(result_payload, replacements), replacements


def _build_workflow_images_zip(
    execution_id: str,
    image_urls: List[str],
    request_base_url: str,
    inherited_headers: Dict[str, str],
) -> tuple[bytes, Dict[str, Any]]:
    return _build_workflow_images_zip_shared(
        execution_id=execution_id,
        image_urls=image_urls,
        trusted_base_url=request_base_url,
        inherited_headers=inherited_headers,
    )


def _build_workflow_image_previews(
    execution_id: str,
    image_urls: List[str],
    request_base_url: str,
    inherited_headers: Dict[str, str],
    limit: Optional[int] = None,
    max_bytes_per_image: Optional[int] = None,
    max_total_bytes: Optional[int] = None,
) -> Dict[str, Any]:
    return _build_workflow_image_previews_shared(
        execution_id=execution_id,
        image_urls=image_urls,
        trusted_base_url=request_base_url,
        inherited_headers=inherited_headers,
        limit=limit,
        max_bytes_per_image=max_bytes_per_image,
        max_total_bytes=max_total_bytes,
    )


def _build_workflow_media_zip(
    execution_id: str,
    media_kind: str,
    media_urls: List[str],
    request_base_url: str,
    inherited_headers: Dict[str, str],
    user_id: Optional[str] = None,
) -> tuple[bytes, Dict[str, Any]]:
    return _build_workflow_media_zip_shared(
        execution_id=execution_id,
        media_kind=media_kind,
        media_urls=media_urls,
        trusted_base_url=request_base_url,
        inherited_headers=inherited_headers,
        user_id=user_id,
    )


def _build_workflow_media_previews(
    execution_id: str,
    media_kind: str,
    media_urls: List[str],
    request_base_url: str,
    preview_path_template: str,
    limit: Optional[int] = None,
) -> Dict[str, Any]:
    return _build_workflow_media_previews_shared(
        execution_id=execution_id,
        media_kind=media_kind,
        media_urls=media_urls,
        trusted_base_url=request_base_url,
        preview_path_template=preview_path_template,
        limit=limit,
    )


def _resolve_workflow_history_trusted_base_url() -> str:
    configured_base_url = str(settings.workflow_history_public_base_url or "").strip()
    try:
        return _resolve_trusted_workflow_history_base_url_shared(
            configured_base_url=configured_base_url,
            fallback_port=settings.port,
        )
    except ValueError as exc:
        logger.warning(
            "[Workflow] Invalid WORKFLOW_HISTORY_PUBLIC_BASE_URL=%r; fallback to loopback origin: %s",
            configured_base_url,
            exc,
        )
        return _resolve_trusted_workflow_history_base_url_shared(
            configured_base_url=None,
            fallback_port=settings.port,
        )


def _excel_safe_sheet_name(name: str, used_names: set[str]) -> str:
    normalized = re.sub(r"[\\/*?:\[\]]+", " ", str(name or "").strip())
    normalized = re.sub(r"\s+", " ", normalized).strip() or "Sheet"
    base_name = normalized[:31] if len(normalized) > 31 else normalized
    candidate = base_name
    index = 2
    while candidate in used_names:
        suffix = f"-{index}"
        keep_len = max(1, 31 - len(suffix))
        candidate = f"{base_name[:keep_len]}{suffix}"
        index += 1
    used_names.add(candidate)
    return candidate


def _to_excel_cell_value(value: Any, max_length: int = 32000) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value

    if isinstance(value, str):
        normalized = value
    else:
        try:
            normalized = json.dumps(value, ensure_ascii=False)
        except Exception:
            normalized = str(value)

    if len(normalized) <= max_length:
        return normalized
    return f"{normalized[:max_length]}...[truncated:{len(normalized) - max_length}]"


def _flatten_payload_for_excel(
    payload: Any,
    *,
    prefix: str = "result",
    max_rows: int = 8000,
    max_depth: int = 10,
) -> List[Tuple[str, Any, str]]:
    rows: List[Tuple[str, Any, str]] = []

    def walk(value: Any, path: str, depth: int):
        if len(rows) >= max_rows:
            return
        if depth > max_depth:
            rows.append((path or "result", "[max depth reached]", "meta"))
            return

        if isinstance(value, dict):
            if not value:
                rows.append((path or "result", "{}", "dict"))
                return
            for key, item in value.items():
                next_path = f"{path}.{key}" if path else str(key)
                walk(item, next_path, depth + 1)
            return

        if isinstance(value, list):
            if not value:
                rows.append((path or "result", "[]", "list"))
                return
            max_list_items = 120
            for idx, item in enumerate(value[:max_list_items]):
                next_path = f"{path}[{idx}]"
                walk(item, next_path, depth + 1)
                if len(rows) >= max_rows:
                    return
            if len(value) > max_list_items and len(rows) < max_rows:
                rows.append(
                    (
                        f"{path}.__truncated__",
                        f"{len(value) - max_list_items} items omitted",
                        "meta",
                    )
                )
            return

        rows.append((path or "result", value, type(value).__name__))

    walk(payload, prefix, 0)
    if not rows:
        rows.append((prefix or "result", "", "empty"))
    return rows


def _looks_like_table_rows(value: Any) -> bool:
    if not isinstance(value, list) or not value:
        return False
    preview = value[: min(len(value), 50)]
    if not preview:
        return False
    if not all(isinstance(item, dict) for item in preview):
        return False
    has_columns = any(bool(item) for item in preview)
    return has_columns


def _collect_table_candidates(
    payload: Any,
    *,
    root_path: str = "result",
    max_tables: int = 8,
    max_rows_per_table: int = 3000,
) -> List[Dict[str, Any]]:
    tables: List[Dict[str, Any]] = []

    def walk(value: Any, path: str, depth: int):
        if len(tables) >= max_tables:
            return
        if depth > 10:
            return

        if _looks_like_table_rows(value):
            rows = value if isinstance(value, list) else []
            columns: List[str] = []
            seen_columns = set()
            for row in rows[:max_rows_per_table]:
                if not isinstance(row, dict):
                    continue
                for column_key in row.keys():
                    normalized_key = str(column_key)
                    if normalized_key in seen_columns:
                        continue
                    seen_columns.add(normalized_key)
                    columns.append(normalized_key)

            if columns:
                row_payloads: List[Dict[str, Any]] = []
                for row in rows[:max_rows_per_table]:
                    if isinstance(row, dict):
                        row_payloads.append({column: row.get(column) for column in columns})
                tables.append({
                    "path": path,
                    "columns": columns,
                    "rows": row_payloads,
                    "row_count": len(rows),
                    "truncated": len(rows) > max_rows_per_table,
                })
            return

        if isinstance(value, dict):
            for key, item in value.items():
                next_path = f"{path}.{key}" if path else str(key)
                walk(item, next_path, depth + 1)
            return

        if isinstance(value, list):
            max_branches = min(len(value), 16)
            for idx, item in enumerate(value[:max_branches]):
                next_path = f"{path}[{idx}]"
                walk(item, next_path, depth + 1)

    walk(payload, root_path, 0)
    return tables


def _build_workflow_analysis_excel_bytes(
    execution: WorkflowExecution,
    result_payload: Any,
    node_executions: List[NodeExecution],
) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except Exception as exc:
        raise RuntimeError("Excel export requires openpyxl dependency") from exc

    workbook = Workbook()
    bold_font = Font(bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    used_sheet_names: set[str] = set()

    summary_ws = workbook.active
    summary_ws.title = _excel_safe_sheet_name("Summary", used_sheet_names)
    duration_ms = None
    if execution.started_at and execution.completed_at:
        duration_ms = max(0, int(execution.completed_at) - int(execution.started_at))

    result_summary = _build_workflow_result_summary(result_payload)
    summary_rows = [
        ("Execution ID", execution.id),
        ("Status", execution.status),
        ("Started At", execution.started_at),
        ("Completed At", execution.completed_at),
        ("Duration (ms)", duration_ms if duration_ms is not None else ""),
        ("Error", execution.error or ""),
        ("Result Preview", result_summary.get("text_preview") or ""),
        ("Result Image Count", result_summary.get("image_count") or 0),
        ("Result Image URLs", "\n".join(result_summary.get("image_urls") or [])),
        ("Exported At", int(time.time() * 1000)),
    ]
    for row_index, (key, value) in enumerate(summary_rows, start=1):
        key_cell = summary_ws.cell(row=row_index, column=1, value=key)
        value_cell = summary_ws.cell(row=row_index, column=2, value=_to_excel_cell_value(value))
        key_cell.font = bold_font
        value_cell.alignment = wrap_alignment
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 120

    flattened_source = result_payload
    if isinstance(result_payload, dict):
        if "finalOutput" in result_payload:
            flattened_source = result_payload.get("finalOutput")
        elif "final_output" in result_payload:
            flattened_source = result_payload.get("final_output")

    flattened_ws = workbook.create_sheet(_excel_safe_sheet_name("Flattened", used_sheet_names))
    flattened_ws.append(["path", "value", "valueType"])
    for header_cell in flattened_ws[1]:
        header_cell.font = bold_font
    for path, value, value_type in _flatten_payload_for_excel(flattened_source, prefix="result"):
        flattened_ws.append([path, _to_excel_cell_value(value), value_type])
    flattened_ws.column_dimensions["A"].width = 64
    flattened_ws.column_dimensions["B"].width = 120
    flattened_ws.column_dimensions["C"].width = 18
    for row in flattened_ws.iter_rows(min_row=2, max_col=3):
        row[1].alignment = wrap_alignment

    node_ws = workbook.create_sheet(_excel_safe_sheet_name("NodeExecutions", used_sheet_names))
    node_ws.append(
        [
            "nodeId",
            "nodeType",
            "status",
            "startedAt",
            "completedAt",
            "durationMs",
            "error",
            "input",
            "output",
        ]
    )
    for header_cell in node_ws[1]:
        header_cell.font = bold_font

    for node_execution in node_executions:
        started_at = int(node_execution.started_at or 0)
        completed_at = int(node_execution.completed_at or 0)
        node_duration = completed_at - started_at if started_at and completed_at else ""
        parsed_input = _safe_json_loads(node_execution.input_json, None)
        parsed_output = _safe_json_loads(node_execution.output_json, None)
        node_ws.append(
            [
                node_execution.node_id,
                node_execution.node_type,
                node_execution.status,
                node_execution.started_at,
                node_execution.completed_at,
                node_duration,
                node_execution.error or "",
                _to_excel_cell_value(parsed_input if parsed_input is not None else (node_execution.input_json or "")),
                _to_excel_cell_value(parsed_output if parsed_output is not None else (node_execution.output_json or "")),
            ]
        )

    node_ws.column_dimensions["A"].width = 28
    node_ws.column_dimensions["B"].width = 16
    node_ws.column_dimensions["C"].width = 12
    node_ws.column_dimensions["D"].width = 18
    node_ws.column_dimensions["E"].width = 18
    node_ws.column_dimensions["F"].width = 12
    node_ws.column_dimensions["G"].width = 28
    node_ws.column_dimensions["H"].width = 80
    node_ws.column_dimensions["I"].width = 80
    for row in node_ws.iter_rows(min_row=2, min_col=7, max_col=9):
        for cell in row:
            cell.alignment = wrap_alignment

    table_candidates = _collect_table_candidates(flattened_source, root_path="result")
    if not table_candidates and flattened_source is not result_payload:
        table_candidates = _collect_table_candidates(result_payload, root_path="result")
    for index, table in enumerate(table_candidates, start=1):
        path = str(table.get("path") or f"table_{index}")
        tail_name = path.split(".")[-1].replace("[", "_").replace("]", "")
        sheet_name = _excel_safe_sheet_name(f"T{index}-{tail_name}", used_sheet_names)
        table_ws = workbook.create_sheet(sheet_name)
        columns = [str(column) for column in table.get("columns") or []]
        if not columns:
            continue

        table_ws.append(columns)
        for header_cell in table_ws[1]:
            header_cell.font = bold_font

        for row_payload in table.get("rows") or []:
            if not isinstance(row_payload, dict):
                continue
            table_ws.append([_to_excel_cell_value(row_payload.get(column)) for column in columns])

        if table.get("truncated"):
            table_ws.append(
                [f"[truncated] total rows={table.get('row_count')} exported={len(table.get('rows') or [])}"]
                + [""] * (max(0, len(columns) - 1))
            )
        for column_cells in table_ws.columns:
            first_column = column_cells[0].column_letter
            table_ws.column_dimensions[first_column].width = min(48, max(14, len(str(column_cells[0].value or "")) + 4))
        for row in table_ws.iter_rows(min_row=2, max_col=len(columns)):
            for cell in row:
                cell.alignment = wrap_alignment

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _extract_text_preview(payload: Any, max_length: int = 200) -> str:
    return _extract_text_preview_shared(payload, max_length=max_length, strip_markdown_fence=True)


def _normalize_runtime_hint(value: Any) -> Optional[str]:
    return _normalize_runtime_hint_shared(value)


def _extract_runtime_hints(
    payload: Any,
    _seen: Optional[set] = None,
    _depth: int = 0,
    _allow_scalar: bool = False,
) -> List[str]:
    return _extract_runtime_hints_shared(
        payload,
        _seen=_seen,
        _depth=_depth,
        _allow_scalar=_allow_scalar,
    )


def _pick_primary_runtime(runtime_hints: List[str]) -> str:
    return _pick_primary_runtime_shared(runtime_hints)


def _extract_trace_summary(payload: Any) -> Dict[str, Any]:
    return _extract_trace_summary_shared(payload)


def _extract_cost_summary(payload: Any) -> Dict[str, Any]:
    return _extract_cost_summary_shared(payload)


def _extract_workflow_output_candidates(payload: Any) -> List[Any]:
    if not isinstance(payload, dict):
        return [payload]

    looks_like_workflow_result = any(
        key in payload
        for key in ("finalOutput", "final_output", "outputs", "outputsMap", "finalNodeId", "final_node_id", "nodeStates")
    )

    final_output = payload.get("finalOutput")
    if final_output is None:
        final_output = payload.get("final_output")
    if final_output is not None:
        return [final_output]

    outputs_payload = payload.get("outputs")
    if outputs_payload is None:
        outputs_payload = payload.get("outputsMap")
    if isinstance(outputs_payload, dict):
        final_node_id = str(payload.get("finalNodeId") or payload.get("final_node_id") or "").strip()
        if final_node_id and not _is_non_result_workflow_node_id(final_node_id):
            final_node_output = outputs_payload.get(final_node_id)
            if final_node_output is not None:
                return [final_node_output]

        candidates: List[Any] = []
        for node_id, node_output in outputs_payload.items():
            if _is_non_result_workflow_node_id(node_id):
                continue
            candidates.append(node_output)
        if candidates:
            return candidates

    if looks_like_workflow_result:
        return []
    return [payload]


def _extract_workflow_video_result_metadata(result_payload: Any) -> Dict[str, Any]:
    summary: Dict[str, Any] = {
        "continuation_strategy": "",
        "video_extension_count": 0,
        "video_extension_applied": 0,
        "total_duration_seconds": 0,
        "continued_from_video": False,
        "subtitle_mode": "none",
        "subtitle_file_count": 0,
    }

    def _coerce_non_negative_int(value: Any) -> Optional[int]:
        if value is None or isinstance(value, bool):
            return None
        try:
            normalized = int(str(value).strip())
        except (TypeError, ValueError):
            return None
        return max(0, normalized)

    def _normalize_subtitle_mode(value: Any) -> str:
        candidate = str(value or "").strip().lower()
        return candidate or "none"

    def _count_subtitle_sidecars(items: Any) -> int:
        if not isinstance(items, list):
            return 0
        count = 0
        for item in items:
            if isinstance(item, dict):
                mime_type = str(item.get("mime_type") or item.get("mimeType") or "").strip().lower()
                file_format = str(item.get("format") or "").strip().lower()
                source_url = str(item.get("url") or item.get("data_url") or item.get("dataUrl") or "").strip().lower()
                if (
                    mime_type.startswith("text/")
                    or file_format in {"vtt", "srt"}
                    or source_url.endswith((".vtt", ".srt"))
                ):
                    count += 1
            elif isinstance(item, str):
                normalized = item.strip().lower()
                if normalized.endswith((".vtt", ".srt")) or normalized.startswith("data:text/"):
                    count += 1
        return count

    def _walk(payload: Any) -> None:
        if isinstance(payload, list):
            for item in payload:
                _walk(item)
            return

        if not isinstance(payload, dict):
            return

        continuation_strategy = str(
            payload.get("continuation_strategy") or payload.get("continuationStrategy") or ""
        ).strip()
        if continuation_strategy and not summary["continuation_strategy"]:
            summary["continuation_strategy"] = continuation_strategy

        video_extension_count = _coerce_non_negative_int(
            payload.get("video_extension_count") if "video_extension_count" in payload else payload.get("videoExtensionCount")
        )
        if video_extension_count is not None:
            summary["video_extension_count"] = max(summary["video_extension_count"], video_extension_count)

        video_extension_applied = _coerce_non_negative_int(
            payload.get("video_extension_applied")
            if "video_extension_applied" in payload
            else payload.get("videoExtensionApplied")
        )
        if video_extension_applied is not None:
            summary["video_extension_applied"] = max(summary["video_extension_applied"], video_extension_applied)

        total_duration_seconds = _coerce_non_negative_int(
            payload.get("total_duration_seconds")
            if "total_duration_seconds" in payload
            else payload.get("totalDurationSeconds")
        )
        if total_duration_seconds is not None:
            summary["total_duration_seconds"] = max(summary["total_duration_seconds"], total_duration_seconds)

        continued_from_video = payload.get("continued_from_video")
        if continued_from_video is None:
            continued_from_video = payload.get("continuedFromVideo")
        if bool(continued_from_video):
            summary["continued_from_video"] = True

        subtitle_mode = _normalize_subtitle_mode(
            payload.get("subtitle_mode") if "subtitle_mode" in payload else payload.get("subtitleMode")
        )
        if subtitle_mode != "none" or summary["subtitle_mode"] == "none":
            summary["subtitle_mode"] = subtitle_mode

        subtitle_file_count = _count_subtitle_sidecars(
            payload.get("sidecar_files") if "sidecar_files" in payload else payload.get("sidecarFiles")
        )
        if subtitle_file_count > 0:
            summary["subtitle_file_count"] = max(summary["subtitle_file_count"], subtitle_file_count)

        for value in payload.values():
            _walk(value)

    for candidate in _extract_workflow_output_candidates(result_payload):
        _walk(candidate)

    if summary["subtitle_file_count"] > 0 and summary["subtitle_mode"] == "none":
        summary["subtitle_mode"] = "vtt"

    return summary


def _build_workflow_result_summary(result_payload: Any) -> Dict[str, Any]:
    image_urls = _extract_workflow_output_media_urls(result_payload, "image")
    audio_urls = _extract_workflow_output_media_urls(result_payload, "audio")
    video_urls = _extract_workflow_output_media_urls(result_payload, "video")
    summary_image_urls = [
        image_url for image_url in image_urls
        if isinstance(image_url, str) and not image_url.strip().lower().startswith("data:image/")
    ]
    summary_audio_urls = [
        audio_url for audio_url in audio_urls
        if isinstance(audio_url, str) and not audio_url.strip().lower().startswith("data:audio/")
    ]
    summary_video_urls = [
        video_url for video_url in video_urls
        if isinstance(video_url, str) and not video_url.strip().lower().startswith("data:video/")
    ]
    text_preview = _extract_text_preview(result_payload)
    runtime_hints = _extract_runtime_hints(result_payload)
    trace_summary = _extract_trace_summary(result_payload)
    cost_summary = _extract_cost_summary(result_payload)
    video_metadata = _extract_workflow_video_result_metadata(result_payload)
    return {
        "has_result": result_payload is not None,
        "text_preview": text_preview,
        "image_count": len(image_urls),
        "image_urls": summary_image_urls,
        "audio_count": len(audio_urls),
        "audio_urls": summary_audio_urls,
        "video_count": len(video_urls),
        "video_urls": summary_video_urls,
        "runtime_hints": runtime_hints,
        "primary_runtime": _pick_primary_runtime(runtime_hints),
        "trace": trace_summary,
        "cost": cost_summary,
        **video_metadata,
    }


def _build_workflow_snapshot(db: Session, execution: WorkflowExecution) -> Dict[str, Any]:
    node_executions = db.query(NodeExecution).filter(
        NodeExecution.execution_id == execution.id
    ).order_by(
        NodeExecution.started_at.asc(),
        NodeExecution.id.asc(),
    ).all()

    node_statuses: Dict[str, str] = {}
    node_progress: Dict[str, int] = {}
    node_results: Dict[str, Any] = {}
    node_errors: Dict[str, str] = {}

    for ne in node_executions:
        status = _normalize_node_status(ne.status)
        node_statuses[ne.node_id] = status

        if status == "running":
            node_progress[ne.node_id] = 30
        elif status in ("completed", "failed", "skipped"):
            node_progress[ne.node_id] = 100
        else:
            node_progress[ne.node_id] = 0

        if ne.output_json:
            node_results[ne.node_id] = _serialize_workflow_result(_safe_json_loads(ne.output_json, {}))
        if ne.error:
            node_errors[ne.node_id] = ne.error

    return {
        "executionId": execution.id,
        "status": _normalize_workflow_status(execution.status),
        "startedAt": execution.started_at,
        "completedAt": execution.completed_at,
        "error": execution.error,
        "result": _serialize_workflow_result(_safe_json_loads(execution.result_json, {})),
        "nodeStatuses": node_statuses,
        "nodeProgress": node_progress,
        "nodeResults": node_results,
        "nodeErrors": node_errors,
        "runtimeMetrics": _build_runtime_metrics_snapshot(execution.id),
    }


def _extract_workflow_payload(execution: WorkflowExecution) -> Dict[str, Any]:
    payload = _safe_json_loads(execution.workflow_json, {})
    if not isinstance(payload, dict):
        return {"schema_version": 2, "nodes": [], "edges": [], "meta": {}}

    nodes = payload.get("nodes")
    edges = payload.get("edges")
    meta = payload.get("meta")
    schema_version = payload.get("schemaVersion") or payload.get("schema_version") or 2
    if not isinstance(nodes, list):
        nodes = []
    if not isinstance(edges, list):
        edges = []
    if not isinstance(meta, dict):
        meta = {}
    return {
        "schema_version": schema_version,
        "nodes": nodes,
        "edges": edges,
        "meta": meta,
    }


def _extract_input_payload(execution: WorkflowExecution) -> Dict[str, Any]:
    payload = _safe_json_loads(execution.input_json, {})
    if isinstance(payload, dict):
        return payload
    return {}


def _build_checkpoint_summary(checkpoint: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    return _build_checkpoint_summary_helper(checkpoint)


def _create_pause_checkpoint(
    request_payload: Dict[str, Any],
    node_events: List[Dict[str, Any]],
    *,
    paused_at: int,
    reason: str,
) -> Dict[str, Any]:
    return _create_pause_checkpoint_helper(
        request_payload=request_payload,
        node_events=node_events,
        paused_at=paused_at,
        reason=reason,
    )


def _build_resume_request_payload(
    execution: WorkflowExecution,
    checkpoint: Optional[Dict[str, Any]],
) -> Dict[str, Any]:
    return _build_resume_request_payload_helper(
        checkpoint,
        workflow_payload=_extract_workflow_payload(execution),
        input_payload=_extract_input_payload(execution),
    )


def _derive_execution_title(meta: Dict[str, Any], input_payload: Dict[str, Any]) -> str:
    for key in ("title", "name"):
        value = meta.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    for key in ("task", "prompt", "text", "input"):
        value = input_payload.get(key)
        if isinstance(value, str) and value.strip():
            trimmed = value.strip()
            return trimmed[:80] if len(trimmed) > 80 else trimmed
    return "未命名工作流"


def _build_node_summary(status_counter: Dict[str, int]) -> Dict[str, int]:
    aggregated = {
        "pending": 0,
        "running": 0,
        "completed": 0,
        "failed": 0,
        "skipped": 0,
    }
    for raw_status, raw_count in (status_counter or {}).items():
        normalized_key = _normalize_node_status(raw_status)
        aggregated[normalized_key] = int(aggregated.get(normalized_key, 0) or 0) + int(raw_count or 0)
    normalized = {
        "pending": int(aggregated.get("pending", 0)),
        "running": int(aggregated.get("running", 0)),
        "completed": int(aggregated.get("completed", 0)),
        "failed": int(aggregated.get("failed", 0)),
        "skipped": int(aggregated.get("skipped", 0)),
    }
    normalized["total"] = sum(normalized.values())
    return normalized


def _extract_template_tags(template: Dict[str, Any]) -> set[str]:
    tags: List[str] = []
    raw_tags = template.get("tags")
    if isinstance(raw_tags, list):
        tags.extend([str(item).strip().lower() for item in raw_tags if str(item).strip()])

    config = template.get("config")
    if isinstance(config, dict):
        meta = config.get("_templateMeta")
        if isinstance(meta, dict) and isinstance(meta.get("tags"), list):
            tags.extend([str(item).strip().lower() for item in meta.get("tags") if str(item).strip()])

    return set(tags)


def _extract_template_starter_key(template: Dict[str, Any]) -> str:
    direct_starter_key = str(
        template.get("starter_key")
        or template.get("starterKey")
        or ""
    ).strip().lower()
    if direct_starter_key:
        return direct_starter_key

    config = template.get("config")
    if not isinstance(config, dict):
        return ""
    meta = config.get("_templateMeta")
    if not isinstance(meta, dict):
        return ""
    starter_key = str(meta.get("starterKey") or meta.get("starter_key") or "").strip().lower()
    return starter_key


def _build_template_coverage_report(templates: List[Dict[str, Any]]) -> Dict[str, Any]:
    category_counter: Dict[str, int] = {}
    runtime_scope_counter: Dict[str, int] = {
        "provider-neutral": 0,
        "google-runtime": 0,
    }
    starter_keys = set()
    adk_template_count = 0

    normalized_templates: List[Dict[str, Any]] = []
    for template in templates:
        category = str(template.get("category") or "未分类").strip() or "未分类"
        category_counter[category] = int(category_counter.get(category, 0) or 0) + 1

        tags = _extract_template_tags(template)
        starter_key = _extract_template_starter_key(template)
        if starter_key:
            starter_keys.add(starter_key)
        is_google_runtime = "adk" in tags or "google-adk" in tags or starter_key.startswith("adk_sample_")
        runtime_scope_counter["google-runtime" if is_google_runtime else "provider-neutral"] += 1
        if is_google_runtime:
            adk_template_count += 1

        normalized_templates.append({
            "id": str(template.get("id") or ""),
            "name": str(template.get("name") or ""),
            "tags": tags,
            "starter_key": starter_key,
            "category": category,
        })

    def has_template(predicate) -> bool:
        return any(predicate(template) for template in normalized_templates)

    scenario_rules = [
        {
            "id": "text_to_image",
            "name": "文生图流程",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "image_generation_v1",
                        "product_image_creation_v1",
                        "product_image_prompt_matrix_v1",
                    }
                    or "text-to-image" in template["tags"]
                    or "image-gen" in template["tags"]
                )
            ),
        },
        {
            "id": "image_to_image",
            "name": "图生图流程",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "image_editing_v1",
                        "product_image_edit_parallel_v1",
                        "product_image_understand_edit_v1",
                        "product_image_batch_edit_v1",
                    }
                    or "image-to-image" in template["tags"]
                    or "image-edit" in template["tags"]
                )
            ),
        },
        {
            "id": "image_understanding",
            "name": "图片理解流程",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "image_understanding_v1",
                        "product_image_understand_edit_v1",
                    }
                    or "vision-understand" in template["tags"]
                    or "image-analysis" in template["tags"]
                )
            ),
        },
        {
            "id": "text_to_video",
            "name": "文生视频流程",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "video_generation_v1",
                    }
                    or "text-to-video" in template["tags"]
                    or "video-gen" in template["tags"]
                )
            ),
        },
        {
            "id": "text_to_speech",
            "name": "文生语音流程",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "audio_generation_v1",
                    }
                    or "text-to-speech" in template["tags"]
                    or "audio-gen" in template["tags"]
                    or "tts" in template["tags"]
                )
            ),
        },
        {
            "id": "parallel_image_pipeline",
            "name": "并行多分支图片生产",
            "matched": has_template(
                lambda template: (
                    "parallel" in template["tags"]
                    and ("image" in template["tags"] or "图像工作流" in template["category"])
                )
            ),
        },
        {
            "id": "multimodal_asset_pack",
            "name": "多模态素材包工作流",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "multimodal_asset_pack_v1",
                    }
                    or ("multimodal" in template["tags"] and "asset-pack" in template["tags"])
                )
            ),
        },
        {
            "id": "amazon_ads_optimization",
            "name": "Amazon 广告优化",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "amazon_ad_analysis_v1",
                        "amazon_ads_decision_board_v1",
                        "amazon_ads_bid_budget_guardrail_v1",
                    }
                    or ("amazon" in template["tags"] and "ads" in template["tags"])
                )
            ),
        },
        {
            "id": "amazon_listing_optimization",
            "name": "Amazon Listing 产出",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "amazon_listing_optimization_v1",
                        "amazon_multi_seller_listing_forge_v1",
                        "amazon_listing_refine_feedback_loop_v1",
                    }
                    or ("amazon" in template["tags"] and "listing" in template["tags"])
                )
            ),
        },
        {
            "id": "excel_data_analysis",
            "name": "Excel/报表数据分析",
            "matched": has_template(
                lambda template: (
                    template["starter_key"] in {
                        "table_analysis_v1",
                        "amazon_data_selection_v1",
                        "amazon_ad_analysis_v1",
                    }
                    or "data-analysis" in template["tags"]
                    or "excel-analysis" in template["tags"]
                )
            ),
        },
        {
            "id": "adk_official_samples",
            "name": "Google ADK 官方样例（Google runtime 专属）",
            "matched": adk_template_count > 0,
        },
    ]

    missing_scenarios = [rule["id"] for rule in scenario_rules if not rule["matched"]]
    return {
        "template_count": len(templates),
        "category_count": len(category_counter),
        "categories": category_counter,
        "runtime_scope_counts": runtime_scope_counter,
        "starter_keys": sorted(starter_keys),
        "adk_template_count": adk_template_count,
        "scenarios": [
            {
                "id": rule["id"],
                "name": rule["name"],
                "covered": bool(rule["matched"]),
            }
            for rule in scenario_rules
        ],
        "missing_scenarios": missing_scenarios,
        "is_complete": len(missing_scenarios) == 0,
    }


def _normalize_workflow_node_type(node: Dict[str, Any]) -> str:
    data = (node or {}).get("data") if isinstance(node, dict) else {}
    node_type = ""
    if isinstance(data, dict):
        node_type = str(data.get("type") or "").strip()
    if not node_type:
        node_type = str((node or {}).get("type") or "").strip()
    return node_type.lower().replace("-", "_")


async def _persist_node_event(
    db: Session,
    execution_id: str,
    nodes: List[Dict[str, Any]],
    event_type: str,
    data: Dict[str, Any],
    now: int,
    node_execution_cache: Optional[Dict[str, List[NodeExecution]]] = None,
):
    node_id = data.get("nodeId", "")
    if not node_id:
        return

    if event_type == "node_start":
        input_payload = data.get("input")
        entry = NodeExecution(
            id=generate_uuid(),
            execution_id=execution_id,
            node_id=node_id,
            node_type=_get_node_type(nodes, node_id),
            status="running",
            input_json=json.dumps(input_payload, ensure_ascii=False) if input_payload is not None else None,
            started_at=data.get("timestamp", now),
        )
        db.add(entry)
        if node_execution_cache is not None:
            node_execution_cache.setdefault(node_id, []).append(entry)
        db.commit()
        return

    if event_type in ("node_complete", "node_error", "node_skipped"):
        existing: Optional[NodeExecution] = None
        if node_execution_cache is not None:
            cached_rows = node_execution_cache.get(node_id) or []
            for row in reversed(cached_rows):
                if str(getattr(row, "status", "")).lower() == "running" and getattr(row, "completed_at", None) is None:
                    existing = row
                    break
            if existing is None and cached_rows:
                existing = cached_rows[-1]

        if not existing:
            existing = db.query(NodeExecution).filter(
                NodeExecution.execution_id == execution_id,
                NodeExecution.node_id == node_id,
                NodeExecution.status == "running",
            ).order_by(
                NodeExecution.started_at.desc(),
                NodeExecution.id.desc(),
            ).first()

        if not existing:
            existing = NodeExecution(
                id=generate_uuid(),
                execution_id=execution_id,
                node_id=node_id,
                node_type=_get_node_type(nodes, node_id),
                status="running",
                started_at=now,
            )
            db.add(existing)
            if node_execution_cache is not None:
                node_execution_cache.setdefault(node_id, []).append(existing)

        target_node_status = "completed"
        if event_type == "node_error":
            target_node_status = "failed"
        elif event_type == "node_skipped":
            target_node_status = "skipped"

        try:
            _transition_node_status(existing, target_node_status)
        except InvalidWorkflowStateTransitionError:
            logger.warning(
                "[Workflow] Ignored illegal node status transition execution=%s node=%s event=%s current=%s target=%s",
                execution_id,
                node_id,
                event_type,
                _normalize_node_status(existing.status),
                target_node_status,
            )
            return

        existing.completed_at = int(time.time() * 1000)
        if existing.input_json is None and data.get("input") is not None:
            existing.input_json = json.dumps(data.get("input"), ensure_ascii=False)
        if event_type == "node_complete":
            existing.output_json = json.dumps(data.get("output", {}), ensure_ascii=False)
            existing.error = None
        elif event_type == "node_error":
            existing.error = data.get("error", "")
        else:
            existing.output_json = json.dumps({"skipped": True, "reason": data.get("reason", "")}, ensure_ascii=False)
            existing.error = None
        db.commit()


async def _sync_template_sample_result(
    db: Session,
    user_id: str,
    execution_id: str,
    request_payload: Dict[str, Any],
    result_payload: Any,
    input_payload: Optional[Dict[str, Any]],
) -> None:
    meta = request_payload.get("meta")
    if not isinstance(meta, dict):
        return

    template_id = str(meta.get("template_id") or "").strip()
    if not template_id:
        return

    result_summary = _build_workflow_result_summary(result_payload)
    template_service = WorkflowTemplateService(db=db)
    updated_template = await template_service.update_template_sample_result(
        user_id=user_id,
        template_id=template_id,
        result_payload=result_payload,
        result_summary=result_summary,
        execution_id=execution_id,
        input_payload=input_payload if isinstance(input_payload, dict) else None,
    )
    if not updated_template:
        logger.warning(
            "[Workflow] Template sample result sync skipped: template not found or no access (template_id=%s, user=%s)",
            template_id,
            user_id,
        )


class WorkflowCancelledError(Exception):
    """Raised when workflow cancellation marker is observed."""


class WorkflowPausedError(Exception):
    """Raised when workflow pause marker is observed."""


@dataclass(frozen=True)
class WorkflowCancelTransitionResult:
    execution: Optional[WorkflowExecution]
    status: str
    transitioned: bool


def _mark_execution_cancelled_in_db(
    db: Session,
    execution_id: str,
    user_id: str,
    *,
    now: int,
    default_error: str,
) -> WorkflowCancelTransitionResult:
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        return WorkflowCancelTransitionResult(
            execution=None,
            status="pending",
            transitioned=False,
        )

    current_status = _normalize_workflow_status(execution.status)
    transitioned = False
    if current_status not in WORKFLOW_TERMINAL_STATUSES:
        _transition_workflow_status(execution, "cancelled")
        transitioned = True
        execution.completed_at = now
        if not execution.error:
            execution.error = default_error

        db.query(NodeExecution).filter(
            NodeExecution.execution_id == execution_id,
            NodeExecution.status == "running",
        ).update(
            {
                "status": "skipped",
                "completed_at": now,
                "error": "Node cancelled because workflow execution was cancelled.",
            },
            synchronize_session=False,
        )
        db.query(NodeExecution).filter(
            NodeExecution.execution_id == execution_id,
            NodeExecution.status == "pending",
        ).update(
            {
                "status": "skipped",
                "completed_at": now,
                "error": "Node skipped due to workflow cancellation.",
            },
            synchronize_session=False,
        )
    elif current_status == "cancelled":
        if execution.completed_at is None:
            execution.completed_at = now
        if not execution.error:
            execution.error = default_error

    return WorkflowCancelTransitionResult(
        execution=execution,
        status=_normalize_workflow_status(execution.status),
        transitioned=transitioned,
    )


def _mark_execution_paused_in_db(
    db: Session,
    execution_id: str,
    user_id: str,
    *,
    now: int,
) -> Optional[WorkflowExecution]:
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        return None

    current_status = _normalize_workflow_status(execution.status)
    if current_status in WORKFLOW_TERMINAL_STATUSES:
        return execution

    try:
        _transition_workflow_status(execution, "paused")
    except InvalidWorkflowStateTransitionError:
        execution.status = "paused"
    execution.completed_at = None
    execution.error = None
    execution.result_json = None
    db.query(NodeExecution).filter(
        NodeExecution.execution_id == execution_id,
        NodeExecution.status == "running",
    ).update(
        {
            "status": "skipped",
            "completed_at": now,
            "error": "Node paused because workflow execution was paused.",
        },
        synchronize_session=False,
    )
    return execution


async def _execute_workflow_internal(
    db: Session,
    execution_id: str,
    user_id: str,
    request_payload: Dict[str, Any],
    publish_events: bool = False,
) -> Dict[str, Any]:
    nodes = request_payload.get("nodes", [])
    edges = request_payload.get("edges", [])
    workflow_input = request_payload.get("input") or {}
    now = int(time.time() * 1000)

    llm_service = AgentLLMService(user_id=user_id, db=db)
    engine = WorkflowEngine(db=db, llm_service=llm_service)
    node_events: List[Dict[str, Any]] = []
    node_execution_cache: Dict[str, List[NodeExecution]] = {}
    cancel_error_message = "Execution cancelled by user."
    pause_error_message = "Execution paused by user."

    async def ensure_not_cancelled() -> None:
        if await _workflow_runtime_store.is_cancel_requested(execution_id):
            raise WorkflowCancelledError(cancel_error_message)

    async def ensure_not_paused() -> None:
        if await _workflow_runtime_store.is_pause_requested(execution_id):
            raise WorkflowPausedError(pause_error_message)

    async def on_event(event_type: str, data: Dict[str, Any]):
        await ensure_not_cancelled()
        await ensure_not_paused()
        serialized_data = _serialize_event_payload(event_type, data)
        node_event = {"type": event_type, **serialized_data}
        node_events.append(node_event)
        await _persist_node_event(
            db=db,
            execution_id=execution_id,
            nodes=nodes,
            event_type=event_type,
            data=serialized_data,
            now=now,
            node_execution_cache=node_execution_cache,
        )
        if publish_events:
            await _publish_runtime_event(execution_id, event_type, serialized_data)

    try:
        await ensure_not_cancelled()
        await ensure_not_paused()
        result = await engine.execute(
            nodes=nodes,
            edges=edges,
            initial_input=workflow_input,
            on_event=on_event,
        )
        await ensure_not_cancelled()
        await ensure_not_paused()
        persisted_result, image_replacements = await _persist_workflow_result_images(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            result_payload=result,
        )
        persisted_result, video_replacements = await _persist_workflow_result_media(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            result_payload=persisted_result,
            media_kind="video",
        )
        persisted_result, audio_replacements = await _persist_workflow_result_media(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            result_payload=persisted_result,
            media_kind="audio",
        )
        serialized_result = _serialize_workflow_result(persisted_result)
        if image_replacements:
            node_events = _replace_payload_image_urls(node_events, image_replacements)
        if video_replacements:
            node_events = _replace_payload_media_urls(node_events, video_replacements, "video")
        if audio_replacements:
            node_events = _replace_payload_media_urls(node_events, audio_replacements, "audio")
        await ensure_not_cancelled()
        await ensure_not_paused()
        result_preview_audio_urls = _build_result_preview_media_urls(
            execution_id,
            "audio",
            audio_replacements,
        )
        result_preview_video_urls = _build_result_preview_media_urls(
            execution_id,
            "video",
            video_replacements,
        )

        execution = db.query(WorkflowExecution).filter(
            WorkflowExecution.id == execution_id,
            WorkflowExecution.user_id == user_id,
        ).first()
        if execution:
            completed_at = int(time.time() * 1000)
            try:
                _transition_workflow_status(execution, "completed")
            except InvalidWorkflowStateTransitionError as transition_error:
                db.rollback()
                db.expire_all()
                latest = db.query(WorkflowExecution).filter(
                    WorkflowExecution.id == execution_id,
                    WorkflowExecution.user_id == user_id,
                ).first()
                latest_status = _normalize_workflow_status(getattr(latest, "status", None), default="failed")
                if latest_status == "cancelled":
                    raise WorkflowCancelledError(
                        str(getattr(latest, "error", "") or transition_error or cancel_error_message)
                    )
                if latest_status == "paused":
                    raise WorkflowPausedError(
                        str(getattr(latest, "error", "") or transition_error or pause_error_message)
                    )
                raise
            execution.result_json = json.dumps(serialized_result, ensure_ascii=False)
            execution.completed_at = completed_at
            execution.error = None
            db.commit()

        try:
            await _sync_template_sample_result(
                db=db,
                user_id=user_id,
                execution_id=execution_id,
                request_payload=request_payload,
                result_payload=serialized_result,
                input_payload=workflow_input if isinstance(workflow_input, dict) else None,
            )
        except Exception as sync_exc:
            logger.warning(
                "[Workflow] Failed to sync template sample result (execution=%s): %s",
                execution_id,
                sync_exc,
                exc_info=True,
            )

        if publish_events:
            await _publish_runtime_event(execution_id, "workflow_complete", _serialize_event_payload("workflow_complete", {
                "executionId": execution_id,
                "status": "completed",
                "result": serialized_result,
                "resultSummary": _build_workflow_result_summary(serialized_result),
                "resultPreviewAudioUrls": result_preview_audio_urls,
                "resultPreviewVideoUrls": result_preview_video_urls,
                "completedAt": int(time.time() * 1000),
            }))

        return {
            "execution_id": execution_id,
            "status": "completed",
            "result": serialized_result,
            "resultSummary": _build_workflow_result_summary(serialized_result),
            "resultPreviewAudioUrls": result_preview_audio_urls,
            "resultPreviewVideoUrls": result_preview_video_urls,
            "events": node_events,
        }
    except WorkflowPausedError as pause_exc:
        paused_at = int(time.time() * 1000)
        checkpoint = _create_pause_checkpoint(
            request_payload=request_payload,
            node_events=node_events,
            paused_at=paused_at,
            reason="pause_requested",
        )
        execution = _mark_execution_paused_in_db(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            now=paused_at,
        )
        if execution:
            db.commit()
        await _workflow_runtime_store.mark_paused(
            execution_id,
            paused=True,
            checkpoint=checkpoint,
            updated_at=paused_at,
        )
        checkpoint_summary = _build_checkpoint_summary(checkpoint)

        if publish_events:
            await _publish_runtime_event(execution_id, "workflow_paused", {
                "executionId": execution_id,
                "status": "paused",
                "pausedAt": paused_at,
                "error": str(pause_exc) or pause_error_message,
                "checkpoint": checkpoint_summary,
                "runtimeMetrics": _build_runtime_metrics_snapshot(execution_id),
            })

        return {
            "execution_id": execution_id,
            "status": "paused",
            "result": None,
            "events": node_events,
            "checkpoint": checkpoint_summary,
            "error": str(pause_exc) or pause_error_message,
        }
    except WorkflowCancelledError as cancel_exc:
        cancelled_at = int(time.time() * 1000)
        cancel_transition = _mark_execution_cancelled_in_db(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            now=cancelled_at,
            default_error=str(cancel_exc) or cancel_error_message,
        )
        execution = cancel_transition.execution
        if execution and cancel_transition.status == "cancelled":
            db.commit()

        await _workflow_runtime_store.mark_done(
            execution_id,
            done=cancel_transition.status in WORKFLOW_TERMINAL_STATUSES,
            updated_at=cancelled_at,
        )

        if publish_events and cancel_transition.transitioned:
            await _publish_runtime_event(execution_id, "workflow_cancelled", {
                "executionId": execution_id,
                "status": "cancelled",
                "error": execution.error if execution else str(cancel_exc),
                "completedAt": cancelled_at,
            })

        return {
            "execution_id": execution_id,
            "status": cancel_transition.status,
            "result": None,
            "events": node_events,
            "error": execution.error if execution else str(cancel_exc),
        }
    except asyncio.CancelledError:
        cancelled_at = int(time.time() * 1000)
        cancel_transition = _mark_execution_cancelled_in_db(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            now=cancelled_at,
            default_error=cancel_error_message,
        )
        execution = cancel_transition.execution
        if execution and cancel_transition.status == "cancelled":
            db.commit()
        await _workflow_runtime_store.mark_done(
            execution_id,
            done=cancel_transition.status in WORKFLOW_TERMINAL_STATUSES,
            updated_at=cancelled_at,
        )
        raise
    except Exception as e:
        logger.error(f"[Workflow] Execution failed: {e}", exc_info=True)
        failure_status = "failed"
        failure_error = str(e)
        completed_at = int(time.time() * 1000)
        execution = db.query(WorkflowExecution).filter(
            WorkflowExecution.id == execution_id,
            WorkflowExecution.user_id == user_id,
        ).first()
        if execution:
            try:
                _transition_workflow_status(execution, "failed")
                execution.error = failure_error
                execution.completed_at = completed_at
                db.commit()
            except InvalidWorkflowStateTransitionError:
                db.rollback()
                db.expire_all()
                latest = db.query(WorkflowExecution).filter(
                    WorkflowExecution.id == execution_id,
                    WorkflowExecution.user_id == user_id,
                ).first()
                if latest:
                    failure_status = _normalize_workflow_status(latest.status, default="failed")
                    failure_error = str(latest.error or failure_error)
                    completed_at = int(latest.completed_at or completed_at)
        await _workflow_runtime_store.mark_done(execution_id, done=True)

        if publish_events:
            event_name = _resolve_terminal_workflow_event(failure_status)
            await _publish_runtime_event(execution_id, event_name, {
                "executionId": execution_id,
                "status": failure_status,
                "error": failure_error,
                "completedAt": completed_at,
            })
        if failure_status == "cancelled":
            return {
                "execution_id": execution_id,
                "status": "cancelled",
                "result": None,
                "events": node_events,
                "error": failure_error,
            }
        raise


async def _run_workflow_in_background(
    execution_id: str,
    user_id: str,
    request_payload: Dict[str, Any]
):
    db = SessionLocal()
    try:
        await _execute_workflow_internal(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            request_payload=request_payload,
            publish_events=True,
        )
    except Exception:
        # 错误已在 _execute_workflow_internal 中记录和发布
        pass
    finally:
        done_flag = True
        try:
            latest = db.query(WorkflowExecution).filter(
                WorkflowExecution.id == execution_id,
                WorkflowExecution.user_id == user_id,
            ).first()
            latest_status = _normalize_workflow_status(getattr(latest, "status", None), default="failed")
            done_flag = latest_status in WORKFLOW_TERMINAL_STATUSES
        except Exception:
            done_flag = True
        await _workflow_runtime_store.mark_done(
            execution_id,
            done=done_flag,
            updated_at=int(time.time() * 1000),
        )
        await _cleanup_execution_runtime(
            execution_id,
            clear_store=done_flag,
            clear_local_runtime=done_flag,
            clear_task=True,
            clear_resume_lock=True,
        )
        db.close()


# ==================== Workflow Execution ====================

@router.post("/api/workflows/execute")
async def execute_workflow(
    request: WorkflowExecuteRequest,
    raw_request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """
    执行工作流

    - async_mode=False: 同步执行，返回完整结果（兼容旧调用）
    - async_mode=True: 异步执行，立即返回 execution_id，再通过 SSE 订阅状态
    """
    normalized_nodes = _normalize_workflow_nodes(request.nodes)
    normalized_input = _normalize_workflow_input_payload(request.input or {})
    normalized_meta = dict(request.meta or {})
    idempotency_key = _resolve_execute_idempotency_key(normalized_meta, raw_request)
    if idempotency_key:
        normalized_meta["idempotencyKey"] = idempotency_key
        normalized_meta["idempotency_key"] = idempotency_key
    validation_error = _validate_workflow_execute_payload(normalized_nodes, request.edges)
    if validation_error:
        raise HTTPException(status_code=400, detail=validation_error)

    now = int(time.time() * 1000)
    execution_id: Optional[str] = None

    def create_execution_record(new_execution_id: str) -> None:
        execution = WorkflowExecution(
            id=new_execution_id,
            user_id=user_id,
            idempotency_key=idempotency_key or None,
            workflow_json=json.dumps({
                "schemaVersion": 2,
                "nodes": normalized_nodes,
                "edges": request.edges,
                "meta": normalized_meta,
            }, ensure_ascii=False),
            input_json=json.dumps(normalized_input, ensure_ascii=False),
            status="running",
            started_at=now,
        )
        db.add(execution)
        db.commit()

    if idempotency_key:
        lock = _get_idempotency_lock(user_id, idempotency_key)
        async with lock:
            existing_execution = _find_execution_by_idempotency_key(
                db,
                user_id=user_id,
                idempotency_key=idempotency_key,
            )
            if existing_execution:
                await _workflow_runtime_store.touch_local(existing_execution.id, updated_at=now)
                _schedule_runtime_store_sync(
                    _workflow_runtime_store.touch(existing_execution.id, updated_at=now),
                    execution_id=existing_execution.id,
                    action="touch",
                )
                return _build_idempotent_execute_response(existing_execution)

            execution_id = generate_uuid()
            try:
                create_execution_record(execution_id)
            except IntegrityError:
                db.rollback()
                existing_execution = _find_execution_by_idempotency_key(
                    db,
                    user_id=user_id,
                    idempotency_key=idempotency_key,
                )
                if existing_execution:
                    await _workflow_runtime_store.touch_local(existing_execution.id, updated_at=now)
                    _schedule_runtime_store_sync(
                        _workflow_runtime_store.touch(existing_execution.id, updated_at=now),
                        execution_id=existing_execution.id,
                        action="touch",
                    )
                    return _build_idempotent_execute_response(existing_execution)
                raise
    else:
        execution_id = generate_uuid()
        create_execution_record(execution_id)

    if not execution_id:
        raise HTTPException(status_code=500, detail="Failed to create workflow execution record")

    _get_local_runtime(execution_id)
    await _workflow_runtime_store.initialize_execution_local(execution_id, updated_at=now)
    _schedule_runtime_store_sync(
        _workflow_runtime_store.initialize_execution(execution_id, updated_at=now),
        execution_id=execution_id,
        action="initialize",
    )

    request_payload = {
        "nodes": normalized_nodes,
        "edges": request.edges,
        "input": normalized_input,
        "meta": normalized_meta,
        "async_mode": bool(request.async_mode),
    }
    if request.async_mode:
        task = asyncio.create_task(
            _run_workflow_in_background(
                execution_id=execution_id,
                user_id=user_id,
                request_payload=request_payload,
            )
        )
        _execution_tasks[execution_id] = task
        return {
            "execution_id": execution_id,
            "status": "running",
        }

    try:
        sync_result = await _execute_workflow_internal(
            db=db,
            execution_id=execution_id,
            user_id=user_id,
            request_payload=request_payload,
            publish_events=False,
        )
    except HTTPException:
        await _workflow_runtime_store.mark_done(execution_id, done=True, updated_at=int(time.time() * 1000))
        await _cleanup_execution_runtime(
            execution_id,
            clear_store=True,
            clear_local_runtime=True,
            clear_task=False,
            clear_resume_lock=True,
        )
        raise
    except Exception as e:
        await _workflow_runtime_store.mark_done(execution_id, done=True, updated_at=int(time.time() * 1000))
        await _cleanup_execution_runtime(
            execution_id,
            clear_store=True,
            clear_local_runtime=True,
            clear_task=False,
            clear_resume_lock=True,
        )
        raise HTTPException(
            status_code=500,
            detail={
                "code": "workflow_execution_failed",
                "message": f"Workflow execution failed: {str(e) or 'unknown error'}",
            },
        ) from e
    sync_status = _normalize_workflow_status(sync_result.get("status"), default="failed")
    await _workflow_runtime_store.mark_done(
        execution_id,
        done=sync_status in WORKFLOW_TERMINAL_STATUSES,
        updated_at=int(time.time() * 1000),
    )
    await _cleanup_execution_runtime(
        execution_id,
        clear_store=sync_status in WORKFLOW_TERMINAL_STATUSES,
        clear_local_runtime=sync_status in WORKFLOW_TERMINAL_STATUSES,
        clear_task=False,
        clear_resume_lock=sync_status in WORKFLOW_TERMINAL_STATUSES,
    )
    return sync_result


@router.get("/api/workflows/{execution_id}/state")
async def get_workflow_execution_state(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """获取工作流执行统一状态快照（用于前端轮询）。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    snapshot = _build_workflow_snapshot(db, execution)

    runtime_updated_at = 0
    checkpoint_payload: Optional[Dict[str, Any]] = None
    try:
        runtime_state = await _workflow_runtime_store.get_state(execution_id)
        runtime_updated_at = int(runtime_state.updated_at or 0)
        if isinstance(runtime_state.checkpoint, dict):
            checkpoint_payload = runtime_state.checkpoint
    except Exception:
        logger.debug("[Workflow] load runtime state failed for execution=%s", execution_id, exc_info=True)

    response = _build_execution_state_payload(
        snapshot,
        runtime_updated_at=runtime_updated_at,
        checkpoint_payload=checkpoint_payload,
    )
    return response


@router.get("/api/workflows/{execution_id}/debug/execution-state-runtime")
async def get_workflow_execution_state_runtime_debug(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """返回 execution_state 推送运行时统计（只读调试接口）。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    runtime = _execution_runtime_local.get(execution_id)
    metrics = _build_execution_state_push_metrics_snapshot(
        None if runtime is None else runtime.execution_state_push_metrics
    )
    last_emitted_state_version = 0
    if runtime and isinstance(runtime.execution_state_last_emitted, dict):
        last_emitted_state_version = _safe_int(runtime.execution_state_last_emitted.get("stateVersion"), default=0)

    return {
        "execution_id": execution_id,
        "runtime_active": runtime is not None,
        "execution_state_runtime": {
            **metrics,
            "last_state_version": last_emitted_state_version,
            "has_cached_state": bool(runtime and isinstance(runtime.execution_state_cache, dict)),
        },
    }


@router.get("/api/workflows/{execution_id}/status")
async def workflow_status_stream(
    execution_id: str,
    request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """通过 SSE 订阅工作流执行统一状态（仅 execution_state）。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()

    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    runtime = _get_local_runtime(execution_id)

    async def _build_execution_state_with_runtime(
        current_execution: WorkflowExecution,
    ) -> Dict[str, Any]:
        snapshot = _build_workflow_snapshot(db, current_execution)
        runtime_updated_at = 0
        checkpoint_payload: Optional[Dict[str, Any]] = None
        try:
            runtime_state = await _workflow_runtime_store.get_state(execution_id)
            runtime_updated_at = _safe_int(runtime_state.updated_at, default=0)
            if isinstance(runtime_state.checkpoint, dict):
                checkpoint_payload = runtime_state.checkpoint
        except Exception:
            logger.debug("[Workflow] load runtime state failed for execution=%s", execution_id, exc_info=True)

        execution_state_payload = _build_execution_state_payload(
            snapshot,
            runtime_updated_at=runtime_updated_at,
            checkpoint_payload=checkpoint_payload,
        )
        _cache_execution_state_payload(execution_id, execution_state_payload)
        return execution_state_payload

    def _build_not_found_execution_state(*, emitted_at: int) -> Dict[str, Any]:
        payload = {
            "executionId": execution_id,
            "status": "failed",
            "finalStatus": "failed",
            "isTerminal": True,
            "stateVersion": max(0, _safe_int(emitted_at, default=0)),
            "startedAt": None,
            "completedAt": max(0, _safe_int(emitted_at, default=0)),
            "error": "Workflow execution not found",
            "result": None,
            "nodeStatuses": {},
            "nodeProgress": {},
            "nodeResults": {},
            "nodeErrors": {},
            "runtimeMetrics": _build_runtime_metrics_snapshot(execution_id),
            "clientPolicy": dict(WORKFLOW_EXECUTION_CLIENT_POLICY),
        }
        _cache_execution_state_payload(execution_id, payload)
        return payload

    async def event_generator():
        queue: asyncio.Queue = asyncio.Queue(maxsize=200)
        runtime.subscribers.append(queue)
        await _workflow_runtime_store.touch(execution_id, updated_at=int(time.time() * 1000))

        try:
            db.expire_all()
            current = db.query(WorkflowExecution).filter(
                WorkflowExecution.id == execution_id,
                WorkflowExecution.user_id == user_id,
            ).first()
            if not current:
                yield _format_sse(
                    "execution_state",
                    _build_not_found_execution_state(emitted_at=int(time.time() * 1000)),
                )
                return

            execution_state_payload = await _build_execution_state_with_runtime(current)
            yield _format_sse("execution_state", execution_state_payload)

            if bool(execution_state_payload.get("isTerminal")):
                return

            while True:
                if await request.is_disconnected():
                    break

                try:
                    message = await asyncio.wait_for(queue.get(), timeout=20.0)
                    if not isinstance(message, dict):
                        continue
                    if message.get("event") != "execution_state":
                        continue
                    state_data = message.get("data")
                    if not isinstance(state_data, dict):
                        continue
                    yield _format_sse("execution_state", state_data)
                    if bool(state_data.get("isTerminal")):
                        break
                except asyncio.TimeoutError:
                    db.expire_all()
                    latest = db.query(WorkflowExecution).filter(
                        WorkflowExecution.id == execution_id,
                        WorkflowExecution.user_id == user_id,
                    ).first()
                    if not latest:
                        yield _format_sse(
                            "execution_state",
                            _build_not_found_execution_state(emitted_at=int(time.time() * 1000)),
                        )
                        break

                    latest_status = _normalize_workflow_status(latest.status)
                    if latest_status in WORKFLOW_TERMINAL_STATUSES or latest_status == "paused":
                        latest_execution_state = await _build_execution_state_with_runtime(latest)
                        yield _format_sse("execution_state", latest_execution_state)
                        break

                    yield ": keepalive\n\n"
        finally:
            if queue in runtime.subscribers:
                runtime.subscribers.remove(queue)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@router.get("/api/workflows/history")
async def list_workflow_history(
    limit: int = 20,
    offset: int = 0,
    status: Optional[str] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """列出工作流执行历史（用于侧边栏加载与回放）"""
    safe_limit = max(1, min(limit, 100))
    safe_offset = max(0, offset)

    query = db.query(WorkflowExecution).filter(WorkflowExecution.user_id == user_id)
    if status:
        normalized_status = _normalize_workflow_status(status, default="")
        if normalized_status:
            query = query.filter(WorkflowExecution.status == normalized_status)

    total = query.count()
    executions = query.order_by(WorkflowExecution.started_at.desc()).offset(safe_offset).limit(safe_limit).all()

    execution_ids = [execution.id for execution in executions]
    node_counter_map: Dict[str, Dict[str, int]] = {}
    if execution_ids:
        counter_rows = db.query(
            NodeExecution.execution_id,
            NodeExecution.status,
            func.count(NodeExecution.id),
        ).filter(
            NodeExecution.execution_id.in_(execution_ids)
        ).group_by(
            NodeExecution.execution_id,
            NodeExecution.status,
        ).all()

        for execution_id, node_status, count in counter_rows:
            status_counter = node_counter_map.setdefault(execution_id, {})
            normalized_node_status = _normalize_node_status(node_status)
            status_counter[normalized_node_status] = int(status_counter.get(normalized_node_status, 0) or 0) + int(count)

    payload = []
    for execution in executions:
        workflow_payload = _extract_workflow_payload(execution)
        input_payload = _extract_input_payload(execution)
        parsed_result = _safe_json_loads(execution.result_json, None)
        result_payload = _serialize_workflow_result(parsed_result) if parsed_result is not None else None
        nodes = workflow_payload.get("nodes", [])
        edges = workflow_payload.get("edges", [])
        meta = workflow_payload.get("meta", {})
        title = _derive_execution_title(meta, input_payload)

        duration_ms = None
        if execution.started_at and execution.completed_at:
            duration_ms = max(0, execution.completed_at - execution.started_at)

        payload.append({
            "id": execution.id,
            "status": _normalize_workflow_status(execution.status),
            "title": title,
            "source": str(meta.get("source") or "").strip(),
            "task": input_payload.get("task") or input_payload.get("prompt") or "",
            "started_at": execution.started_at,
            "completed_at": execution.completed_at,
            "duration_ms": duration_ms,
            "error": execution.error,
            "node_summary": _build_node_summary(node_counter_map.get(execution.id, {})),
            "workflow_summary": {
                "node_count": len(nodes),
                "edge_count": len(edges),
            },
            "result_summary": _build_workflow_result_summary(result_payload),
        })

    return {
        "executions": payload,
        "count": len(payload),
        "total": total,
        "limit": safe_limit,
        "offset": safe_offset,
    }


@router.get("/api/workflows/history/{execution_id}")
async def get_workflow_history_detail(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """获取单次工作流执行详情（含 nodes/edges，可直接加载到编辑器）"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    workflow_payload = _extract_workflow_payload(execution)
    input_payload = _extract_input_payload(execution)
    parsed_result = _safe_json_loads(execution.result_json, None)
    result_payload = _serialize_workflow_result(parsed_result) if parsed_result is not None else None
    response_result_payload = _sanitize_history_detail_payload(result_payload)
    meta = workflow_payload.get("meta", {})
    title = _derive_execution_title(meta, input_payload)

    node_executions = db.query(NodeExecution).filter(
        NodeExecution.execution_id == execution_id
    ).order_by(
        NodeExecution.started_at.asc(),
        NodeExecution.id.asc(),
    ).all()

    node_status_counter: Dict[str, int] = {}
    node_payload: List[Dict[str, Any]] = []
    for node_execution in node_executions:
        status_key = _normalize_node_status(node_execution.status)
        node_status_counter[status_key] = node_status_counter.get(status_key, 0) + 1
        node_payload.append({
            "id": node_execution.id,
            "node_id": node_execution.node_id,
            "node_type": node_execution.node_type,
            "status": status_key,
            "input": _sanitize_history_detail_payload(
                _safe_json_loads(node_execution.input_json, {})
            ),
            "output": _sanitize_history_detail_payload(
                _serialize_workflow_result(_safe_json_loads(node_execution.output_json, {}))
            ),
            "error": node_execution.error,
            "started_at": node_execution.started_at,
            "completed_at": node_execution.completed_at,
        })

    duration_ms = None
    if execution.started_at and execution.completed_at:
        duration_ms = max(0, execution.completed_at - execution.started_at)

    return {
        "id": execution.id,
        "status": _normalize_workflow_status(execution.status),
        "title": title,
        "task": input_payload.get("task") or input_payload.get("prompt") or "",
        "started_at": execution.started_at,
        "completed_at": execution.completed_at,
        "duration_ms": duration_ms,
        "error": execution.error,
        "meta": meta,
        "workflow": {
            "schema_version": workflow_payload.get("schema_version") or 2,
            "nodes": workflow_payload.get("nodes", []),
            "edges": workflow_payload.get("edges", []),
        },
        "input": input_payload,
        "result": response_result_payload,
        "result_summary": _build_workflow_result_summary(result_payload),
        "node_summary": _build_node_summary(node_status_counter),
        "node_executions": node_payload,
    }


def _collect_history_request_forward_headers(request: Request) -> Dict[str, str]:
    inherited_headers: Dict[str, str] = {}
    for header_key in ("authorization", "cookie"):
        header_value = request.headers.get(header_key)
        if isinstance(header_value, str) and header_value.strip():
            inherited_headers[header_key.title()] = header_value.strip()
    return inherited_headers


def _load_workflow_history_result_payload(execution: WorkflowExecution) -> Any:
    parsed_result = _safe_json_loads(execution.result_json, None)
    return _serialize_workflow_result(parsed_result) if parsed_result is not None else None


@router.get("/api/workflows/history/{execution_id}/images/download")
async def download_workflow_history_images(
    execution_id: str,
    request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """批量下载单次工作流执行结果中的图片（ZIP）。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    image_urls = _extract_workflow_output_media_urls(result_payload, "image")
    if not image_urls:
        raise HTTPException(status_code=404, detail="No downloadable images found in workflow result")

    inherited_headers = _collect_history_request_forward_headers(request)
    trusted_base_url = _resolve_workflow_history_trusted_base_url()

    zip_bytes, manifest = await asyncio.to_thread(
        _build_workflow_images_zip,
        execution_id,
        image_urls,
        trusted_base_url,
        inherited_headers,
    )

    downloaded_count = int(manifest.get("downloadedCount") or 0)
    if downloaded_count <= 0:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "No images were downloadable. See manifest for skipped reasons.",
                "manifest": manifest,
            },
        )

    file_name = f"workflow-{execution_id[:8]}-images.zip"
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{file_name}"',
            "X-Workflow-Downloaded-Images": str(downloaded_count),
            "X-Workflow-Skipped-Images": str(int(manifest.get("skippedCount") or 0)),
        },
    )


@router.get("/api/workflows/history/{execution_id}/images/preview")
async def preview_workflow_history_images(
    execution_id: str,
    request: Request,
    limit: Optional[int] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """按需预览单次工作流结果图，返回 data URL（用于历史面板快速核验）。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    image_urls = _extract_workflow_output_media_urls(result_payload, "image")

    inherited_headers = _collect_history_request_forward_headers(request)
    trusted_base_url = _resolve_workflow_history_trusted_base_url()

    preview_payload = await asyncio.to_thread(
        _build_workflow_image_previews,
        execution_id,
        image_urls,
        trusted_base_url,
        inherited_headers,
        int(limit) if limit is not None else None,
        WORKFLOW_HISTORY_PREVIEW_MAX_BYTES_PER_IMAGE,
        WORKFLOW_HISTORY_PREVIEW_MAX_TOTAL_BYTES,
    )
    return preview_payload


@router.get("/api/workflows/history/{execution_id}/audio/download")
async def download_workflow_history_audio(
    execution_id: str,
    request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    audio_urls = _extract_workflow_output_media_urls(result_payload, "audio")
    if not audio_urls:
        raise HTTPException(status_code=404, detail="No downloadable audio found in workflow result")

    inherited_headers = _collect_history_request_forward_headers(request)
    trusted_base_url = _resolve_workflow_history_trusted_base_url()

    zip_bytes, manifest = await asyncio.to_thread(
        _build_workflow_media_zip,
        execution_id,
        "audio",
        audio_urls,
        trusted_base_url,
        inherited_headers,
        user_id,
    )

    downloaded_count = int(manifest.get("downloadedCount") or 0)
    if downloaded_count <= 0:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "No audio files were downloadable. See manifest for skipped reasons.",
                "manifest": manifest,
            },
        )

    file_name = f"workflow-{execution_id[:8]}-audio.zip"
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{file_name}"',
            "X-Workflow-Downloaded-Audio": str(downloaded_count),
            "X-Workflow-Skipped-Audio": str(int(manifest.get("skippedCount") or 0)),
        },
    )


@router.get("/api/workflows/history/{execution_id}/video/download")
async def download_workflow_history_video(
    execution_id: str,
    request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    video_urls = _extract_workflow_output_media_urls(result_payload, "video")
    if not video_urls:
        raise HTTPException(status_code=404, detail="No downloadable video found in workflow result")

    inherited_headers = _collect_history_request_forward_headers(request)
    trusted_base_url = _resolve_workflow_history_trusted_base_url()

    zip_bytes, manifest = await asyncio.to_thread(
        _build_workflow_media_zip,
        execution_id,
        "video",
        video_urls,
        trusted_base_url,
        inherited_headers,
        user_id,
    )

    downloaded_count = int(manifest.get("downloadedCount") or 0)
    if downloaded_count <= 0:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "No video files were downloadable. See manifest for skipped reasons.",
                "manifest": manifest,
            },
        )

    file_name = f"workflow-{execution_id[:8]}-video.zip"
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{file_name}"',
            "X-Workflow-Downloaded-Video": str(downloaded_count),
            "X-Workflow-Skipped-Video": str(int(manifest.get("skippedCount") or 0)),
        },
    )


@router.get("/api/workflows/history/{execution_id}/audio/preview")
async def preview_workflow_history_audio(
    execution_id: str,
    request: Request,
    limit: Optional[int] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    audio_urls = _extract_workflow_output_media_urls(result_payload, "audio")

    trusted_base_url = _resolve_workflow_history_trusted_base_url()
    preview_payload = await asyncio.to_thread(
        _build_workflow_media_previews,
        execution_id,
        "audio",
        audio_urls,
        trusted_base_url,
        f"/api/workflows/history/{execution_id}/audio/items/{{index}}",
        int(limit) if limit is not None else None,
    )
    return preview_payload


@router.get("/api/workflows/history/{execution_id}/video/preview")
async def preview_workflow_history_video(
    execution_id: str,
    request: Request,
    limit: Optional[int] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    video_urls = _extract_workflow_output_media_urls(result_payload, "video")

    trusted_base_url = _resolve_workflow_history_trusted_base_url()
    preview_payload = await asyncio.to_thread(
        _build_workflow_media_previews,
        execution_id,
        "video",
        video_urls,
        trusted_base_url,
        f"/api/workflows/history/{execution_id}/video/items/{{index}}",
        int(limit) if limit is not None else None,
    )
    return preview_payload


@router.get("/api/workflows/history/{execution_id}/audio/items/{item_index}")
async def stream_workflow_history_audio_item(
    execution_id: str,
    item_index: int,
    request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    audio_urls = _extract_workflow_output_media_urls(result_payload, "audio")
    if not audio_urls:
        raise HTTPException(status_code=404, detail="No previewable audio found in workflow result")

    try:
        audio_url = _resolve_workflow_media_item_shared(audio_urls, item_index)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    inherited_headers = _collect_history_request_forward_headers(request)
    trusted_base_url = _resolve_workflow_history_trusted_base_url()
    try:
        binary, mime_type, _ = await asyncio.to_thread(
            _download_workflow_media_binary_shared,
            audio_url,
            trusted_base_url,
            inherited_headers,
            "audio",
            WORKFLOW_HISTORY_AUDIO_PREVIEW_MAX_BYTES,
            user_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    return StreamingResponse(
        io.BytesIO(binary),
        media_type=mime_type,
        headers={"Accept-Ranges": "bytes"},
    )


@router.get("/api/workflows/history/{execution_id}/video/items/{item_index}")
async def stream_workflow_history_video_item(
    execution_id: str,
    item_index: int,
    request: Request,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    result_payload = _load_workflow_history_result_payload(execution)
    video_urls = _extract_workflow_output_media_urls(result_payload, "video")
    if not video_urls:
        raise HTTPException(status_code=404, detail="No previewable video found in workflow result")

    try:
        video_url = _resolve_workflow_media_item_shared(video_urls, item_index)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    inherited_headers = _collect_history_request_forward_headers(request)
    trusted_base_url = _resolve_workflow_history_trusted_base_url()
    try:
        binary, mime_type, _ = await asyncio.to_thread(
            _download_workflow_media_binary_shared,
            video_url,
            trusted_base_url,
            inherited_headers,
            "video",
            WORKFLOW_HISTORY_VIDEO_PREVIEW_MAX_BYTES,
            user_id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    return StreamingResponse(
        io.BytesIO(binary),
        media_type=mime_type,
        headers={"Accept-Ranges": "bytes"},
    )


@router.get("/api/workflows/history/{execution_id}/analysis/download")
async def download_workflow_history_analysis_excel(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """下载单次工作流执行分析结果为 Excel。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    parsed_result = _safe_json_loads(execution.result_json, None)
    result_payload = _serialize_workflow_result(parsed_result) if parsed_result is not None else {}
    node_executions = db.query(NodeExecution).filter(
        NodeExecution.execution_id == execution_id
    ).order_by(
        NodeExecution.started_at.asc(),
        NodeExecution.id.asc(),
    ).all()

    try:
        excel_bytes = await asyncio.to_thread(
            _build_workflow_analysis_excel_bytes,
            execution,
            result_payload,
            node_executions,
        )
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc))

    file_name = f"workflow-{execution_id[:8]}-analysis.xlsx"
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{file_name}"',
            "X-Workflow-Node-Count": str(len(node_executions)),
        },
    )


@router.delete("/api/workflows/history/{execution_id}")
async def delete_workflow_history(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """删除工作流执行历史（包含节点执行明细）"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    if execution.status == "running":
        raise HTTPException(status_code=409, detail="Cannot delete running execution")

    db.query(NodeExecution).filter(
        NodeExecution.execution_id == execution_id
    ).delete(synchronize_session=False)
    db.delete(execution)
    db.commit()

    await _cleanup_execution_runtime(
        execution_id,
        clear_store=True,
        clear_local_runtime=True,
        clear_task=True,
        clear_resume_lock=True,
    )
    return {"success": True}


@router.post("/api/workflows/history/{execution_id}/pause")
async def pause_workflow_history_execution(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """请求暂停运行中的工作流执行（在安全检查点生效）。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    current_status = _normalize_workflow_status(execution.status)
    if current_status in WORKFLOW_TERMINAL_STATUSES:
        return {
            "success": True,
            "execution_id": execution_id,
            "status": current_status,
            "already_terminal": True,
        }

    state = await _workflow_runtime_store.get_state(execution_id)
    if current_status == "paused" or bool(state.paused):
        return {
            "success": True,
            "execution_id": execution_id,
            "status": "paused",
            "already_paused": True,
            "checkpoint": _build_checkpoint_summary(state.checkpoint),
            "runtime_metrics": _build_runtime_metrics_snapshot(execution_id),
        }

    if current_status != "running":
        raise HTTPException(status_code=409, detail=f"Only running execution can be paused (current: {current_status})")

    now = int(time.time() * 1000)
    await _workflow_runtime_store.request_pause(execution_id, updated_at=now)
    try:
        await _publish_runtime_event(execution_id, "workflow_pause_requested", {
            "executionId": execution_id,
            "status": "running",
            "pauseRequestedAt": now,
        })
    except Exception:
        logger.debug("[Workflow] publish pause-requested event failed for execution=%s", execution_id, exc_info=True)

    return {
        "success": True,
        "execution_id": execution_id,
        "status": "pause_requested",
        "pause_requested": True,
    }


@router.post("/api/workflows/history/{execution_id}/resume")
async def resume_workflow_history_execution(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """恢复已暂停工作流执行（基于 checkpoint 重启）。"""
    lock = _get_resume_lock(execution_id)
    async with lock:
        db.expire_all()
        execution = db.query(WorkflowExecution).filter(
            WorkflowExecution.id == execution_id,
            WorkflowExecution.user_id == user_id,
        ).first()
        if not execution:
            raise HTTPException(status_code=404, detail="Workflow execution not found")

        current_status = _normalize_workflow_status(execution.status)
        if current_status in WORKFLOW_TERMINAL_STATUSES:
            return {
                "success": True,
                "execution_id": execution_id,
                "status": current_status,
                "already_terminal": True,
            }

        state = await _workflow_runtime_store.get_state(execution_id)
        existing_task = _execution_tasks.get(execution_id)
        has_active_task = bool(existing_task and not existing_task.done())
        is_stale_running = current_status == "running" and not has_active_task

        if current_status == "running" and has_active_task:
            return {
                "success": True,
                "execution_id": execution_id,
                "status": "running",
                "already_running": True,
            }
        if current_status != "paused" and not bool(state.paused) and not is_stale_running:
            raise HTTPException(status_code=409, detail=f"Only paused execution can be resumed (current: {current_status})")

        request_payload = _build_resume_request_payload(execution, state.checkpoint)
        validation_error = _validate_workflow_execute_payload(
            request_payload.get("nodes") if isinstance(request_payload, dict) else [],
            request_payload.get("edges") if isinstance(request_payload, dict) else [],
        )
        if validation_error:
            raise HTTPException(status_code=400, detail=f"Resume checkpoint is invalid: {validation_error}")

        now = int(time.time() * 1000)
        db.query(NodeExecution).filter(
            NodeExecution.execution_id == execution_id
        ).delete(synchronize_session=False)
        try:
            _transition_workflow_status(execution, "running")
        except InvalidWorkflowStateTransitionError:
            execution.status = "running"
        execution.completed_at = None
        execution.error = None
        execution.result_json = None
        db.commit()

        _get_local_runtime(execution_id)
        task = asyncio.create_task(
            _run_workflow_in_background(
                execution_id=execution_id,
                user_id=user_id,
                request_payload=request_payload,
            )
        )
        _execution_tasks[execution_id] = task

        try:
            await _workflow_runtime_store.mark_running(
                execution_id,
                clear_checkpoint=True,
                updated_at=now,
            )
        except Exception:
            logger.debug("[Workflow] runtime mark_running failed for execution=%s", execution_id, exc_info=True)

        try:
            await _publish_runtime_event(execution_id, "workflow_resumed", {
                "executionId": execution_id,
                "status": "running",
                "resumedAt": now,
                "checkpoint": _build_checkpoint_summary(state.checkpoint),
                "resumeStrategy": "restart",
            })
        except Exception:
            logger.debug("[Workflow] publish resumed event failed for execution=%s", execution_id, exc_info=True)

        return {
            "success": True,
            "execution_id": execution_id,
            "status": "running",
            "resume_strategy": "restart",
            "recovered_from_stale_running": is_stale_running,
        }


@router.post("/api/workflows/history/{execution_id}/cancel")
async def cancel_workflow_history_execution(
    execution_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """取消运行中的工作流执行，并保留历史记录。"""
    execution = db.query(WorkflowExecution).filter(
        WorkflowExecution.id == execution_id,
        WorkflowExecution.user_id == user_id,
    ).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Workflow execution not found")

    current_status = _normalize_workflow_status(execution.status)
    if current_status in WORKFLOW_TERMINAL_STATUSES:
        return {
            "success": True,
            "execution_id": execution_id,
            "status": current_status,
            "already_terminal": True,
        }

    now = int(time.time() * 1000)
    await _workflow_runtime_store.request_cancel(execution_id, updated_at=now)
    task_cancelled = False
    task = _execution_tasks.get(execution_id)
    if task and not task.done():
        task.cancel()
        task_cancelled = True

    cancel_transition = _mark_execution_cancelled_in_db(
        db=db,
        execution_id=execution_id,
        user_id=user_id,
        now=now,
        default_error="Execution cancelled by user.",
    )
    execution = cancel_transition.execution
    final_status = cancel_transition.status if execution else current_status

    if execution and final_status == "cancelled":
        db.commit()

    done_flag = final_status in WORKFLOW_TERMINAL_STATUSES
    await _workflow_runtime_store.mark_done(execution_id, done=done_flag, updated_at=now)
    await _cleanup_execution_runtime(
        execution_id,
        clear_store=done_flag,
        clear_local_runtime=done_flag,
        clear_task=done_flag or task_cancelled,
        clear_resume_lock=done_flag,
    )

    if cancel_transition.transitioned:
        try:
            await _publish_runtime_event(execution_id, "workflow_cancelled", {
                "executionId": execution_id,
                "status": "cancelled",
                "error": execution.error if execution else "Execution cancelled by user.",
                "completedAt": now,
            })
        except Exception:
            logger.debug("[Workflow] publish cancel event failed for execution=%s", execution_id, exc_info=True)

    return {
        "success": True,
        "execution_id": execution_id,
        "status": final_status,
        "cancel_transitioned": cancel_transition.transitioned,
        "task_cancelled": task_cancelled,
    }


async def _clear_all_workflow_history_for_user(db: Session, user_id: str) -> Dict[str, Any]:
    execution_rows = db.query(
        WorkflowExecution.id,
        WorkflowExecution.status,
    ).filter(
        WorkflowExecution.user_id == user_id
    ).all()

    execution_ids = [str(execution_id) for execution_id, _ in execution_rows if execution_id]
    running_execution_ids = [
        str(execution_id)
        for execution_id, status in execution_rows
        if execution_id and str(status or "").lower() == "running"
    ]

    cancelled_running_ids: List[str] = []
    for execution_id in running_execution_ids:
        task = _execution_tasks.get(execution_id)
        if task and not task.done():
            task.cancel()
            cancelled_running_ids.append(execution_id)

    node_deleted = 0
    execution_deleted = 0
    if execution_ids:
        node_deleted = db.query(NodeExecution).filter(
            NodeExecution.execution_id.in_(execution_ids)
        ).delete(synchronize_session=False)

    execution_deleted = db.query(WorkflowExecution).filter(
        WorkflowExecution.user_id == user_id
    ).delete(synchronize_session=False)

    db.commit()

    for execution_id in execution_ids:
        await _cleanup_execution_runtime(
            execution_id,
            clear_store=True,
            clear_local_runtime=True,
            clear_task=True,
            clear_resume_lock=True,
        )

    return {
        "execution_ids": execution_ids,
        "running_execution_ids": running_execution_ids,
        "cancelled_running_ids": cancelled_running_ids,
        "node_deleted_count": int(node_deleted or 0),
        "execution_deleted_count": int(execution_deleted or 0),
    }


async def _rebuild_workflow_templates_for_user(
    db: Session,
    user_id: str,
    recreate_starters: bool = True,
) -> Dict[str, Any]:
    deleted_count = db.query(WorkflowTemplate).filter(
        WorkflowTemplate.user_id == user_id
    ).delete(synchronize_session=False)
    db.commit()

    created_templates: List[Dict[str, Any]] = []
    if recreate_starters:
        service = WorkflowTemplateService(db=db)
        created_templates = await service.ensure_starter_templates(user_id=user_id)

    return {
        "deleted_count": int(deleted_count or 0),
        "created_count": len(created_templates),
        "templates": created_templates,
    }


@router.delete("/api/workflows/history")
async def clear_workflow_history(
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """清空当前用户的全部工作流执行历史（含节点执行明细）"""
    summary = await _clear_all_workflow_history_for_user(db=db, user_id=user_id)
    return {
        "success": True,
        "execution_deleted_count": summary["execution_deleted_count"],
        "node_deleted_count": summary["node_deleted_count"],
        "running_execution_count": len(summary["running_execution_ids"]),
        "cancelled_running_count": len(summary["cancelled_running_ids"]),
    }


@router.post("/api/workflows/templates/rebuild")
async def rebuild_workflow_templates(
    request: Optional[WorkflowTemplateRebuildRequest] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """删除并重建当前用户工作流模板（默认重建 Starter 模板）"""
    recreate_starters = True if request is None else bool(request.recreate_starters)
    summary = await _rebuild_workflow_templates_for_user(
        db=db,
        user_id=user_id,
        recreate_starters=recreate_starters,
    )
    return {
        "success": True,
        "deleted_count": summary["deleted_count"],
        "created_count": summary["created_count"],
        "templates": summary["templates"],
    }


@router.post("/api/workflows/reset")
async def reset_workflow_data(
    request: Optional[WorkflowResetRequest] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """一键重置工作流数据：清空历史 + 重建模板"""
    history_summary = await _clear_all_workflow_history_for_user(db=db, user_id=user_id)
    template_summary = await _rebuild_workflow_templates_for_user(
        db=db,
        user_id=user_id,
        recreate_starters=True if request is None else bool(request.recreate_starters),
    )

    return {
        "success": True,
        "history": {
            "execution_deleted_count": history_summary["execution_deleted_count"],
            "node_deleted_count": history_summary["node_deleted_count"],
            "running_execution_count": len(history_summary["running_execution_ids"]),
            "cancelled_running_count": len(history_summary["cancelled_running_ids"]),
        },
        "templates": {
            "deleted_count": template_summary["deleted_count"],
            "created_count": template_summary["created_count"],
            "items": template_summary["templates"],
        },
    }


# ==================== Mode Workflow Presets ====================

@router.get("/api/workflows/mode-presets")
async def list_mode_workflow_presets(
    user_id: str = Depends(require_current_user),
):
    """列出可直接载入编辑器的模式工作流预设摘要。"""
    _ = user_id
    items = [_build_mode_preset_summary(item) for item in MODE_WORKFLOW_PRESET_ITEMS]
    return {
        "items": items,
        "count": len(items),
    }


@router.get("/api/workflows/mode-presets/{mode_id}")
async def get_mode_workflow_preset(
    mode_id: str,
    user_id: str = Depends(require_current_user),
):
    """获取单个模式工作流预设（包含完整 nodes/edges）。"""
    _ = user_id
    normalized_mode_id = _normalize_mode_preset_id(mode_id)
    preset = MODE_WORKFLOW_PRESETS.get(normalized_mode_id)
    if not preset:
        raise HTTPException(status_code=404, detail=f"Mode workflow preset not found: {mode_id}")
    return copy.deepcopy(preset)


# ==================== Agent Helpers ====================

def _load_provider_models(user_id: str, db: Session) -> List[Dict[str, Any]]:
    profiles = db.query(ConfigProfile).filter(ConfigProfile.user_id == user_id).all()

    providers: List[Dict[str, Any]] = []
    media_only_tokens = (
        "veo", "sora", "luma", "video", "tts", "speech", "audio", "whisper", "embedding",
        "segmentation", "upscale", "try-on", "tryon", "aqa",
    )
    video_generation_tokens = ("veo", "sora", "luma", "video")
    audio_generation_tokens = ("tts", "speech")
    image_generation_tokens = (
        "imagen", "image", "dall", "wanx", "-t2i", "z-image",
        "flux", "midjourney", "nano-banana",
    )
    image_edit_tokens = (
        "capability", "ingredients", "edit", "inpaint",
        "outpaint", "mask", "background", "recontext",
    )

    def parse_saved_models(raw_saved_models: Any) -> List[Dict[str, Any]]:
        saved_models = raw_saved_models or []
        if isinstance(saved_models, str):
            try:
                saved_models = json.loads(saved_models)
            except Exception:
                saved_models = []

        normalized: List[Dict[str, Any]] = []
        if not isinstance(saved_models, list):
            return normalized

        for item in saved_models:
            if isinstance(item, dict):
                model_id = str(item.get("id") or item.get("model_id") or "").strip()
                model_name = str(item.get("name") or model_id).strip()
                capabilities = item.get("capabilities") if isinstance(item.get("capabilities"), dict) else {}
            else:
                model_id = str(item or "").strip()
                model_name = model_id
                capabilities = {}

            if not model_id:
                continue

            lower_id = model_id.lower()
            is_tongyi_vl = (
                "-vl-" in lower_id
                or lower_id.endswith("-vl")
                or "qwen-vl" in lower_id
                or "qwen2-vl" in lower_id
                or "qwen2.5-vl" in lower_id
            )
            is_google_chat_image_edit = (
                ("gemini" in lower_id and "image" in lower_id)
                or any(token in lower_id for token in ("flash-image", "pro-image", "nano-banana"))
            )
            is_tongyi_image_edit = (
                any(token in lower_id for token in ("wan2.6-image", "qwen-image-edit", "-i2i"))
                or (lower_id.startswith("wan") and "image" in lower_id and "-t2i" not in lower_id)
            )
            is_image_edit = any(token in lower_id for token in image_edit_tokens) or (
                "imagen" in lower_id and "generate" not in lower_id
            ) or is_google_chat_image_edit or is_tongyi_image_edit
            is_image_generate = (
                any(token in lower_id for token in image_generation_tokens)
                and not any(token in lower_id for token in media_only_tokens)
                and not is_image_edit
                and not is_tongyi_vl
            )
            is_video_generate = any(token in lower_id for token in video_generation_tokens)
            is_audio_generate = (
                any(token in lower_id for token in audio_generation_tokens)
                or lower_id.endswith("-tts")
                or "-tts-" in lower_id
            )
            supports_chat = True

            if any(token in lower_id for token in media_only_tokens):
                supports_chat = False
            if (
                any(token in lower_id for token in ("imagen", "dall", "wanx", "-t2i", "z-image", "midjourney", "flux"))
                and "gemini" not in lower_id
            ):
                supports_chat = False
            if is_image_edit and "gemini" not in lower_id:
                supports_chat = False
            if is_tongyi_vl:
                supports_chat = True
            is_vision_understand = bool(
                supports_chat and (
                    is_google_chat_image_edit
                    or is_tongyi_vl
                    or ("gemini" in lower_id and any(token in lower_id for token in ("flash", "pro", "image")))
                    or any(token in lower_id for token in ("gpt-4o", "claude-3", "qwen-vl", "qwen2-vl", "qwen2.5-vl"))
                )
            )

            supported_tasks: List[str] = []
            if supports_chat:
                supported_tasks.extend(["chat", "data-analysis"])
            if is_vision_understand:
                supported_tasks.append("vision-understand")
            if is_image_generate:
                supported_tasks.append("image-gen")
            if is_image_edit:
                supported_tasks.append("image-edit")
            if is_video_generate:
                supported_tasks.append("video-gen")
            if is_audio_generate:
                supported_tasks.append("audio-gen")

            if not supported_tasks and not any(token in lower_id for token in media_only_tokens):
                supported_tasks.append("chat")

            dedup_tasks: List[str] = []
            for task in supported_tasks:
                if task not in dedup_tasks:
                    dedup_tasks.append(task)

            normalized.append({
                "id": model_id,
                "name": model_name or model_id,
                "supported_tasks": dedup_tasks,
                "capabilities": capabilities,
            })
        return normalized

    def default_models_for_provider(provider_id: str) -> List[Dict[str, Any]]:
        lowered = str(provider_id or "").strip().lower()
        if lowered.startswith("google"):
            return [
                {"id": "gemini-2.5-flash", "name": "Gemini 2.5 Flash", "supported_tasks": ["chat", "data-analysis", "vision-understand"], "capabilities": {}},
                {"id": "gemini-2.5-flash-image", "name": "Gemini 2.5 Flash Image", "supported_tasks": ["chat", "data-analysis", "vision-understand", "image-edit"], "capabilities": {}},
                {"id": "imagen-3.0-generate-002", "name": "Imagen 3 Generate", "supported_tasks": ["image-gen"], "capabilities": {}},
            ]
        if lowered.startswith("tongyi") or lowered.startswith("dashscope"):
            return [
                {"id": "qwen-plus", "name": "Qwen Plus", "supported_tasks": ["chat", "data-analysis"], "capabilities": {}},
                {"id": "qwen2.5-vl-7b-instruct", "name": "Qwen2.5 VL 7B Instruct", "supported_tasks": ["chat", "data-analysis", "vision-understand"], "capabilities": {}},
                {"id": "wan2.6-t2i", "name": "Wan 2.6 T2I", "supported_tasks": ["image-gen"], "capabilities": {}},
                {"id": "wan2.6-image", "name": "Wan 2.6 Image", "supported_tasks": ["image-edit"], "capabilities": {}},
            ]
        if lowered.startswith("openai"):
            return [
                {"id": "gpt-4o-mini", "name": "GPT-4o mini", "supported_tasks": ["chat", "data-analysis"], "capabilities": {}},
                {"id": "gpt-4o", "name": "GPT-4o", "supported_tasks": ["chat", "data-analysis", "vision-understand"], "capabilities": {}},
                {"id": "dall-e-3", "name": "DALL-E 3", "supported_tasks": ["image-gen"], "capabilities": {}},
                {"id": "tts-1", "name": "TTS 1", "supported_tasks": ["audio-gen"], "capabilities": {}},
            ]
        return []

    provider_bucket: Dict[str, Dict[str, Any]] = {}

    for profile in profiles:
        if not profile.api_key:
            continue

        provider_id = str(profile.provider_id or "").strip()
        if not provider_id:
            continue
        provider_key = provider_id.lower()

        bucket = provider_bucket.get(provider_key)
        if not bucket:
            bucket = {
                "provider_id": provider_id,
                "provider_name": str(profile.name or provider_id).strip(),
                "latest_updated_at": int(getattr(profile, "updated_at", 0) or 0),
                "model_map": {},
            }
            provider_bucket[provider_key] = bucket
        else:
            current_updated_at = int(getattr(profile, "updated_at", 0) or 0)
            if current_updated_at >= int(bucket.get("latest_updated_at") or 0):
                bucket["provider_name"] = str(profile.name or provider_id).strip()
                bucket["latest_updated_at"] = current_updated_at

        model_map: Dict[str, Dict[str, Any]] = bucket["model_map"]
        for model in parse_saved_models(profile.saved_models):
            model_id = str(model.get("id") or "").strip()
            if not model_id:
                continue
            existing = model_map.get(model_id)
            if not existing:
                model_map[model_id] = model
                continue

            existing_tasks = list(existing.get("supported_tasks") or [])
            incoming_tasks = list(model.get("supported_tasks") or [])
            for task in incoming_tasks:
                if task not in existing_tasks:
                    existing_tasks.append(task)
            existing["supported_tasks"] = existing_tasks

            existing_caps = existing.get("capabilities")
            if not isinstance(existing_caps, dict):
                existing_caps = {}
            incoming_caps = model.get("capabilities")
            if isinstance(incoming_caps, dict):
                existing["capabilities"] = {**existing_caps, **incoming_caps}

    for bucket in provider_bucket.values():
        all_models = list(bucket.get("model_map", {}).values())
        if not all_models:
            all_models = default_models_for_provider(str(bucket.get("provider_id") or ""))
        all_models.sort(key=lambda item: str(item.get("name") or item.get("id") or ""))
        chat_models = [model for model in all_models if "chat" in (model.get("supported_tasks") or [])]
        image_generation_models = [
            model for model in all_models
            if "image-gen" in (model.get("supported_tasks") or [])
        ]
        image_edit_models = [
            model for model in all_models
            if "image-edit" in (model.get("supported_tasks") or [])
        ]
        video_generation_models = [
            model for model in all_models
            if "video-gen" in (model.get("supported_tasks") or [])
        ]
        audio_generation_models = [
            model for model in all_models
            if "audio-gen" in (model.get("supported_tasks") or [])
        ]
        default_models_by_task: Dict[str, str] = {}
        for task_name in AGENT_MODEL_SELECTION_POLICY["tasks"]:
            selected_id = ""
            for model in all_models:
                supported_tasks = model.get("supported_tasks")
                if isinstance(supported_tasks, list) and task_name in supported_tasks:
                    selected_id = str(model.get("id") or "").strip()
                    if selected_id:
                        break
            if selected_id:
                default_models_by_task[task_name] = selected_id

        providers.append({
            "provider_id": bucket.get("provider_id"),
            "provider_name": bucket.get("provider_name"),
            "all_models": all_models,
            "models": chat_models,
            "image_generation_models": image_generation_models,
            "image_edit_models": image_edit_models,
            "video_generation_models": video_generation_models,
            "audio_generation_models": audio_generation_models,
            "default_models_by_task": default_models_by_task,
        })

    providers.sort(key=lambda item: str(item.get("provider_name") or item.get("provider_id") or ""))

    return providers

# ==================== Agent CRUD ====================

@router.get("/api/agents")
async def list_agents(
    include_inactive: bool = False,
    search: Optional[str] = None,
    status: Optional[str] = None,
    task_type: Optional[str] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """列出用户的 Agent 列表，并返回后端计算的任务数量。"""

    ensure_seed_agents(db, user_id, seeds=get_default_seed_agents())

    normalized_status = _normalize_agent_list_status(status)
    normalized_task_type = _normalize_agent_task_filter(task_type)
    normalized_search = str(search or "").strip()

    query = db.query(AgentRegistry).filter(
        AgentRegistry.user_id == user_id,
    )
    if normalized_status == "active":
        query = query.filter(AgentRegistry.status == "active")
    elif normalized_status == "inactive":
        query = query.filter(AgentRegistry.status == "inactive")
    elif not include_inactive:
        query = query.filter(AgentRegistry.status == "active")

    if normalized_search:
        pattern = f"%{normalized_search}%"
        query = query.filter(
            (AgentRegistry.name.ilike(pattern))
            | (AgentRegistry.description.ilike(pattern))
            | (AgentRegistry.provider_id.ilike(pattern))
            | (AgentRegistry.model_id.ilike(pattern))
        )

    scoped_agents = query.order_by(
        AgentRegistry.updated_at.desc(),
        AgentRegistry.created_at.desc(),
    ).all()
    scoped_agent_payloads = [agent.to_dict() for agent in scoped_agents]
    task_counts = _build_agent_task_counts(scoped_agent_payloads)
    agents = _filter_agent_payloads_by_task(scoped_agent_payloads, normalized_task_type)

    active_count = db.query(AgentRegistry).filter(
        AgentRegistry.user_id == user_id,
        AgentRegistry.status == "active",
    ).count()
    inactive_count = db.query(AgentRegistry).filter(
        AgentRegistry.user_id == user_id,
        AgentRegistry.status == "inactive",
    ).count()

    return {
        "agents": agents,
        "count": len(agents),
        "active_count": int(active_count or 0),
        "inactive_count": int(inactive_count or 0),
        "task_counts": task_counts,
        "include_inactive": include_inactive,
        "search": normalized_search,
        "status": normalized_status,
        "task_type": normalized_task_type,
    }


@router.post("/api/agents")
async def create_agent(
    request: CreateAgentRequest,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """创建 Agent"""
    normalized_name = request.name.strip()
    if not normalized_name:
        raise HTTPException(status_code=400, detail="Agent name cannot be empty")
    normalized_agent_type = _normalize_agent_registry_type(request.agent_type)
    raw_agent_type = str(request.agent_type or "custom").strip().lower()
    if raw_agent_type not in VALID_AGENT_TYPES:
        raise HTTPException(status_code=400, detail=f"Unsupported agentType: {normalized_agent_type}")
    normalized_provider_id = str(request.provider_id or "").strip()
    normalized_model_id = str(request.model_id or "").strip()
    if not normalized_provider_id:
        raise HTTPException(status_code=400, detail="providerId is required")
    if not normalized_model_id:
        raise HTTPException(status_code=400, detail="modelId is required")
    if _agent_type_requires_google_provider(normalized_agent_type) and not normalized_provider_id.lower().startswith("google"):
        raise HTTPException(status_code=400, detail="adk/google-adk agentType currently requires a Google provider")

    duplicated = db.query(AgentRegistry).filter(
        AgentRegistry.user_id == user_id,
        AgentRegistry.status == "active",
        func.lower(AgentRegistry.name) == normalized_name.lower(),
    ).first()
    if duplicated:
        raise HTTPException(status_code=409, detail=f"Agent name already exists: {normalized_name}")

    now = int(time.time() * 1000)
    normalized_agent_card = _validate_and_normalize_agent_card(request.agent_card)
    agent = AgentRegistry(
        id=generate_uuid(),
        user_id=user_id,
        name=normalized_name,
        description=request.description or "",
        agent_type=normalized_agent_type,
        agent_card_json=json.dumps(normalized_agent_card, ensure_ascii=False) if normalized_agent_card else None,
        provider_id=normalized_provider_id,
        model_id=normalized_model_id,
        system_prompt=request.system_prompt or "",
        temperature=request.temperature or 0.7,
        max_tokens=request.max_tokens or 4096,
        icon=request.icon or "🤖",
        color=request.color or "#14b8a6",
        status="active",
        created_at=now,
        updated_at=now
    )
    db.add(agent)
    db.commit()
    db.refresh(agent)

    logger.info(f"[Agent] Created agent '{agent.name}' ({agent.id}) for user {user_id}")
    return agent.to_dict()


@router.put("/api/agents/{agent_id}")
async def update_agent(
    agent_id: str,
    request: UpdateAgentRequest,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """更新 Agent"""
    valid_statuses = {"active", "inactive"}
    agent = db.query(AgentRegistry).filter(
        AgentRegistry.id == agent_id,
        AgentRegistry.user_id == user_id
    ).first()

    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    update_data = request.dict(exclude_unset=True, exclude_none=True)
    if "name" in update_data:
        normalized_name = str(update_data["name"]).strip()
        if not normalized_name:
            raise HTTPException(status_code=400, detail="Agent name cannot be empty")
        duplicated = db.query(AgentRegistry).filter(
            AgentRegistry.user_id == user_id,
            AgentRegistry.status == "active",
            AgentRegistry.id != agent_id,
            func.lower(AgentRegistry.name) == normalized_name.lower(),
        ).first()
        if duplicated:
            raise HTTPException(status_code=409, detail=f"Agent name already exists: {normalized_name}")
        update_data["name"] = normalized_name

    if "provider_id" in update_data:
        provider_id = str(update_data.get("provider_id") or "").strip()
        if not provider_id:
            raise HTTPException(status_code=400, detail="providerId cannot be empty")
        update_data["provider_id"] = provider_id

    if "model_id" in update_data:
        model_id = str(update_data.get("model_id") or "").strip()
        if not model_id:
            raise HTTPException(status_code=400, detail="modelId cannot be empty")
        update_data["model_id"] = model_id

    if "agent_type" in update_data:
        raw_agent_type = str(update_data.get("agent_type") or "").strip().lower()
        normalized_agent_type = _normalize_agent_registry_type(raw_agent_type)
        if raw_agent_type not in VALID_AGENT_TYPES:
            raise HTTPException(status_code=400, detail=f"Unsupported agentType: {normalized_agent_type}")
        update_data["agent_type"] = normalized_agent_type

    effective_agent_type = _normalize_agent_registry_type(update_data.get("agent_type") or agent.agent_type or "custom")
    effective_provider_id = str(update_data.get("provider_id") or agent.provider_id or "").strip()
    if _agent_type_requires_google_provider(effective_agent_type) and not effective_provider_id.lower().startswith("google"):
        raise HTTPException(status_code=400, detail="adk/google-adk agentType currently requires a Google provider")

    if "status" in update_data:
        normalized_status = str(update_data.get("status") or "").strip().lower()
        if normalized_status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status: {normalized_status}")
        update_data["status"] = normalized_status

    if "agent_card" in update_data:
        raw_agent_card = update_data.pop("agent_card")
        if raw_agent_card is None:
            update_data["agent_card_json"] = None
        else:
            normalized_agent_card = _validate_and_normalize_agent_card(raw_agent_card)
            update_data["agent_card_json"] = json.dumps(normalized_agent_card, ensure_ascii=False)

    for key, value in update_data.items():
        setattr(agent, key, value)
    agent.updated_at = int(time.time() * 1000)

    db.commit()
    db.refresh(agent)

    return agent.to_dict()


@router.delete("/api/agents/{agent_id}")
async def delete_agent(
    agent_id: str,
    hard_delete: bool = False,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """删除 Agent（默认软删除，hard_delete=true 时物理删除）"""
    agent = db.query(AgentRegistry).filter(
        AgentRegistry.id == agent_id,
        AgentRegistry.user_id == user_id
    ).first()

    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    if hard_delete:
        db.delete(agent)
        db.commit()
        return {"success": True, "deleted_mode": "hard"}

    agent.status = "inactive"
    agent.updated_at = int(time.time() * 1000)
    db.commit()

    return {"success": True, "deleted_mode": "soft"}


@router.post("/api/agents/{agent_id}/restore")
async def restore_agent(
    agent_id: str,
    rename_on_conflict: bool = False,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """恢复已软删除的 Agent"""
    agent = db.query(AgentRegistry).filter(
        AgentRegistry.id == agent_id,
        AgentRegistry.user_id == user_id
    ).first()

    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")

    duplicated = db.query(AgentRegistry).filter(
        AgentRegistry.user_id == user_id,
        AgentRegistry.status == "active",
        AgentRegistry.id != agent_id,
        func.lower(AgentRegistry.name) == str(agent.name or "").lower(),
    ).first()
    if duplicated:
        if not rename_on_conflict:
            raise HTTPException(
                status_code=409,
                detail=f"Cannot restore agent because active agent with same name exists: {agent.name}",
            )

        base_name = str(agent.name or "").strip() or "Agent"
        candidate_name = f"{base_name} (restored)"
        suffix_index = 2
        while db.query(AgentRegistry).filter(
            AgentRegistry.user_id == user_id,
            AgentRegistry.status == "active",
            AgentRegistry.id != agent_id,
            func.lower(AgentRegistry.name) == candidate_name.lower(),
        ).first():
            candidate_name = f"{base_name} (restored {suffix_index})"
            suffix_index += 1
        agent.name = candidate_name

    agent.status = "active"
    agent.updated_at = int(time.time() * 1000)
    db.commit()
    db.refresh(agent)
    return {"success": True, "agent": agent.to_dict()}


# ==================== Agent Available Models ====================

@router.get("/api/workflows/execution-policy")
async def get_workflow_execution_policy(
    user_id: str = Depends(require_current_user),
):
    """返回前端执行状态同步策略（由后端统一下发）。"""
    _ = user_id
    return dict(WORKFLOW_EXECUTION_CLIENT_POLICY)


@router.get("/api/agents/available-models")
async def get_available_models_for_agents(
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """
    获取用户已配置的所有提供商及其模型列表（用于 Agent 编辑器）
    
    从 config_profiles 表读取用户已配置的提供商，
    返回每个提供商的 saved_models 列表。
    """
    providers = _load_provider_models(user_id=user_id, db=db)

    return {
        "providers": providers,
        "selection_policy": dict(AGENT_MODEL_SELECTION_POLICY),
    }


@router.get("/api/agents/{agent_id}")
async def get_agent(
    agent_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    """获取单个 Agent 详情"""
    agent = db.query(AgentRegistry).filter(
        AgentRegistry.id == agent_id,
        AgentRegistry.user_id == user_id
    ).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    return agent.to_dict()


# ==================== Workflow Templates ====================

@router.post("/api/workflows/templates")
async def create_workflow_template(
    request: WorkflowTemplateCreateRequest,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    service = WorkflowTemplateService(db=db)
    try:
        template = await service.create_template(
            user_id=user_id,
            name=request.name,
            description=request.description,
            category=request.category or "general",
            workflow_type=request.workflow_type or "graph",
            tags=request.tags,
            config=request.config,
            is_public=bool(request.is_public),
        )
        return template
    except ValueError as error:
        message = str(error)
        status_code = 409 if "already exists" in message else 400
        raise HTTPException(status_code=status_code, detail=message)


@router.get("/api/workflows/templates")
async def list_workflow_templates(
    category: Optional[str] = None,
    workflow_type: Optional[str] = None,
    search: Optional[str] = None,
    include_public: bool = True,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    service = WorkflowTemplateService(db=db)
    templates = await service.list_templates(
        user_id=user_id,
        category=category,
        workflow_type=workflow_type,
        search=search,
        include_public=include_public,
    )
    return {"templates": templates, "count": len(templates)}


@router.get("/api/workflows/template-categories")
async def list_workflow_template_categories(
    include_public: bool = True,
    ensure_defaults: bool = True,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    service = WorkflowTemplateService(db=db)
    categories = await service.list_categories(
        user_id=user_id,
        include_public=include_public,
        ensure_defaults=ensure_defaults,
    )
    return {"categories": categories, "count": len(categories)}


@router.post("/api/workflows/template-categories")
async def create_workflow_template_category(
    request: WorkflowTemplateCategoryCreateRequest,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    service = WorkflowTemplateService(db=db)
    try:
        category = await service.create_category(
            user_id=user_id,
            name=request.name,
        )
        return category
    except ValueError as error:
        message = str(error)
        status_code = 409 if "already exists" in message.lower() else 400
        raise HTTPException(status_code=status_code, detail=message)


@router.get("/api/workflows/templates/coverage")
async def get_workflow_template_coverage(
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    """
    返回当前用户模板覆盖报告（用于检查核心业务场景是否已覆盖）。
    """
    service = WorkflowTemplateService(db=db)
    templates = await service.list_templates(
        user_id=user_id,
        include_public=False,
    )
    coverage = _build_template_coverage_report(templates)
    return {
        "coverage": coverage,
        "templates": {
            "count": len(templates),
        },
    }


@router.post("/api/workflows/templates/seed")
async def seed_workflow_templates(
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    service = WorkflowTemplateService(db=db)
    created_templates = await service.ensure_starter_templates(user_id=user_id)
    return {
        "created_count": len(created_templates),
        "templates": created_templates,
    }


@router.get("/api/workflows/templates/{template_id}")
async def get_workflow_template(
    template_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    service = WorkflowTemplateService(db=db)
    template = await service.get_template(template_id=template_id, user_id=user_id)
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return template


@router.post("/api/workflows/templates/{template_id}/copy")
async def copy_workflow_template(
    template_id: str,
    request: Optional[WorkflowTemplateCopyRequest] = None,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db),
):
    service = WorkflowTemplateService(db=db)
    try:
        copied_template = await service.copy_template(
            user_id=user_id,
            template_id=template_id,
            name=request.name if request else None,
        )
        return copied_template
    except ValueError as error:
        message = str(error)
        lowered = message.lower()
        if "not found" in lowered:
            raise HTTPException(status_code=404, detail=message)
        if "already exists" in lowered:
            raise HTTPException(status_code=409, detail=message)
        raise HTTPException(status_code=400, detail=message)


@router.put("/api/workflows/templates/{template_id}")
async def update_workflow_template(
    template_id: str,
    request: WorkflowTemplateUpdateRequest,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    service = WorkflowTemplateService(db=db)
    try:
        template = await service.update_template(
            user_id=user_id,
            template_id=template_id,
            name=request.name,
            description=request.description,
            category=request.category,
            workflow_type=request.workflow_type,
            tags=request.tags,
            config=request.config,
            is_public=request.is_public,
        )
        return template
    except ValueError as error:
        message = str(error)
        lowered = message.lower()
        if "not found" in lowered:
            raise HTTPException(status_code=404, detail=message)
        if "already exists" in lowered:
            raise HTTPException(status_code=409, detail=message)
        raise HTTPException(status_code=400, detail=message)


@router.delete("/api/workflows/templates/{template_id}")
async def delete_workflow_template(
    template_id: str,
    user_id: str = Depends(require_current_user),
    db: Session = Depends(get_db)
):
    service = WorkflowTemplateService(db=db)
    try:
        success = await service.delete_template(user_id=user_id, template_id=template_id)
        return {"success": success}
    except ValueError as error:
        raise HTTPException(status_code=404, detail=str(error))


# ==================== Helpers ====================

def _get_node_type(nodes: List[Dict], node_id: str) -> str:
    for n in nodes:
        if n.get("id") == node_id:
            return n.get("data", {}).get("type") or n.get("type", "unknown")
    return "unknown"
